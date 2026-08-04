"""
工厂工作台 - FastAPI 后端服务
前后端分离架构 - 后端接口层
端口: 8933
"""

import os, sys, json, datetime, sqlite3, socket, io, re
from fastapi import FastAPI, File, UploadFile, Form, Query, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
import openpyxl

# ====== 配置 ======
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(os.path.dirname(BASE_DIR), '_scan_data.db')
UPLOAD_DIR = r'D:\工作文件\上传文件'
VERSION = 'v1.2'
REGIONS = ['东北部', '东南部', '西部', '中北部', '中南部']

os.makedirs(UPLOAD_DIR, exist_ok=True)

app = FastAPI(title="工厂工作台API", version="2.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ====== 工具函数 ======
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def get_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return '127.0.0.1'

def today_str():
    d = datetime.date.today()
    return d.isoformat()

def now_str():
    return datetime.datetime.now().isoformat()[:19]

# ====== 数据库初始化 ======
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS batches (
        id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, created_at TEXT,
        items_count INTEGER, regions TEXT, status TEXT DEFAULT 'active')''')
    c.execute('''CREATE TABLE IF NOT EXISTS documents (
        id INTEGER PRIMARY KEY AUTOINCREMENT, batch_id INTEGER, doc_number TEXT,
        region TEXT, total_boxes INTEGER, total_weight REAL, total_volume REAL,
        carrier TEXT, scanned INTEGER DEFAULT 0)''')
    c.execute('''CREATE TABLE IF NOT EXISTS scans (
        id INTEGER PRIMARY KEY AUTOINCREMENT, batch_id INTEGER, doc_number TEXT,
        worker TEXT, result TEXT, expected_qty INTEGER, scanned_at TEXT,
        region TEXT, note TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS job_batches (
        id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, created_at TEXT,
        items_count INTEGER)''')
    c.execute('''CREATE TABLE IF NOT EXISTS job_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT, batch_id INTEGER, sku TEXT,
        product_name TEXT, qty INTEGER, customer TEXT, notes TEXT,
        status TEXT DEFAULT 'pending', worker TEXT, started_at TEXT,
        completed_at TEXT, completed_qty INTEGER, priority INTEGER DEFAULT 0)''')
    try: c.execute('ALTER TABLE job_items ADD COLUMN paused_seconds INTEGER DEFAULT 0')
    except: pass
    try: c.execute('ALTER TABLE job_items ADD COLUMN job_number TEXT DEFAULT ""')
    except: pass
    try: c.execute('ALTER TABLE job_items ADD COLUMN done_qty INTEGER DEFAULT 0')
    except: pass
    try: c.execute('ALTER TABLE job_items ADD COLUMN est_hours REAL DEFAULT 0')
    except: pass
    try: c.execute('ALTER TABLE job_items ADD COLUMN est_min REAL DEFAULT 0')
    except: pass
    try: c.execute('ALTER TABLE job_items ADD COLUMN est_time TEXT DEFAULT ""')
    except: pass
    c.execute('''CREATE TABLE IF NOT EXISTS efficiency (
        id INTEGER PRIMARY KEY AUTOINCREMENT, sku TEXT UNIQUE, rate REAL,
        note TEXT, created_at TEXT)''')
    try: c.execute('ALTER TABLE efficiency ADD COLUMN type TEXT DEFAULT "manual"')
    except: pass
    try: c.execute('ALTER TABLE efficiency ADD COLUMN updated_at TEXT DEFAULT ""')
    except: pass
    conn.commit()
    conn.close()

init_db()

# ====== 发货Excel导入逻辑 ======
def import_shipment(fp, batch_name):
    wb = openpyxl.load_workbook(fp)
    ws1 = wb[wb.sheetnames[0]]
    doc_info = {}
    regions_used = set()
    for r in range(2, ws1.max_row + 1):
        doc = str(ws1.cell(r, 1).value or '').strip()
        if not doc: continue
        region = str(ws1.cell(r, 2).value or '').strip()
        carrier = str(ws1.cell(r, 5).value or '').strip() if ws1.max_column >= 5 else ''
        doc_info[doc] = {'region': region, 'carrier': carrier}
        if region: regions_used.add(region)

    ws2 = wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else ws1
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    now = now_str()
    c.execute('INSERT INTO batches (name, created_at, items_count, regions, status) VALUES (?,?,0,?,"active")',
              (batch_name, now, ','.join(sorted(regions_used))))
    bid = c.lastrowid
    count = 0
    for r in range(2, ws2.max_row + 1):
        doc = str(ws2.cell(r, 1).value or '').strip()
        boxes = ws2.cell(r, 2).value
        weight = ws2.cell(r, 3).value
        volume = ws2.cell(r, 4).value
        if not doc or not boxes: continue
        try: boxes = int(float(boxes))
        except: boxes = 0
        try: weight = float(weight)
        except: weight = 0
        try: volume = float(volume)
        except: volume = 0
        info = doc_info.get(doc, {})
        region = info.get('region', '')
        carrier = info.get('carrier', '')
        c.execute('''INSERT INTO documents (batch_id, doc_number, region, total_boxes, total_weight, total_volume, carrier)
                     VALUES (?,?,?,?,?,?,?)''', (bid, doc, region, boxes, weight, volume, carrier))
        count += 1
    c.execute('UPDATE batches SET items_count=? WHERE id=?', (count, bid))
    conn.commit(); conn.close()
    return bid, count, regions_used


# ====== 作业Excel导入逻辑 ======
def import_jobs(fp, batch_name):
    wb = openpyxl.load_workbook(fp)
    ws = wb.active
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    now = now_str()
    c.execute('INSERT INTO job_batches (name, created_at, items_count) VALUES (?,?,0)', (batch_name, now))
    batch_id = c.lastrowid
    count = 0
    for r in range(2, ws.max_row + 1):
        sku = str(ws.cell(r, 1).value or '').strip()
        if not sku: continue
        name = str(ws.cell(r, 2).value or '').strip()
        try: qty = int(float(ws.cell(r, 3).value or 0))
        except: qty = 0
        customer = str(ws.cell(r, 4).value or '').strip() if ws.max_column >= 4 else ''
        notes = str(ws.cell(r, 5).value or '').strip() if ws.max_column >= 5 else ''
        job_number = str(ws.cell(r, 6).value or '').strip() if ws.max_column >= 6 else ''
        est_hours = float(ws.cell(r, 7).value or 0) if ws.max_column >= 7 else 0
        c.execute('''INSERT INTO job_items (batch_id, sku, product_name, qty, customer, notes, job_number, est_hours)
                     VALUES (?,?,?,?,?,?,?,?)''',
                  (batch_id, sku, name, qty, customer, notes, job_number, est_hours))
        count += 1
    c.execute('UPDATE job_batches SET items_count=? WHERE id=?', (count, batch_id))
    conn.commit(); conn.close()
    return batch_id, count


# ====== API 端点 ======

@app.get("/health")
async def health():
    return {"status": "ok", "version": VERSION}

@app.get("/get_ip")
async def server_ip():
    return {"ip": get_ip()}

# ---- 车间加工相关 ----
@app.get("/workshop_data")
async def workshop_data(people: str = "1"):
    conn = get_db()
    c = conn.cursor()
    c.execute('''SELECT id, batch_id, sku, product_name as name, qty, customer, notes,
                 status, worker, started_at as started, completed_at as completed,
                 done_qty, priority, job_number, est_hours, est_min, est_time
                 FROM job_items ORDER BY priority DESC, id DESC''')
    rows = c.fetchall()
    items = []
    for r in rows:
        item = dict(r)
        if item.get('est_hours') and item.get('est_hours') > 0:
            ppl = int(people) if people.isdigit() else 1
            if ppl > 0:
                item['est_min'] = round(item['est_hours'] * 60 / ppl, 1)
                total_min = item['est_min']
                h = int(total_min // 60)
                m = int(total_min % 60)
                item['est_time'] = f"{h}小时{m}分钟" if h > 0 else f"{m}分钟"
        items.append(item)
    conn.close()
    return JSONResponse(content=items)

@app.get("/job_batches")
async def job_batches():
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM job_batches ORDER BY id DESC')
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return JSONResponse(content=rows)

@app.get("/calc_capacity")
async def calc_capacity(sku: str = "", qty: int = 0, people: int = 1):
    if not sku or qty <= 0:
        return JSONResponse({"status": "error", "message": "参数无效"})
    conn = get_db()
    c = conn.cursor()
    # 优先查手工录入的效率
    c.execute('SELECT rate, note FROM efficiency WHERE sku=?', (sku.upper(),))
    eff = c.fetchone()
    # 也从完成的job_items计算
    c.execute('''SELECT SUM(completed_qty) as tq, COUNT(*) as cnt
                 FROM job_items WHERE sku=? AND status='completed' AND completed_qty > 0''', (sku,))
    hist = c.fetchone()
    conn.close()

    rate = None
    source = "无历史记录"
    if eff and eff['rate']:
        rate = eff['rate']
        source = f"手动录入 ({eff['note']})" if eff['note'] else "手动录入"
    elif hist and hist['tq'] and hist['cnt']:
        rate = round(hist['tq'] / hist['cnt'] / people, 1) if people > 0 else 0
        source = f"加工历史 ({hist['cnt']}次, 共{hist['tq']}件)"

    if rate and rate > 0:
        hours = round(qty / (rate * people), 2)
        now = datetime.datetime.now()
        end_dt = now + datetime.timedelta(hours=hours)
        return JSONResponse({
            "status": "ok", "hours": hours, "rate": rate,
            "source": source, "end_time": end_dt.strftime("%m/%d %H:%M")
        })
    return JSONResponse({"status": "error", "message": "该SKU无历史效率数据，请先录入产能"})


@app.get("/get_efficiency")
async def get_efficiency():
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM efficiency ORDER BY created_at DESC')
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return JSONResponse(content=rows)


@app.get("/get_job_efficiency")
async def get_job_efficiency(sku: str = ""):
    conn = get_db()
    c = conn.cursor()
    c.execute('''SELECT sku, COUNT(*) as cnt, SUM(completed_qty) as total_qty
                 FROM job_items WHERE sku=? AND status='completed' AND completed_qty > 0
                 GROUP BY sku''', (sku.upper(),))
    r = c.fetchone()
    conn.close()
    if r:
        return JSONResponse({"status": "ok", "sku": r['sku'], "cnt": r['cnt'], "total_qty": r['total_qty']})
    return JSONResponse({"status": "error", "message": "无记录"})


@app.get("/query_efficiency")
async def query_efficiency(sku: str = ""):
    if not sku: return JSONResponse(content=[])
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM efficiency WHERE sku LIKE ? ORDER BY created_at DESC",
              (f"%{sku.upper()}%",))
    manual = [dict(r) for r in c.fetchall()]
    # 也查自动计算的（从completed jobs）
    c.execute('''SELECT sku, ROUND(SUM(completed_qty)*1.0/COUNT(*),1) as rate,
                 MAX(completed_at) as created_at, 'auto' as type, '加工完成自动计算' as note
                 FROM job_items WHERE sku LIKE ? AND status='completed' AND completed_qty > 0
                 GROUP BY sku ORDER BY created_at DESC''', (f"%{sku.upper()}%",))
    auto = [dict(r) for r in c.fetchall()]
    # 合并，手动优先
    manual_skus = {m['sku'].upper() for m in manual}
    auto_filtered = [a for a in auto if a['sku'].upper() not in manual_skus]
    result = manual + auto_filtered
    return JSONResponse(content=result)


# ---- 扫码发货相关 ----
@app.get("/scan_info")
async def scan_info(code: str = ""):
    if not code: return JSONResponse({"status": "error", "message": "请输入单号"})
    conn = get_db()
    c = conn.cursor()
    c.execute('''SELECT d.*, b.name as batch_name FROM documents d
                 JOIN batches b ON d.batch_id = b.id
                 WHERE d.doc_number=? AND b.status='active' ''', (code.strip(),))
    r = c.fetchone()
    if r:
        d = dict(r)
        # 查已扫数量
        c.execute('SELECT COUNT(*) as cnt FROM scans WHERE doc_number=? AND result="ok"', (code.strip(),))
        s = c.fetchone()
        d['scanned_count'] = s['cnt'] if s else 0
        conn.close()
        return JSONResponse({"status": "ok", "data": d})
    conn.close()
    return JSONResponse({"status": "error", "message": "未找到此发货单号"})


@app.get("/scan_check")
async def scan_check(code: str = "", worker: str = ""):
    if not code or not worker:
        return JSONResponse({"status": "error", "message": "参数不完整"})
    conn = get_db()
    c = conn.cursor()
    c.execute('''SELECT d.*, b.name as batch_name FROM documents d
                 JOIN batches b ON d.batch_id = b.id
                 WHERE d.doc_number=? AND b.status='active' ''', (code.strip(),))
    r = c.fetchone()
    if not r:
        conn.close()
        return JSONResponse({"status": "error", "message": "未找到此发货单号"})
    d = dict(r)
    now = now_str()
    c.execute('INSERT INTO scans (batch_id, doc_number, worker, result, expected_qty, scanned_at, region, note) VALUES (?,?,?,?,?,?,?,?)',
              (d['batch_id'], d['doc_number'], worker, 'ok', d['total_boxes'], now, d.get('region', ''), ''))
    conn.commit()
    # 检查是否全部扫完
    c.execute('SELECT COUNT(*) as cnt FROM scans WHERE doc_number=? AND result="ok"', (code.strip(),))
    sc = c.fetchone()
    conn.close()
    return JSONResponse({
        "status": "ok",
        "message": f"✅ {d['doc_number']} 扫码成功",
        "scanned": sc['cnt'] if sc else 0,
        "expected": d['total_boxes']
    })


@app.get("/scan_stats")
async def scan_stats(people: str = "1"):
    conn = get_db()
    c = conn.cursor()
    now = now_str()
    today = now[:10]
    c.execute('SELECT COUNT(*) as cnt FROM scans WHERE scanned_at LIKE ?', (f"{today}%",))
    today_count = c.fetchone()['cnt']
    c.execute('SELECT COUNT(DISTINCT doc_number) as cnt FROM scans WHERE result="ok"')
    total_docs = c.fetchone()['cnt']
    c.execute('SELECT COUNT(*) as cnt FROM scans')
    total_scans = c.fetchone()['cnt']
    c.execute('SELECT COUNT(*) as cnt FROM batches WHERE status="active"')
    active_batches = c.fetchone()['cnt']
    conn.close()
    return JSONResponse({
        "today": today_count, "total_docs": total_docs,
        "total_scans": total_scans, "active_batches": active_batches
    })


@app.get("/scan_batches")
async def scan_batches():
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM batches ORDER BY id DESC')
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return JSONResponse(content=rows)


@app.get("/scan_history")
async def scan_history(batch: str = "", worker: str = ""):
    conn = get_db()
    c = conn.cursor()
    q = "SELECT * FROM scans WHERE 1=1"
    params = []
    if batch:
        q += " AND batch_id=?"
        params.append(batch)
    if worker:
        q += " AND worker=?"
        params.append(worker)
    q += " ORDER BY scanned_at DESC LIMIT 500"
    c.execute(q, params)
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return JSONResponse(content=rows)


# ---- 物流跟踪 ----
@app.get("/track")
async def track(code: str = ""):
    if not code: return JSONResponse({"status": "error", "message": "请输入单号"})
    conn = get_db()
    c = conn.cursor()
    c.execute('''SELECT s.*, d.region, d.total_boxes, d.carrier, b.name as batch_name
                 FROM scans s LEFT JOIN documents d ON s.doc_number = d.doc_number
                 LEFT JOIN batches b ON s.batch_id = b.id
                 WHERE s.doc_number=? ORDER BY s.scanned_at DESC''', (code.strip(),))
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return JSONResponse(content=rows)


# ---- POST /run 统一处理 ----
@app.post("/run")
async def run_action(
    action: str = Form(""),
    batch_name: str = Form(""),
    batch_id: str = Form(""),
    worker: str = Form(""),
    done_qty: str = Form("0"),
    priority: str = Form("0"),
    sku: str = Form(""),
    qty: str = Form(""),
    people: str = Form("1"),
    hours: str = Form(""),
    note: str = Form(""),
    ids: str = Form(""),
    file: UploadFile = None
):
    conn = get_db(); c = conn.cursor()
    result = {"status": "error", "message": "未知操作"}

    # ---- 上传文件类 ----
    if action in ('shipment', 'import_jobs', 'complete'):
        if not file:
            conn.close()
            return JSONResponse({"status": "error", "message": "请选择文件"})
        fname = file.filename
        save_path = os.path.join(UPLOAD_DIR, os.path.basename(fname))
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        content = await file.read()
        with open(save_path, 'wb') as f: f.write(content)

        if action == 'shipment':
            bid, cnt, regions = import_shipment(save_path, batch_name or fname.replace('.xlsx',''))
            ip = get_ip()
            region_info = '、'.join(sorted(regions)) if regions else '无'
            result = {"status": "ok", "message": f"✅ 导入成功！批次#{bid}，{cnt}条记录\n区域：{region_info}\n\n工人扫码地址：http://{ip}:8933/scan"}
        elif action == 'import_jobs':
            bid, cnt = import_jobs(save_path, batch_name or fname.replace('.xlsx',''))
            result = {"status": "ok", "message": f"✅ 导入成功！{cnt}条加工任务"}
        elif action == 'complete':
            # 完成并上传结果文件
            if fname.lower().endswith('.xlsx'):
                # 读取完成文件更新job_items
                wb = openpyxl.load_workbook(save_path)
                ws = wb.active
                updated = 0
                for r in range(2, ws.max_row + 1):
                    sku_cell = str(ws.cell(r, 1).value or '').strip()
                    try: done = int(float(ws.cell(r, 2).value or 0))
                    except: done = 0
                    if sku_cell and done > 0:
                        c.execute('''UPDATE job_items SET status='completed', completed_at=?,
                                     completed_qty=completed_qty+?, done_qty=done_qty+?
                                     WHERE sku=? AND status='processing' ''',
                                  (now_str(), done, done, sku_cell))
                        updated += c.rowcount
                conn.commit()
                result = {"status": "ok", "message": f"✅ 完成 {updated} 条"}
            else:
                result = {"status": "error", "message": "请上传Excel文件"}
        conn.close()
        return JSONResponse(content=result)

    # ---- 纯数据操作 ----
    if action == 'start_job':
        item_id = int(batch_name) if batch_name.isdigit() else 0
        if item_id:
            c.execute("UPDATE job_items SET status='processing', worker=?, started_at=? WHERE id=? AND status='pending'",
                      (worker, now_str(), item_id))
            conn.commit()
            result = {"status": "ok", "message": "✅ 已开始加工"}

    elif action == 'complete_job':
        item_id = int(batch_name) if batch_name.isdigit() else 0
        try: dq = int(float(done_qty))
        except: dq = 0
        if item_id and dq > 0:
            now = now_str()
            c.execute('''UPDATE job_items SET status='completed', completed_at=?,
                         completed_qty=completed_qty+?, done_qty=done_qty+? WHERE id=?''',
                      (now, dq, dq, item_id))
            # 自动计算效率
            c.execute('SELECT sku, completed_qty, worker FROM job_items WHERE id=?', (item_id,))
            j = c.fetchone()
            if j:
                c.execute('SELECT started_at FROM job_items WHERE id=?', (item_id,))
                st = c.fetchone()
                if st and st[0]:
                    try:
                        start_dt = datetime.datetime.fromisoformat(st[0])
                        elapsed = (datetime.datetime.now() - start_dt).total_seconds() / 3600
                        if elapsed > 0.01:
                            workers = len(j['worker'].split(',')) if j['worker'] else 1
                            rate = round(dq / elapsed / max(1, workers), 1)
                            c.execute('''INSERT OR REPLACE INTO efficiency (sku, rate, note, created_at, type)
                                         VALUES (?,?,?,?,?)''',
                                      (j['sku'].upper(), rate, f'自动计算(完成{dq}件)', now, 'auto'))
                    except: pass
            conn.commit()
            result = {"status": "ok", "message": f"✅ 已完成 {dq} 件"}

    elif action == 'set_priority':
        item_id = int(batch_name) if batch_name.isdigit() else 0
        pri = int(priority) if priority.isdigit() else 0
        if item_id:
            c.execute('UPDATE job_items SET priority=? WHERE id=?', (pri, item_id))
            conn.commit()
            result = {"status": "ok"}

    elif action == 'cancel_job':
        item_id = int(batch_name) if batch_name.isdigit() else 0
        if item_id:
            c.execute("UPDATE job_items SET status='pending', worker='', started_at='' WHERE id=?", (item_id,))
            conn.commit()
            result = {"status": "ok", "message": "✅ 已取消"}

    elif action == 'delete_jobs':
        id_list = [int(x) for x in ids.split(',') if x.strip().isdigit()]
        if id_list:
            placeholders = ','.join(['?' for _ in id_list])
            c.execute(f'DELETE FROM job_items WHERE id IN ({placeholders})', id_list)
            conn.commit()
            result = {"status": "ok", "message": f"✅ 已删除 {len(id_list)} 项"}

    elif action == 'pause_job':
        item_id = int(batch_name) if batch_name.isdigit() else 0
        if item_id:
            c.execute("UPDATE job_items SET status='paused', notes=? WHERE id=? AND status='processing'",
                      (f"PAUSED:{now_str()}", item_id))
            conn.commit()
            result = {"status": "ok", "message": "⏸ 已暂停"}

    elif action == 'resume_job':
        item_id = int(batch_name) if batch_name.isdigit() else 0
        if item_id:
            c.execute("UPDATE job_items SET status='processing' WHERE id=? AND status='paused'", (item_id,))
            conn.commit()
            result = {"status": "ok", "message": "▶ 已恢复"}

    elif action == 'save_efficiency':
        if sku and qty:
            try:
                q = float(qty); p = float(people); h = float(hours)
                if h > 0 and p > 0:
                    rate = round(q / h / p, 1)
                    c.execute('''INSERT OR REPLACE INTO efficiency (sku, rate, note, created_at, type, updated_at)
                                 VALUES (?,?,?,?,?,?)''',
                              (sku.upper(), rate, note, now_str(), 'manual', now_str()))
                    conn.commit()
                    result = {"status": "ok", "message": f"✅ 已保存", "rate": rate}
                else:
                    result = {"status": "error", "message": "参数无效"}
            except:
                result = {"status": "error", "message": "数值格式错误"}

    elif action == 'delete_efficiency':
        if sku:
            c.execute('DELETE FROM efficiency WHERE sku=?', (sku.upper(),))
            conn.commit()
            result = {"status": "ok"}

    conn.close()
    return JSONResponse(content=result)



# ====== 静态文件服务 ======
from fastapi.staticfiles import StaticFiles
import os.path as _osp
_frontend = _osp.join(_osp.dirname(_osp.abspath(__file__)), '..', 'frontend')
if _osp.exists(_frontend):
    app.mount("/static", StaticFiles(directory=_frontend), name="static")

# ====== 页面路由 ======
@app.get("/")
async def index():
    fp = _osp.join(_frontend, 'index.html')
    if _osp.exists(fp):
        return FileResponse(fp)
    return HTMLResponse("<h1>前端文件未找到</h1>")

@app.get("/workshop_board")
async def workshop_board_page():
    fp = _osp.join(_frontend, 'workshop_board.html')
    if _osp.exists(fp): return FileResponse(fp)
    raise HTTPException(404)

@app.get("/workshop")
async def workshop_page():
    fp = _osp.join(_frontend, 'workshop.html')
    if _osp.exists(fp): return FileResponse(fp)
    raise HTTPException(404)

@app.get("/workshop_admin")
async def workshop_admin_page():
    fp = _osp.join(_frontend, 'workshop_admin.html')
    if _osp.exists(fp): return FileResponse(fp)
    raise HTTPException(404)

@app.get("/scan")
async def scan_page():
    fp = _osp.join(_frontend, 'scan.html')
    if _osp.exists(fp): return FileResponse(fp)
    raise HTTPException(404)

@app.get("/scan_admin")
async def scan_admin_page():
    fp = _osp.join(_frontend, 'scan_admin.html')
    if _osp.exists(fp): return FileResponse(fp)
    raise HTTPException(404)

@app.get("/scan_history_page")
async def scan_history_page():
    fp = _osp.join(_frontend, 'scan_history.html')
    if _osp.exists(fp): return FileResponse(fp)
    raise HTTPException(404)

# 其他静态页面通配
@app.get("/{page}.html")
async def html_page(page: str):
    fp = _osp.join(_frontend, f'{page}.html')
    if _osp.exists(fp):
        return FileResponse(fp)
    raise HTTPException(404)
FRONTEND_DIR = os.path.join(BASE_DIR, '..', 'frontend')

@app.get("/")
async def index():
    fp = os.path.join(FRONTEND_DIR, 'index.html')
    if os.path.exists(fp):
        return FileResponse(fp)
    return HTMLResponse("<h1>前端文件未找到</h1>")


if __name__ == '__main__':
    import uvicorn
    print(f"🚀 工厂工作台API启动: http://{get_ip()}:8933")
    uvicorn.run(app, host="0.0.0.0", port=8933)