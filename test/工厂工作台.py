import http.server, json, os, re, winreg, urllib.parse, io, webbrowser, sys, threading, time
import urllib.request, socket, sqlite3, datetime

# ====== 数据库 ======
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '_scan_data.db')
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS batches (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, created_at TEXT, items_count INTEGER, regions TEXT, status TEXT DEFAULT \'active\')''')
    c.execute('''CREATE TABLE IF NOT EXISTS documents (id INTEGER PRIMARY KEY AUTOINCREMENT, batch_id INTEGER, doc_number TEXT, region TEXT, total_boxes INTEGER, total_weight REAL, total_volume REAL, carrier TEXT, scanned INTEGER DEFAULT 0)''')
    c.execute('''CREATE TABLE IF NOT EXISTS scans (id INTEGER PRIMARY KEY AUTOINCREMENT, batch_id INTEGER, doc_number TEXT, worker TEXT, result TEXT, expected_qty INTEGER, scanned_at TEXT, region TEXT, note TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS job_batches (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, created_at TEXT, items_count INTEGER)''')
    c.execute('''CREATE TABLE IF NOT EXISTS job_items (id INTEGER PRIMARY KEY AUTOINCREMENT, batch_id INTEGER, sku TEXT, product_name TEXT, qty INTEGER, customer TEXT, notes TEXT, status TEXT DEFAULT \'pending\', worker TEXT, started_at TEXT, completed_at TEXT, completed_qty INTEGER, priority INTEGER DEFAULT 0)''')
    try: c.execute('ALTER TABLE job_items ADD COLUMN paused_seconds INTEGER DEFAULT 0')
    except: pass
    try: c.execute('ALTER TABLE job_items ADD COLUMN job_number TEXT DEFAULT \'\'')
    except: pass
    c.execute('''CREATE TABLE IF NOT EXISTS efficiency (id INTEGER PRIMARY KEY AUTOINCREMENT, sku TEXT UNIQUE, rate REAL, note TEXT, created_at TEXT)''')
    conn.commit(); conn.close()
init_db()

# 5个区域
REGIONS = ['东北部','东南部','西部','中北部','中南部']

def import_shipment(fp, batch_name):
    """导入发货Excel，按货件单号+区域存储，供工人扫码核对"""
    import openpyxl
    wb = openpyxl.load_workbook(fp)
    
    # 读取sheet1：发货单号 → 区域/物流商
    ws1 = wb[wb.sheetnames[0]]
    doc_info = {}  # 发货单号 → {'region': '', 'carrier': ''}
    regions_used = set()
    for r in range(2, ws1.max_row + 1):
        doc = str(ws1.cell(r, 1).value or '').strip()
        center = str(ws1.cell(r, 2).value or '').strip()  # 物流中心编码
        carrier = str(ws1.cell(r, 3).value or '').strip()  # 物流商
        if not doc: continue
        region = ''
        for rg in REGIONS:
            if center.startswith(rg):
                region = rg
                break
        if not region:
            for rg in REGIONS:
                if doc.startswith(rg) or doc.find(rg) != -1:
                    region = rg
                    break
        doc_info[doc] = {'region': region, 'carrier': carrier}
        if region: regions_used.add(region)
    
    # 读取sheet5：按货件单号汇总箱数/重量/体积
    try: ws5 = wb[wb.sheetnames[4]]
    except: ws5 = wb.active
    
    doc_totals = {}  # doc_number -> {boxes, weight, volume}
    for r in range(2, ws5.max_row + 1):
        doc = str(ws5.cell(r, 1).value or '').strip()
        if not doc: continue
        try: b = int(float(str(ws5.cell(r, 2).value or '0')))
        except: b = 0
        try: w = float(str(ws5.cell(r, 3).value or '0').replace(',','').strip())
        except: w = 0.0
        try: v = float(str(ws5.cell(r, 4).value or '0').replace(',','').strip())
        except: v = 0.0
        if doc not in doc_totals:
            doc_totals[doc] = {'boxes': 0, 'weight': 0.0, 'volume': 0.0}
        doc_totals[doc]['boxes'] += b
        doc_totals[doc]['weight'] += w
        doc_totals[doc]['volume'] += v
    
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    regions_str = ','.join(sorted(regions_used)) if regions_used else ''
    c.execute('INSERT INTO batches (name, created_at, items_count, regions) VALUES (?,?,0,?)',
              (batch_name, now, regions_str))
    bid = c.lastrowid; cnt = 0
    
    for doc, info in doc_info.items():
        if not doc: continue
        totals = doc_totals.get(doc, {'boxes': 0, 'weight': 0.0, 'volume': 0.0})
        region = info['region']
        if not region: continue  # 跳过无区域的数据
        c.execute('INSERT INTO documents (batch_id, doc_number, region, total_boxes, total_weight, total_volume, carrier, scanned) VALUES (?,?,?,?,?,?,?,0)',
                  (bid, doc, region, totals['boxes'], round(totals['weight'], 2), round(totals['volume'], 2), info['carrier']))
        cnt += 1
    
    c.execute('UPDATE batches SET items_count=? WHERE id=?', (cnt, bid))
    conn.commit(); conn.close()
    return bid, cnt, regions_used

def check_document(code, bid, region=''):
    """查询货件单号（前12字匹配），按区域过滤"""
    code12 = code[:12].upper()
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    if region:
        c.execute('SELECT doc_number, region, total_boxes, total_weight, total_volume, carrier FROM documents WHERE substr(doc_number,1,12)=? AND batch_id=? AND region=?',
                  (code12, bid, region))
        r = c.fetchone()
        if r:
            conn.close()
            return r  # (doc_number, region, boxes, weight, volume, carrier)
        # 没找到，看是否在其他区域
        c.execute('SELECT doc_number, region, total_boxes, total_weight, total_volume, carrier FROM documents WHERE substr(doc_number,1,12)=? AND batch_id=?',
                  (code12, bid))
        r2 = c.fetchone()
        conn.close()
        if r2:
            return ('wrong_region', r2[1], r2[2], r2[3], r2[4], r2[5])
        return None
    else:
        c.execute('SELECT doc_number, region, total_boxes, total_weight, total_volume, carrier FROM documents WHERE substr(doc_number,1,12)=? AND batch_id=?',
                  (code12, bid))
        r = c.fetchone(); conn.close(); return r

def record_scan(bid, code, worker, result, exp_qty, region='', note=''):
    code12 = code[:12].upper()
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    c.execute('INSERT INTO scans (batch_id,doc_number,worker,result,expected_qty,scanned_at,region,note) VALUES (?,?,?,?,?,?,?,?)',
              (bid, code12, worker, result, exp_qty, now, region, note))
    if result == 'correct':
        c.execute('UPDATE documents SET scanned=1 WHERE batch_id=? AND substr(doc_number,1,12)=?', (bid, code12))
    conn.commit(); conn.close()

def get_scan_stats(bid, region=''):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    if region:
        c.execute('SELECT doc_number, region, total_boxes, total_weight, total_volume, carrier, scanned FROM documents WHERE batch_id=? AND region=? ORDER BY doc_number', (bid, region))
    else:
        c.execute('SELECT doc_number, region, total_boxes, total_weight, total_volume, carrier, scanned FROM documents WHERE batch_id=? ORDER BY doc_number', (bid,))
    docs = c.fetchall()
    if region:
        c.execute('SELECT doc_number,worker,result,expected_qty,scanned_at,region,note FROM scans WHERE batch_id=? AND region=? ORDER BY id DESC', (bid, region))
    else:
        c.execute('SELECT doc_number,worker,result,expected_qty,scanned_at,region,note FROM scans WHERE batch_id=? ORDER BY id DESC', (bid,))
    scans = c.fetchall()
    c.execute('SELECT name,created_at,items_count,regions,status FROM batches WHERE id=?', (bid,))
    batch = c.fetchone()
    scanned_docs = [d for d in docs if d[6] == 1]
    scanned_nums = len(set(s[0] for s in scans if s[2] == 'correct'))
    correct = sum(1 for s in scans if s[2] == 'correct')
    wrong = sum(1 for s in scans if s[2] != 'correct')
        # Count corrections
    wrong_region_docs = set(s[0] for s in scans if s[2] == 'wrong_region')
    correct_docs = set(s[0] for s in scans if s[2] == 'correct')
    corrections = len(wrong_region_docs & correct_docs)
    total_expected_boxes = sum(d[2] for d in docs)
    scanned_correct_boxes = 0
    for s in scans:
        if s[2] == 'correct':
            for d in docs:
                if d[0] == s[0]:
                    scanned_correct_boxes += d[2]
                    break
    return {'batch':batch,'total':len(docs),'scanned':scanned_nums,
            'correct':correct,'wrong':wrong,'corrections':corrections,
            'total_expected_boxes':total_expected_boxes,'scanned_correct_boxes':scanned_correct_boxes,
            'docs':docs,'scans':scans,
            'unscanned':[d for d in docs if d[6] == 0]}


def import_jobs(fp, batch_name):
    """导入加工单Excel——创建工作任务列表"""
    import openpyxl
    wb = openpyxl.load_workbook(fp)
    ws = wb[wb.sheetnames[0]]
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('INSERT INTO job_batches (name, created_at, items_count) VALUES (?,?,0)', (batch_name, datetime.datetime.now().isoformat()[:19]))
    bid = c.lastrowid
    cnt = 0
    for r in range(2, ws.max_row + 1):
        job_no = str(ws.cell(r, 1).value or '').strip()
        sku = str(ws.cell(r, 12).value or '').strip()
        name = str(ws.cell(r, 10).value or '').strip()
        qty = ws.cell(r, 15).value
        sku = sku.replace('-','-').replace('_','-')
        if not sku: continue
        try: qty = int(float(qty))
        except: qty = 0
        notes = str(ws.cell(r, 24).value or '').strip()
        c.execute('INSERT INTO job_items (batch_id, sku, product_name, qty, job_number, notes) VALUES (?,?,?,?,?,?)', (bid, sku.upper(), name, qty, job_no, notes))
        cnt += 1
    c.execute('UPDATE job_batches SET items_count=? WHERE id=?', (cnt, bid))
    conn.commit(); conn.close()
    msg = '\u2705 \u52a0\u5de5\u5355\u5df2\u5bfc\u5165 (ID:'+str(bid)+')\n\u540d\u79f0: '+batch_name+'\n\u6761\u76ee: '+str(cnt)+'\u9879'
    return msg

def get_latest_batch():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT id,name,created_at,items_count,regions FROM batches ORDER BY id DESC LIMIT 1')
    r = c.fetchone(); conn.close(); return r

def get_batch_summary(bid):
    """Comprehensive batch summary"""
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT name,created_at,items_count,regions,status FROM batches WHERE id=?', (bid,))
    b = c.fetchone()
    if not b: conn.close(); return None
    regions = [r for r in b[3].split(',') if r] if b[3] else []
    region_stats = {}
    for rg in regions:
        region_stats[rg] = get_region_stats(bid, rg)
    c.execute('SELECT doc_number,worker,result,expected_qty,scanned_at,region,note FROM scans WHERE batch_id=? ORDER BY id', (bid,))
    scans = c.fetchall()
    wrong_region_set = set(s[0] for s in scans if s[2] == 'wrong_region')
    correct_set = set(s[0] for s in scans if s[2] == 'correct')
    corrections = len(wrong_region_set & correct_set)
    still_wrong = wrong_region_set - correct_set
    conn.close()
    return {
        'batch':b, 'region_stats':region_stats,
        'total_scans':len(scans), 'correct_scans':sum(1 for s in scans if s[2]=='correct'),
        'wrong_scans':sum(1 for s in scans if s[2]=='wrong_region'),
        'not_found_scans':sum(1 for s in scans if s[2]=='not_found'),
        'corrections':corrections, 'still_wrong':list(still_wrong)
    }

def get_all_batches():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT id,name,created_at,items_count,regions,status FROM batches ORDER BY id DESC')
    r = c.fetchall(); conn.close(); return r

def get_region_stats(bid, region):
    """获取某个区域内的扫码统计"""
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute('SELECT COUNT(*) FROM documents WHERE batch_id=? AND region=?', (bid, region))
    total = c.fetchone()[0]
    c.execute('SELECT COUNT(*) FROM documents WHERE batch_id=? AND region=? AND scanned=1', (bid, region))
    scanned = c.fetchone()[0]
    c.execute('SELECT COUNT(*) FROM scans WHERE batch_id=? AND region=? AND result="correct"', (bid, region))
    correct = c.fetchone()[0]
    c.execute('SELECT COUNT(*) FROM scans WHERE batch_id=? AND region=? AND result!="correct"', (bid, region))
    wrong = c.fetchone()[0]
    conn.close()
    return {'total':total,'scanned':scanned,'correct':correct,'wrong':wrong}

# ====== 端口 ======
# 使服务器能重用TIME_WAIT状态的端口
# allow_reuse_address removed - causes port stealing on Windows
VERSION = 'v20260720'
PORT = 8933
for _ in range(20):
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try: sock.bind(('0.0.0.0', PORT)); sock.close(); break
    except OSError: sock.close(); PORT += 1
else: print('无法找到可用端口'); sys.exit(1)

DESKTOP = r'D:\桌面'
if not os.path.isdir(DESKTOP):
    DESKTOP = os.path.expanduser('~/Desktop')
    if not os.path.isdir(DESKTOP):
        try:
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, r'Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders') as k:
                DESKTOP = os.path.expandvars(winreg.QueryValueEx(k, 'Desktop')[0])
        except: pass
UPLOAD_DIR = r'D:\工作文件\上传文件'
if not os.path.exists(UPLOAD_DIR): os.makedirs(UPLOAD_DIR)

def get_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80)); ip = s.getsockname()[0]; s.close(); return ip
    except: return 'localhost'

CARRIER_NAMES = {'shunfeng':'顺丰','yuantong':'圆通','shentong':'申通','zhongtong':'中通',
    'yunda':'韵达','jtexpress':'极兔','jd':'京东','ems':'EMS','sf':'顺丰'}

# ====== 扫码页面HTML（带区域选择） ======
SCAN_PAGE = '''<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=1.0,user-scalable=no"><title>扫码核对</title><style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:"Microsoft YaHei","PingFang SC",sans-serif;background:#f5f5f5;color:#333;padding:10px;max-width:500px;margin:0 auto}
h1{font-size:18px;text-align:center;padding:8px 0 4px}
.st{text-align:center;font-size:11px;color:#999;margin-bottom:10px}
.sel{width:100%;padding:10px;border:2px solid #1a73e8;border-radius:8px;font-size:15px;text-align:center;margin-bottom:8px;background:#fff;appearance:auto;cursor:pointer}
.sel:focus{outline:none;border-color:#1557b0}
.i{width:100%;padding:10px;border:2px solid #ddd;border-radius:8px;font-size:16px;text-align:center;margin-bottom:8px}
.i:focus{outline:none;border-color:#1a73e8}
.r{padding:20px;border-radius:10px;text-align:center;display:none;margin-top:10px}
.r .i2{font-size:48px;margin-bottom:6px}
.r .s{font-size:20px;font-weight:bold;margin-bottom:4px}
.r .d{font-size:13px;color:#666}
.good{background:#e6f4ea;border:2px solid #34a853}
.bad{background:#fce8e6;border:2px solid #ea4335}
.sc{font-size:12px;margin-top:10px;padding:8px;background:#fff;border-radius:6px;max-height:150px;overflow-y:auto;display:none}
.sc div{padding:3px 0;border-bottom:1px solid #eee}.sc div:last-child{border:none}
	.hd{display:none}
.cr{font-size:11px;color:#999;text-align:center;margin-top:15px}
</style></head><body>
<h1>📋 扫码核对</h1>
<p class="st" id="batchInfo">加载中...</p>
<select class="sel" id="regionSel"><option value="">-- 请选择区域 --</option></select>
<input class="i" id="worker" placeholder="输入你的名字">
<div style="display:flex;gap:6px;padding:0;margin-bottom:8px">
<input class="i" id="codeInput" placeholder="输入条码" style="flex:1;margin-bottom:0" onkeydown="if(event.key==='Enter')manualCheck()">
<button onclick="manualCheck()" style="padding:10px 18px;background:#1a73e8;color:#fff;border:none;border-radius:8px;font-size:15px;cursor:pointer;white-space:nowrap">查询</button>
</div>
<div class="r" id="result"></div>
<div class="sc" id="scanList"></div>
<p class="cr"><a href="/scan_admin">管理后台 →</a></p>
<script>
var w=localStorage.getItem('sw');if(w)document.getElementById('worker').value=w;
document.getElementById('worker').oninput=function(){localStorage.setItem('sw',this.value)};

async function manualCheck(){
    var code=document.getElementById('codeInput').value.trim();
    if(!code){alert('请输入条码');return}
    check(code);
}

async function check(c){
    var rs=document.getElementById('result'),sl=document.getElementById('scanList');
    var rg=document.getElementById('regionSel').value;
    rs.style.display='block';rs.className='r';rs.innerHTML='<div class="i2">\u23f3</div><div class="s">查询中...</div>';
    var r=await fetch('/scan_check?code='+encodeURIComponent(c)+'&worker='+encodeURIComponent(document.getElementById('worker').value.trim())+'&region='+encodeURIComponent(rg));
    var d=await r.json();
    if(d.found && d.match){
        rs.className='r good';
        rs.innerHTML='<div class="i2">\u2705</div><div class="s">\u6b63\u786e\uff01</div><div class="d">'+d.doc_number.substring(0,12)+'<br>\u5171'+d.expected_qty+'\u7bb1\uff0c\u5c5e\u4e8e'+d.region+'</div>';
    } else if(d.found && !d.match){
        rs.className='r bad';
        rs.innerHTML='<div class="i2">\u274c</div><div class="s">\u8d27\u7269\u653e\u9519\uff01</div><div class="d">'+d.code12+'<br>\u6b64\u5355\u5c5e\u4e8e <b>'+d.wrong_region+'</b>\uff0c\u5171'+d.expected_qty+'\u7bb1<br>\u8bf7\u66f4\u6b63\u8d27\u7269\uff0c\u653e\u56de<strong>'+d.wrong_region+'</strong>';
    } else {
        rs.className='r bad';
        rs.innerHTML='<div class="i2">\u2753</div><div class="s">\u672a\u627e\u5230\u6b64\u5355\u53f7</div><div class="d">'+c.substring(0,12)+'<br>\u4e0d\u5728\u5f53\u524d\u6279\u6b21\u6570\u636e\u4e2d</div>';
    }
    sl.style.display='block';
    sl.innerHTML='<b>\u23f1 \u626b\u63cf\u8bb0\u5f55</b>'+d.history.map(h=>'<div>'+h.time.substr(11,8)+' '+h.worker+' \u2192 '+h.code+' '+h.status+'</div>').reverse().join('');
    scanning=false;
    // 自动弹出继续扫码按钮
    document.getElementById('nextBtn').style.display='block';
}

function restartCamera(){
    document.getElementById('result').style.display='none';
    document.getElementById('nextBtn').style.display='none';
    document.getElementById('reader').style.display='block';
    flashOn=false;
    startScan();
}

// 加载区域选项和信息
fetch('/scan_info').then(r=>r.json()).then(d=>{
    document.getElementById('batchInfo').textContent=d.batch_name+' | '+d.total+'\u9879';
    var sel=document.getElementById('regionSel');
    var saved=localStorage.getItem('sr')||'';
    d.regions.forEach(function(rg){
        var o=document.createElement('option');
        o.value=rg;
        var rs=d.region_stats[rg]||{total:0,scanned:0};
        o.textContent=rg+' ('+rs.total+'\u9879, \u5df2\u626b'+rs.scanned+')';
        sel.appendChild(o);
    });
    if(saved && d.regions.includes(saved)){
        sel.value=saved;
        setTimeout(autoStart,300);
    }
});
</script></body></html>'''

SCAN_ADMIN = '''<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"><title>扫码管理</title><style>
body{font-family:"Microsoft YaHei","PingFang SC",sans-serif;background:#f0f2f5;color:#333;padding:20px;max-width:1200px;margin:0 auto}
h1{font-size:20px;margin-bottom:12px}
.st{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:16px}
.st>div{background:#fff;border-radius:8px;padding:12px;text-align:center}
.st .n{font-size:24px;font-weight:bold}.st .l{font-size:11px;color:#999;margin-top:2px}
.rt{display:grid;grid-template-columns:repeat(5,1fr);gap:6px;margin-bottom:16px}
.rt>div{background:#fff;border-radius:6px;padding:8px;text-align:center;font-size:11px;cursor:pointer;border:2px solid transparent}
.rt>div:hover{border-color:#1a73e8}.rt>div.sel{border-color:#1a73e8;background:#e8f0fe}
.rt .rn{font-size:14px;font-weight:bold;color:#1a73e8}
.sel2{width:100%;padding:6px;border:1px solid #ddd;border-radius:4px;font-size:12px;margin-bottom:12px}
.t{width:100%;border-collapse:collapse;margin-bottom:12px;font-size:11px}
.t th,.t td{padding:5px 6px;border:1px solid #ddd;text-align:left}
.t th{background:#1a73e8;color:#fff;white-space:nowrap;position:sticky;top:0}
.g{color:#34a853;font-weight:bold}.r{color:#ea4335;font-weight:bold}
.h2{font-size:14px;margin:10px 0 6px;display:flex;align-items:center;gap:6px}
.h2 .tag{padding:2px 8px;border-radius:3px;color:#fff;font-size:10px}
.ex{background:#fff;border-radius:6px;padding:8px;font-size:11px;margin-bottom:12px;max-height:200px;overflow-y:auto}
</style></head><body>
<h1>📊 扫码管理后台</h1>
<select class="sel2" id="batchSel" onchange="load()"></select>
<div class="rt" id="regionTabs"></div>
<div class="st" id="stats"></div>
<div style="display:flex;gap:20px;flex-wrap:wrap">
<div style="flex:1;min-width:300px"><h2 class="h2">未扫码项</h2><div class="ex" id="unscanned"></div></div>
<div style="flex:2;min-width:400px"><h2 class="h2">扫描记录</h2><div id="history" style="overflow-x:auto"></div></div>
</div>
<p style="font-size:12px;margin-top:10px"><a href="/scan">📱 扫码页面</a> | <a href="/">← 工作台</a></p>
<script>
var curRegion='';
async function load(){
    var bid=document.getElementById('batchSel').value;
    var r=await fetch('/scan_stats?batch='+bid+'&region='+curRegion);var d=await r.json();
    var statsDiv=document.getElementById('stats');
    var pct=d.total>0?Math.round(d.scanned/d.total*100):0;
    var boxPct = d.total_expected_boxes>0 ? Math.round(d.scanned_correct_boxes/d.total_expected_boxes*100) : 0;
    statsDiv.innerHTML='<div><div class="n">'+d.total+'</div><div class="l">总单</div></div>'+
        '<div><div class="n">'+d.scanned+'</div><div class="l">已扫 '+pct+'%</div></div>'+
        '<div><div class="n g">'+d.correct+'</div><div class="l">正确</div></div>'+
        '<div><div class="n r">'+d.wrong+'</div><div class="l">异常</div></div>'+
        '<div><div class="n" style="color:#3949ab">'+d.corrections+'</div><div class="l">更正</div></div>'+
        '<div><div style="font-size:16px;font-weight:bold;color:#555">'+d.scanned_correct_boxes+'/'+d.total_expected_boxes+'箱</div><div class="l">箱数校验 '+boxPct+'%</div></div>';
    var un='';    var un='';d.unscanned.forEach(i=>{un+=i[0].substring(0,12)+' | '+i[1]+' | '+i[2]+'\u7bb1 | '+i[5]+'<br>'});
    document.getElementById('unscanned').innerHTML=un||'<span style="color:#999">\u5168\u90e8\u5df2\u626b \u2705</span>';
    var h='<table class="t"><tr><th>\u8d27\u4ef6\u5355\u53f7</th><th>\u5de5\u4eba</th><th>\u7ed3\u679c</th><th>\u7bb1\u6570</th><th>\u533a\u57df</th><th>\u5907\u6ce8</th><th>\u65f6\u95f4</th></tr>';
    d.scans.forEach(s=>{h+='<tr><td>'+(s[0]||'').substring(0,12)+'</td><td>'+(s[1]||'')+'</td><td class="'+(s[2]=='correct'?'g':'r')+'">'+(s[2]=='correct'?'\u2713':'\u2717')+'</td><td>'+(s[3]||'')+'</td><td>'+(s[5]||'')+'</td><td>'+(s[6]||'')+'</td><td>'+(s[4]||'').substr(11,8)+'</td></tr>'});
    h+='</table>';document.getElementById('history').innerHTML=h;
    // 状态标识
    var sb=document.getElementById('statusBadge'),cb=document.getElementById('completeBtn');
    if(d.batch && d.batch[4]){
        sb.style.display='inline-block';
        if(d.batch[4]=='complete'){sb.textContent='✅ 已完成';sb.style.background='#e6f4ea';sb.style.color='#188038';cb.style.display='none';}
        else{sb.textContent='⏳ 扫描中';sb.style.background='#fef7e0';sb.style.color='#ea8600';cb.style.display='inline-block';}
    }
}
async function markComplete(){
    var bid=document.getElementById('batchSel').value;
    if(!confirm('确认标记此批次为完成？'))return;
    var fd=new FormData();fd.append('action','complete');fd.append('batch_name',bid);
    try{var r=await fetch('/run',{method:'POST',body:fd});var d=await r.json();if(d.status==='ok'){alert('✅');load();}else alert('❌ '+d.message);}
    catch(e){alert('❌ '+e.message);}
}
// 加载批次和区域
fetch('/scan_batches').then(r=>r.json()).then(bs=>{
    var sel=document.getElementById('batchSel');
    bs.forEach(b=>{var o=document.createElement('option');o.value=b[0];o.textContent=b[1]+'('+b[3]+'项)';sel.appendChild(o)});
    // 加载区域页签
    if(bs.length>0){
        fetch('/scan_info?batch='+bs[0][0]).then(r=>r.json()).then(d=>{
            var tabs=document.getElementById('regionTabs');
            var allDiv=document.createElement('div');
            allDiv.innerHTML='<div class="rn">全部</div>全部区域';allDiv.onclick=function(){curRegion='';document.querySelectorAll('.rt>div').forEach(el=>el.classList.remove('sel'));this.classList.add('sel');load();};
            tabs.appendChild(allDiv);
            d.regions.forEach(function(rg){
                var rs=d.region_stats[rg]||{total:0,scanned:0};
                var div=document.createElement('div');
                div.innerHTML='<div class="rn">'+rg+'</div>'+rs.total+'项·已扫'+rs.scanned;
                div.onclick=function(){curRegion=rg;document.querySelectorAll('.rt>div').forEach(el=>el.classList.remove('sel'));this.classList.add('sel');load();};
                tabs.appendChild(div);
            });
            allDiv.classList.add('sel');
            load();
        });
    }
});
</script></body></html>'''

SCAN_HISTORY = '''<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"><title>全部批次</title><style>
body{font-family:"Microsoft YaHei","PingFang SC",sans-serif;background:#f0f2f5;color:#333;padding:20px;max-width:1200px;margin:0 auto}
h1{font-size:20px;margin-bottom:12px}
.t{width:100%;border-collapse:collapse;font-size:11px}
.t th,.t td{padding:6px 8px;border:1px solid #ddd;text-align:left}
.t th{background:#1a73e8;color:#fff;white-space:nowrap;position:sticky;top:0}
.g{color:#34a853;font-weight:bold}.r{color:#ea4335;font-weight:bold}
.bb{display:inline-block;background:#e6f4ea;border:1px solid #34a853;border-radius:4px;padding:2px 8px;font-size:10px;color:#188038}
.br{display:inline-block;background:#fce8e6;border:1px solid #ea4335;border-radius:4px;padding:2px 8px;font-size:10px;color:#d93025}
.cr{font-size:11px;color:#999;margin-top:16px}
</style></head><body><h1>全部批次记录</h1>
<div style="overflow-x:auto"><table class="t" id="batchTable"><tr><th>批次名称</th><th>上传时间</th><th>状态</th><th>总货件</th><th>已扫</th><th>正确</th><th>异常</th><th>更正</th><th>未处理</th><th>各区域</th></tr></table></div>
<p class="cr"><a href="/scan_admin">← 返回</a> | <a href="/">工作台</a></p><script>
fetch('/scan_history').then(r=>r.json()).then(bs=>{
    var t=document.getElementById('batchTable');
    bs.forEach(function(b){
        var tr=document.createElement('tr');
        var rgs='';
        if(b.regions){Object.entries(b.regions).forEach(function(e){rgs+=e[0]+'('+e[1].scanned+'/'+e[1].total+') ';});}
        tr.innerHTML='<td><b>'+b.name+'</b></td><td>'+(b.created_at||'').substring(0,16)+'</td>'+'<td><span class="'+(b.status==='complete'?'bb':'br')+'">'+(b.status==='complete'?'已完成':'扫描中')+'</span></td>'+'<td>'+b.items_count+'</td><td class="g">'+b.total_scans+'</td><td class="g">'+b.correct+'</td>'+'<td class="r">'+b.wrong+'</td><td style="color:#3949ab">'+b.corrections+'</td>'+'<td class="r">'+b.still_wrong+'</td><td style="font-size:10px">'+rgs+'</td>';
        t.appendChild(tr);
    });
    if(!bs.length){var tr2=document.createElement('tr');tr2.innerHTML='<td colspan="10" style="text-align:center;color:#999">暂无数据</td>';t.appendChild(tr2);}
});
</script></body></html>'''

WORKSHOP_PAGE = '''<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=1.0,user-scalable=no"><title>✅车间加工(手机端)</title><style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:"Microsoft YaHei","PingFang SC",sans-serif;background:#f5f5f5;color:#333;padding:0;max-width:500px;margin:0 auto}
.hd{background:linear-gradient(135deg,#2d3748,#4a5568);color:#fff;padding:14px 16px;position:sticky;top:0;z-index:10}
.hd h1{font-size:17px}.hd p{font-size:11px;color:#cbd5e0;margin-top:2px}
.sel{width:100%;padding:8px;border:1px solid #e2e8f0;border-radius:6px;font-size:12px;margin:10px 12px;width:calc(100% - 24px);background:#fff}
.filter-bar{display:flex;gap:4px;margin:4px 12px 8px}
.filter-bar button{flex:1;padding:6px;border:1px solid #3182ce;border-radius:6px;font-size:12px;cursor:pointer;background:#fff;color:#3182ce;font-weight:500}
.filter-bar button.on{background:#e53e3e;border-color:#e53e3e;color:#fff}
.pri-badge{display:inline-block;background:#e53e3e;color:#fff;font-size:9px;padding:1px 5px;border-radius:3px;margin-left:6px;vertical-align:middle}
.li{list-style:none;padding:0;margin:0}
.li li{background:#fff;margin:8px 12px;border-radius:10px;box-shadow:0 1px 3px rgba(0,0,0,.06);overflow:hidden}
.li .top{display:flex;align-items:center;padding:10px 12px 6px;gap:8px}
.li .st{width:8px;height:8px;border-radius:50%;flex-shrink:0}
.li .st.pd{background:#a0aec0}.li .st.pr{background:#d69e2e;animation:bl 1s infinite}
.li .st.cp{background:#38a169}
@keyframes bl{50%{opacity:.4}}
.li .sk{font-size:13px;font-weight:600;flex:1}
.li .nm{font-size:11px;color:#718096;margin:0 12px 8px}
.li .qt{font-size:11px;color:#718096;margin:0 12px 2px}
.li .btns{display:flex;gap:6px;padding:6px 12px 10px}
.btns button{flex:1;padding:8px;border:none;border-radius:6px;font-size:12px;cursor:pointer;font-weight:500}
.btn-s{background:#3182ce;color:#fff}.btn-c{background:#38a169;color:#fff}.btn-x{background:#e53e3e;color:#fff}.btn-d{background:#e2e8f0;color:#718096;cursor:not-allowed}
.people-bar{display:flex;align-items:center;gap:4px;margin:-2px 12px 10px;font-size:12px;color:#4a5568}
.people-bar input{width:56px;padding:3px 6px;border:1px solid #cbd5e0;border-radius:4px;font-size:12px;text-align:center;background:#fff}
.people-bar input:focus{outline:none;border-color:#3182ce;box-shadow:0 0 0 2px rgba(49,130,206,.15)}
.ov{position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,.5);z-index:50;display:none;align-items:center;justify-content:center}
.ov .bx{background:#fff;border-radius:12px;padding:20px;width:300px;max-width:90%}
.ov .bx h3{font-size:15px;margin-bottom:12px}
.ov .bx input{width:100%;padding:10px;border:2px solid #3182ce;border-radius:8px;font-size:16px;text-align:center;margin-bottom:12px}
.ov .bx button{padding:8px 20px;border:none;border-radius:6px;font-size:13px;cursor:pointer;margin:0 4px}
.na{text-align:center;font-size:12px;color:#999;padding:40px 20px}
.cr{font-size:11px;color:#999;text-align:center;padding:12px;margin-top:8px}
	.bt{font-size:11px;color:#cbd5e0;text-decoration:none}
	.pri-item{background:#fffaf0;border-left:3px solid #e53e3e;margin-bottom:6px!important}</style></head><body>
				<div class="hd"><div style="display:flex;align-items:center;gap:8px"><h1 style="flex:1">🔧 车间加工列表(手机端) <span style="font-size:11px;color:#a0aec0;font-weight:400">'''+VERSION+'''</span></h1><a href="#" style="color:#fff;text-decoration:none;font-size:18px" onclick="showQR();return false">📱</a></div><p id="batchInfo">全部加工单</p></div>
			<div class="filter-bar"><button id="f_pending" class="on" onclick="setFilter('pending')">📋 待处理</button><button id="f_processing" onclick="setFilter('processing')">🔧 加工中</button><button id="f_completed" onclick="setFilter('completed')">✅ 已完成</button><button id="f_priority" onclick="setFilter('priority')">⭐ 优先</button></div>
			<div style="margin:0 12px 4px;padding:6px 10px;background:#fffbeb;border:1px solid #f6e05e;border-radius:6px;font-size:11px;color:#975a16">📢 开始、暂停、完工环节，须第一时间点击对应按键（如打包、下班点暂停）</div>
			<div id="priNotice" style="display:none;margin:0 12px 6px;padding:8px 12px;background:#fff5f5;border:1px solid #fed7d7;border-radius:8px;font-size:12px;color:#e53e3e;font-weight:600;text-align:center;cursor:pointer" onclick="setFilter('priority')">⭐ 有 <span id="priCount">0</span> 个优先订单待处理</div>
				<div class="people-bar"><span>👥</span><input id="peopleInput" type="number" min="1" placeholder="人数" onchange="savePeople()"><span style="font-size:11px;color:#999" id="peopleLabel">上班: -</span></div>
		<div style="margin:0 12px 6px"><input id="searchBoxWs" type="text" placeholder="🔍 搜索SKU、品名、单号..." oninput="doSearchWs()" style="width:100%;padding:6px 10px;border:1px solid #e2e8f0;border-radius:6px;font-size:12px;outline:none;box-sizing:border-box"></div>
<ul class="li" id="jobList"></ul>
<div class="ov" id="modal"><div class="bx"><h3 id="mTitle">完成加工</h3><p style="font-size:12px;color:#666;margin-bottom:6px" id="mSku"></p><input id="mQty" type="number" placeholder="输入完成数量" min="1" onkeydown="if(event.key==='Enter')confirmComplete()"><div style="display:flex;gap:6px;margin-bottom:10px"><button class="btn-s" onclick="fillQty()" style="flex:1;font-size:11px;padding:6px">📋 订单数量</button></div><button class="btn-s" onclick="confirmComplete()">确定完成</button><button onclick="closeModal()" style="background:#e2e8f0">取消</button></div></div>
<div class="ov" id="priAlert" style="display:none;z-index:100"><div class="bx" style="text-align:center;padding:20px" onclick="event.stopPropagation()"><div style="font-size:48px;margin-bottom:10px">⭐</div><h2 style="color:#e53e3e;margin-bottom:8px">优先订单！</h2><p style="font-size:14px;color:#4a5568;margin-bottom:16px">有 <span id="priAlertCount" style="font-weight:700;color:#e53e3e">0</span> 个优先订单待处理</p><button class="btn-s" onclick="dismissPriAlert()" style="padding:10px 30px;font-size:15px">知道了</button></div></div>
<script>
var curItemId = 0;
var curFilter = 'pending';
function setFilter(mode){
    curFilter=mode;
    document.querySelectorAll('.filter-bar button').forEach(function(b){b.className=''});
    document.getElementById('f_'+mode).className='on';
    loadJobs();
}
async function loadJobs(){
	    var ppl=localStorage.getItem('default_people')||'1';
	    var r=await fetch('/workshop_data?people='+ppl);var items=await r.json();
	    window.allItems=items;
	    var kw=document.getElementById('searchBoxWs').value.trim().toLowerCase();
	    if(kw){items=items.filter(function(i){return (i.sku&&i.sku.toLowerCase().indexOf(kw)>=0)||(i.name&&i.name.toLowerCase().indexOf(kw)>=0)||(i.job_number&&i.job_number.toLowerCase().indexOf(kw)>=0);});}
	    var ul=document.getElementById('jobList');ul.innerHTML='';
    if(!items.length){ul.innerHTML='<div class="na">暂无加工单<br>请先在工作台上传</div>';return}
    var info=document.getElementById('batchInfo');
    if(info) info.textContent='全部加工单 ('+items.length+'项)';
    // 优先订单提醒
    var priPending=items.filter(function(i){return i.status==='pending' && i.priority==1}).length;
    var el=document.getElementById('priNotice');
    if(el){if(priPending>0){el.style.display='block';document.getElementById('priCount').textContent=priPending}else el.style.display='none'}
	    // 优先单强提醒（持续响铃+弹窗，点"知道了"才停）
		    if(priPending>0 && (!window.lastPriCount || priPending>window.lastPriCount)){
		        window.priAlertDismissed=false;
		        document.getElementById('priAlertCount').textContent=priPending;
		        document.getElementById('priAlert').style.display='flex';
		        startPriBeep();
	    }
	    window.lastPriCount=priPending;
    // 更新筛选按钮数量
    var ti=items;
	    document.getElementById('f_pending').textContent='📋 待处理 ('+ti.filter(function(x){return x.status==='pending'||x.status==='paused'}).length+')';
	    document.getElementById('f_processing').textContent='🔧 加工中 ('+ti.filter(function(x){return x.status==='processing'}).length+')';
	    document.getElementById('f_completed').textContent='✅ 已完成 ('+ti.filter(function(x){return x.status==='completed'}).length+')';
	    document.getElementById('f_priority').textContent='⭐ 优先 ('+ti.filter(function(x){return x.priority==1 && x.status!=='completed'}).length+')';
	    if(curFilter==='all') items=items.filter(function(i){return i.status!=='completed'});
	    if(curFilter==='pending') items=items.filter(function(i){return i.status==='pending'||i.status==='paused'});
	    if(curFilter==='priority') items=items.filter(function(i){return i.priority==1 && i.status!=='completed'});
	    if(curFilter==='processing') items=items.filter(function(i){return i.status==='processing'});
	    if(curFilter==='completed') items=items.filter(function(i){return i.status==='completed'});
	    items.sort(function(a,b){if(a.priority!==b.priority)return b.priority-a.priority;return (a.sku||'').localeCompare(b.sku||'');});
	    items.forEach(function(i){
	        var li=document.createElement('li');
	        if(i.priority) li.className='pri-item';
	        var statusLabel={pending:'待处理',processing:'加工中',paused:'已暂停',completed:'已完成'};
	        var stClass = i.status;
	        var btns='';
		        if(i.status==='pending'){btns='<button class="btn-s" onclick="startJob('+i.id+')">▶ 开始加工</button>';}
		        else if(i.status==='processing'){btns='<button class="btn-c" onclick="openComplete(this)" data-id="'+i.id+'" data-sku="'+i.sku+'" data-qty="'+i.qty+'">✔ 完成</button><button class="btn-x" style="background:#d69e2e" onclick="pauseJob('+i.id+')">⏸ 暂停</button><button class="btn-x" onclick="cancelJob('+i.id+')">✖ 取消</button>';}
		        else if(i.status==='paused'){btns='<button class="btn-s" onclick="resumeJob('+i.id+')">▶ 取消暂停</button><button class="btn-x" onclick="cancelJob('+i.id+')">✖ 取消</button>';}
		        else{btns='<button class="btn-d">✔ 已完成</button>';}
	            var info='<div class="qt">数量: '+i.qty+'</div>';
	            if((i.status==='processing'||i.status==='completed') && i.worker){var wm=i.worker.match(/x(\d+)/);var wc=wm?wm[1]:'?';var st=i.started?i.started.substr(11,5):'?';info+='<div class="qt">👥 '+wc+'人 | 🕐 '+st+' 开始</div>';}
	            if(i.status==='completed' && i.done_qty) info+='<div class="qt">完成: '+i.done_qty+'件 | '+i.completed.substr(11,5)+'</div>';
	        var priBadge = i.priority ? '<span class="pri-badge">⭐优先</span>' : '';
		        var jobNoHtml = i.job_number ? '<span style="font-size:10px;color:#718096">#'+i.job_number+(i.notes?' <span style="color:#a0aec0">'+i.notes+'</span>':'')+' </span>' : '';
		        li.innerHTML='<div class="top"><span class="st '+stClass+'"></span><span class="sk">'+i.sku+priBadge+'</span>'+jobNoHtml+'<span style="font-size:10px;color:#999">'+statusLabel[i.status]+'</span></div>'+'<div class="nm">'+(i.name||'')+'</div>'+info+'<div class="btns">'+btns+'</div>';
		        ul.appendChild(li);
		    });
			    try { syncChannel.postMessage('refresh'); } catch(e) {}
			}
			function doSearchWs(){ loadJobs(); }
		async function startJob(id){
		    var ppl=document.getElementById('peopleInput')?document.getElementById('peopleInput').value:'';
    if(!ppl||parseInt(ppl)<1)ppl='1';
    localStorage.setItem('default_people',ppl);
    var wn='车间工人 x'+ppl;
    var fd=new FormData();fd.append('action','start_job');fd.append('batch_name',id);fd.append('worker',wn);
    try{var r=await fetch('/run',{method:'POST',body:fd});var d=await r.json();if(d.status==='ok'){alert('✅');loadJobs();}else alert('❌ '+d.message);}
    catch(e){alert('❌ '+e.message);}
}
function savePeople(){
    var v=document.getElementById('peopleInput').value;
    if(v&&parseInt(v)>=1){localStorage.setItem('default_people',v);document.getElementById('peopleLabel').textContent='上班: '+v+'人';}
    try { syncChannel.postMessage('refresh'); } catch(e) {}
    if(typeof loadJobs==='function') loadJobs();
    if(typeof loadBoard==='function') loadBoard();
}
async function cancelJob(id){
    if(!confirm('确定取消此加工任务？'))return;
    var fd=new FormData();fd.append('action','cancel_job');fd.append('batch_name',id);
    try{var r=await fetch('/run',{method:'POST',body:fd});var d=await r.json();if(d.status==='ok'){alert('✅');loadJobs();}else alert('❌ '+d.message);}
    catch(e){alert("❌ "+e.message);}
    }
	async function pauseJob(id){
	    var fd=new FormData();fd.append('action','pause_job');fd.append('batch_name',id);
	    try{var r=await fetch('/run',{method:'POST',body:fd});var d=await r.json();if(d.status==='ok'){alert('⏸');loadJobs();}else alert('❌ '+d.message);}
	    catch(e){alert('❌ '+e.message);}
	}
		async function resumeJob(id){
		    var fd=new FormData();fd.append('action','resume_job');fd.append('batch_name',id);
		    try{var r=await fetch('/run',{method:'POST',body:fd});var d=await r.json();if(d.status==='ok'){alert('▶');loadJobs();}else alert('❌ '+d.message);}
		    catch(e){alert('❌ '+e.message);}
		}
		// 优先单强提醒（含系统通知）
		var priBeepTimer=null;
		function startPriBeep(){
		    // 请求系统通知权限并发送锁屏通知
		    if('Notification' in window && Notification.permission!=='denied'){
		        Notification.requestPermission().then(function(p){
		            if(p==='granted'){
		                var n=new Notification('⭐ 优先订单！',{body:'有 '+document.getElementById('priAlertCount').textContent+' 个优先订单待处理',tag:'pri-order',requireInteraction:true});
		                setTimeout(function(){n.close()},10000);
		            }
		        });
		    }
		    if(priBeepTimer) return;
		    function beep(){
		        try{
		            var ac=new(window.AudioContext||window.webkitAudioContext)();
		            var osc=ac.createOscillator();var g=ac.createGain();
		            osc.connect(g);g.connect(ac.destination);
		            osc.frequency.value=880;osc.type='sine';
		            g.gain.setValueAtTime(0.3,ac.currentTime);
		            g.gain.exponentialRampToValueAtTime(0.01,ac.currentTime+0.3);
		            osc.start(ac.currentTime);osc.stop(ac.currentTime+0.3);
		        }catch(e){}
		        if(!window.priAlertDismissed) priBeepTimer=setTimeout(beep,2000);
		    }
		    beep();
		}
		function dismissPriAlert(){
		    window.priAlertDismissed=true;
		    document.getElementById('priAlert').style.display='none';
		    if(priBeepTimer){clearTimeout(priBeepTimer);priBeepTimer=null}
		}
	function openComplete(btn){curItemId=btn.dataset.id;curOrderQty=btn.dataset.qty||'';document.getElementById('mSku').textContent=btn.dataset.sku;document.getElementById('mQty').value='';document.getElementById('modal').style.display='flex';}
	    function fillQty(){document.getElementById('mQty').value=curOrderQty||''}
function closeModal(){document.getElementById('modal').style.display='none';}
async function confirmComplete(){
    var qty=document.getElementById('mQty').value;
    if(!qty||parseInt(qty)<=0){alert('请输入有效数量');return}
    var fd=new FormData();fd.append('action','complete_job');fd.append('batch_name',curItemId);fd.append('done_qty',qty);
    try{var r=await fetch('/run',{method:'POST',body:fd});var d=await r.json();closeModal();if(d.status==='ok'){alert(d.message);loadJobs();}else alert('❌ '+d.message);}
    catch(e){alert('❌ '+e.message);}
}
var savedPpl=localStorage.getItem('default_people');if(savedPpl){document.getElementById('peopleInput').value=savedPpl;document.getElementById('peopleLabel').textContent='上班: '+savedPpl+'人';}
document.getElementById('batchInfo').textContent='全部加工单';
	// 产能计算器
	
		async 		// 跨页面实时同步
		var syncChannel = new BroadcastChannel('workshop_sync');
	syncChannel.onmessage = function(e) { if (e.data === 'refresh') loadJobs(); };
	window.addEventListener('storage', function(e) { if (e.key === 'default_people') loadJobs(); });
	loadJobs();
    // Auto-refresh

function autoRefresh(){loadJobs();setTimeout(autoRefresh,2000);}
	autoRefresh();
		function showQR(){
		    var self=this;
		    fetch('/get_ip').then(function(r){return r.json()}).then(function(d){
		        var url='http://'+d.ip+':8932/workshop';
		        document.getElementById('qrImg').src='https://api.qrserver.com/v1/create-qr-code/?size=300x300&data='+encodeURIComponent(url);
		        document.getElementById('qrModal').style.display='flex';
		    }).catch(function(){
		        var url=window.location.href;
		        document.getElementById('qrImg').src='https://api.qrserver.com/v1/create-qr-code/?size=300x300&data='+encodeURIComponent(url);
		        document.getElementById('qrModal').style.display='flex';
		    });
		}
		function downloadQR(){
		    var url=window.location.href;
		    var a=document.createElement('a');a.href='https://api.qrserver.com/v1/create-qr-code/?size=500x500&data='+encodeURIComponent(url);a.download='workshop_qr.png';a.click();
		}
		</script>
		
	<div class="ov" id="qrModal" style="display:none" onclick="this.style.display='none'"><div class="bx" style="text-align:center" onclick="event.stopPropagation()"><h3 style="margin-bottom:10px">📱 手机端扫码打开</h3><img id="qrImg" src="" style="width:200px;height:200px;border:1px solid #e2e8f0;border-radius:8px;background:#fff;margin-bottom:10px"><br><button class="btn-s" onclick="downloadQR()">⬇ 下载二维码</button><button onclick="document.getElementById('qrModal').style.display='none'" style="background:#e2e8f0;padding:6px 14px;border:none;border-radius:6px;margin-left:6px;cursor:pointer">关闭</button></div></div>
		</body></html>'''
WORKSHOP_ADMIN = '''<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=1.0,user-scalable=no"><title>管理后台</title><style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:"Microsoft YaHei","PingFang SC",sans-serif;background:#f5f5f5;color:#333;padding:0;max-width:500px;margin:0 auto}
.hd{background:linear-gradient(135deg,#2d3748,#4a5568);color:#fff;padding:14px 16px;position:sticky;top:0;z-index:10}
.hd h1{font-size:17px}.hd p{font-size:11px;color:#cbd5e0;margin-top:2px}
.sel{width:100%;padding:8px;border:1px solid #e2e8f0;border-radius:6px;font-size:12px;margin:10px 12px;width:calc(100% - 24px);background:#fff}
.filter-bar{display:flex;gap:4px;margin:4px 12px 8px}
.filter-bar button{flex:1;padding:6px;border:1px solid #3182ce;border-radius:6px;font-size:12px;cursor:pointer;background:#fff;color:#3182ce;font-weight:500}
.filter-bar button.on{background:#e53e3e;border-color:#e53e3e;color:#fff}
.pri-badge{display:inline-block;background:#e53e3e;color:#fff;font-size:9px;padding:1px 5px;border-radius:3px;margin-left:6px;vertical-align:middle}
.li{list-style:none;padding:0;margin:0}
.li li{background:#fff;margin:8px 12px;border-radius:10px;box-shadow:0 1px 3px rgba(0,0,0,.06);overflow:hidden}
.li .top{display:flex;align-items:center;padding:10px 12px 6px;gap:8px}
.li .st{width:8px;height:8px;border-radius:50%;flex-shrink:0}
.li .st.pd{background:#a0aec0}.li .st.pr{background:#d69e2e;animation:bl 1s infinite}
.li .st.cp{background:#38a169}
@keyframes bl{50%{opacity:.4}}
.li .sk{font-size:13px;font-weight:600;flex:1}
.li .nm{font-size:11px;color:#718096;margin:0 12px 8px}
.li .qt{font-size:11px;color:#718096;margin:0 12px 2px}
.li .btns{display:flex;gap:6px;padding:6px 12px 10px}
.btns button{flex:1;padding:8px;border:none;border-radius:6px;font-size:12px;cursor:pointer;font-weight:500}
.btn-s{background:#3182ce;color:#fff}.btn-c{background:#38a169;color:#fff}.btn-x{background:#e53e3e;color:#fff}.btn-d{background:#e2e8f0;color:#718096;cursor:not-allowed}
.people-bar{display:flex;align-items:center;gap:4px;margin:-2px 12px 10px;font-size:12px;color:#4a5568}
.people-bar input{width:56px;padding:3px 6px;border:1px solid #cbd5e0;border-radius:4px;font-size:12px;text-align:center;background:#fff}
.people-bar input:focus{outline:none;border-color:#3182ce;box-shadow:0 0 0 2px rgba(49,130,206,.15)}
.ov{position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,.5);z-index:50;display:none;align-items:center;justify-content:center}
.ov .bx{background:#fff;border-radius:12px;padding:20px;width:300px;max-width:90%}
.ov .bx h3{font-size:15px;margin-bottom:12px}
.ov .bx input{width:100%;padding:10px;border:2px solid #3182ce;border-radius:8px;font-size:16px;text-align:center;margin-bottom:12px}
.ov .bx button{padding:8px 20px;border:none;border-radius:6px;font-size:13px;cursor:pointer;margin:0 4px}
.na{text-align:center;font-size:12px;color:#999;padding:40px 20px}
.cr{font-size:11px;color:#999;text-align:center;padding:12px;margin-top:8px}
.bt{font-size:11px;color:#cbd5e0;text-decoration:none}
    .btn-pri{background:#e53e3e;color:#fff;flex:0!important;padding:8px 10px!important}
    .btn-pri-off{background:#fff;color:#e53e3e;border:1px solid #e53e3e!important;flex:0!important;padding:7px 9px!important}
    .btn-xs{background:#3182ce;color:#fff;padding:4px 8px!important;font-size:11px;flex:0!important;border-radius:4px}
    .progress{height:4px;background:#edf2f7;border-radius:2px;margin:4px 0;overflow:hidden}
    .progress .bar{height:100%;border-radius:2px;background:linear-gradient(90deg,#d69e2e,#ecc94b);transition:width .5s}
.del-bar{display:flex;align-items:center;gap:6px;margin:4px 12px 6px;padding:6px 8px;background:#fff;border-radius:6px;font-size:11px;border:1px solid #fed7d7}
.del-bar label{display:flex;align-items:center;gap:3px;cursor:pointer;color:#4a5568}
.del-bar label input{cursor:pointer}
.del-btn{background:#e53e3e;color:#fff;border:none;border-radius:4px;padding:4px 10px;font-size:11px;cursor:pointer}
.del-btn:hover{background:#c53030}
.cb-item{width:14px;height:14px;cursor:pointer;margin-right:4px;flex-shrink:0}</style></head><body>
	<div class="hd" style="display:flex;align-items:center;gap:10px"><div><h1>🛡 管理后台</h1><p id="batchInfo">全部加工单</p><p id="debugInfo" style="font-size:10px;color:#ff0;text-align:center">加载中...</p></div></div>
<div class="filter-bar"><button id="af_all" class="on" onclick="setFilter('all')">📋 待处理</button><button id="af_processing" onclick="setFilter('processing')">🔧 加工中</button><button id="af_completed" onclick="setFilter('completed')">✅ 已完成</button><button id="af_priority" onclick="setFilter('priority')">⭐ 优先</button></div>
<div class="people-bar"><span>👥</span><input id="peopleInput" type="number" min="1" placeholder="人数" onchange="savePeople()"><span style="font-size:11px;color:#999" id="peopleLabel">上班: -</span></div>
<div style="margin:6px 12px;padding:8px;background:#fff;border-radius:8px;font-size:11px;box-shadow:0 1px 3px rgba(0,0,0,.06)">
<div style="font-weight:600;font-size:12px;margin-bottom:6px">📦 已上传批次</div>
<div id="batchList" style="max-height:120px;overflow-y:auto"></div>
</div>
<div class="del-bar"><label><input type="checkbox" id="selectAll" onchange="toggleAll()"> 全选</label><button class="del-btn" onclick="deleteSelected()">🗑 删除选中</button><span id="selCount" style="color:#999">0项</span></div>
<ul class="li" id="jobList"></ul>
<div class="ov" id="modal"><div class="bx"><h3 id="mTitle">完成加工</h3><p style="font-size:12px;color:#666;margin-bottom:6px" id="mSku"></p><input id="mQty" type="number" placeholder="输入完成数量" min="1" onkeydown="if(event.key==='Enter')confirmComplete()"><button class="btn-s" onclick="confirmComplete()">确定完成</button><button onclick="closeModal()" style="background:#e2e8f0">取消</button></div></div>
<p class="cr"><a class="bt" href="/workshop">车间页面</a> | <a class="bt" href="/">工作台</a></p>
<script>
var curItemId = 0;
var curFilter = 'all';
function setFilter(mode){
    curFilter=mode;
    document.querySelectorAll('.filter-bar button').forEach(function(b){b.className='';b.style.background='#fff';b.style.color='#3182ce'});
    var el=document.getElementById('af_'+mode);
    el.className='on';el.style.background='#e53e3e';el.style.color='#fff';
    loadJobs();
}
async function togglePri(id,pri){
    var fd=new FormData();fd.append('action','set_priority');fd.append('batch_name',id);fd.append('priority',pri);
    try{var r=await fetch('/run',{method:'POST',body:fd});var d=await r.json();if(d.status==='ok')loadJobs();else alert(d.message);}
    catch(e){alert(e.message);}
}
async function loadJobs(){
    var ppl=localStorage.getItem('default_people')||'1';
    var r=await fetch('/workshop_data?people='+ppl);var items=await r.json();
    var ul=document.getElementById('jobList');ul.innerHTML='';
    if(!items.length){ul.innerHTML='<div class="na">暂无加工单<br>请先在工作台上传</div>';return}
    var info=document.getElementById('batchInfo');
    if(info) info.textContent='全部加工单 ('+items.length+'项)';
    // 优先订单提醒
    var priPending=items.filter(function(i){return i.status==='pending' && i.priority==1}).length;
    var el=document.getElementById('priNotice');
    if(el){if(priPending>0){el.style.display='block';document.getElementById('priCount').textContent=priPending}else el.style.display='none'}
    if(curFilter==='all') items=items.filter(function(i){return i.status!=='completed'});
    if(curFilter==='priority') items=items.filter(function(i){return i.priority==1});
    if(curFilter==='processing') items=items.filter(function(i){return i.status==='processing'});
    if(curFilter==='completed') items=items.filter(function(i){return i.status==='completed'});
    items.forEach(function(i){
        var li=document.createElement('li');
        if(i.priority) li.className='pri-item';
        var statusLabel={pending:'待处理',processing:'加工中',paused:'已暂停',completed:'已完成'};
        var stClass = i.status;
            var priBtn = i.priority ? '<button class="btn-pri" onclick="togglePri('+i.id+',0)">★优</button>' : '<button class="btn-pri-off" onclick="togglePri('+i.id+',1)">☆设优</button>';
            if(i.status==='completed') priBtn='';
            var btns='';
                if(i.status==='pending'){btns='<button class="btn-xs" onclick="startJob('+i.id+')">▶ 开始</button>';}
                else if(i.status==='processing'){btns='<button class="btn-c" onclick="openComplete(this)" data-id="'+i.id+'" data-sku="'+i.sku+'">✔ 完成</button><button class="btn-x" style="background:#d69e2e" onclick="pauseJob('+i.id+')">⏸ 暂停</button><button class="btn-x" onclick="cancelJob('+i.id+')">✖ 取消</button>';}
                else{btns='<button class="btn-d">✔ 已完</button>';}
                var info='<div class="qt">数量: '+i.qty+'</div>';
                if(i.est_hours) info+='<div class="qt" style="color:#718096">⏱ 约 '+i.est_hours+'h</div>';
                if(i.status==='processing' && i.worker) info+='<div class="qt">工人: '+i.worker+' | '+i.started.substr(11,5)+'</div>';
                if(i.status==='processing') info+='<div class="qt" style="color:#d69e2e">⏱ '+(i.est_time?'预计完成：<b>'+i.est_time+'</b> (~'+i.est_min+'分钟)':'暂无数据')+'</div>';
                if(i.status==='processing' && i.started && i.est_min>0){
                    var stTime=new Date(i.started).getTime();
                    var elMin=(Date.now()-stTime)/60000;
                    var pct=Math.min(100,Math.round(elMin/i.est_min*100));
                info+='<div class="progress"><div class="bar" style="width:'+Math.max(3,pct)+'%"></div></div>';
            }
            if(i.status==='completed' && i.done_qty) info+='<div class="qt">完成: '+i.done_qty+'件 | '+i.completed.substr(11,5)+'</div>';
            var priBadge = i.priority ? '<span class="pri-badge">⭐优先</span>' : '';
            var jobNoHtml2 = i.job_number ? '<span style="font-size:10px;color:#718096">#'+i.job_number+' </span>' : '';
            li.innerHTML='<div class="top"><input type="checkbox" class="cb-item" data-id="'+i.id+'" onchange="updateSelCount()"><span class="st '+stClass+'"></span>'+jobNoHtml2+'<span class="sk">'+i.sku+priBadge+'</span><span style="font-size:10px;color:#999">'+statusLabel[i.status]+'</span></div>'+'<div class="nm">'+(i.name||'')+'</div>'+info+'<div class="btns">'+priBtn+btns+'</div>';
        ul.appendChild(li);
    });
}
async function startJob(id){
    var ppl=document.getElementById('peopleInput')?document.getElementById('peopleInput').value:'';
    if(!ppl||parseInt(ppl)<1)ppl='1';
    localStorage.setItem('default_people',ppl);
    var wn='车间工人 x'+ppl;
    var fd=new FormData();fd.append('action','start_job');fd.append('batch_name',id);fd.append('worker',wn);
    try{var r=await fetch('/run',{method:'POST',body:fd});var d=await r.json();if(d.status==='ok'){alert('✅');loadJobs();}else alert('❌ '+d.message);}
    catch(e){alert('❌ '+e.message);}
}
function savePeople(){
    var v=document.getElementById('peopleInput').value;
    if(v&&parseInt(v)>=1){localStorage.setItem('default_people',v);document.getElementById('peopleLabel').textContent='上班: '+v+'人';}
    
}
async function cancelJob(id){
    if(!confirm('确定取消此加工任务？'))return;
    var fd=new FormData();fd.append('action','cancel_job');fd.append('batch_name',id);
    try{var r=await fetch('/run',{method:'POST',body:fd});var d=await r.json();if(d.status==='ok'){alert('✅');loadJobs();}else alert('❌ '+d.message);}
    catch(e){alert("❌ "+e.message);}
}
function toggleAll(){
    var checked=document.getElementById('selectAll').checked;
    document.querySelectorAll('.cb-item').forEach(function(c){c.checked=checked;});
    updateSelCount();
    }catch(e){document.getElementById('batchInfo').textContent='❌ 加载失败: '+e.message;console.error(e);}
}
function updateSelCount(){
    var n=document.querySelectorAll('.cb-item:checked').length;
    document.getElementById('selCount').textContent=n+'项';
}
async function deleteSelected(){
    var ids=[];
    document.querySelectorAll('.cb-item:checked').forEach(function(c){ids.push(c.getAttribute('data-id'));});
    if(!ids.length){alert('请先选择要删除的项');return;}
    if(!confirm('确定删除选中的 '+ids.length+' 项？此操作不可撤销！'))return;
    var fd=new FormData();fd.append('action','delete_jobs');fd.append('ids',ids.join(','));
    try{var r=await fetch('/run',{method:'POST',body:fd});var d=await r.json();if(d.status==='ok'){alert(d.message);loadJobs();}else alert('❌ '+d.message);}
    catch(e){alert('❌ '+e.message);}
}
function openComplete(btn){curItemId=btn.dataset.id;document.getElementById('mSku').textContent=btn.dataset.sku;document.getElementById('mQty').value='';document.getElementById('modal').style.display='flex';}
function closeModal(){document.getElementById('modal').style.display='none';}
async function confirmComplete(){
    var qty=document.getElementById('mQty').value;
    if(!qty||parseInt(qty)<=0){alert('请输入有效数量');return}
    var fd=new FormData();fd.append('action','complete_job');fd.append('batch_name',curItemId);fd.append('done_qty',qty);
    try{var r=await fetch('/run',{method:'POST',body:fd});var d=await r.json();closeModal();if(d.status==='ok'){alert(d.message);loadJobs();}else alert('❌ '+d.message);}
    catch(e){alert('❌ '+e.message);}
}
var savedPpl=localStorage.getItem('default_people');if(savedPpl){document.getElementById('peopleInput').value=savedPpl;document.getElementById('peopleLabel').textContent='上班: '+savedPpl+'人';}
document.getElementById('batchInfo').textContent='全部加工单';
loadJobs();
// Batch list
function loadBatchList(){
    fetch('/job_batches').then(function(r){return r.json()}).then(function(bs){
        var div=document.getElementById('batchList');
        if(!div)return;
        var h='';
        bs.forEach(function(b){h+='<div style=\"padding:2px 0;border-bottom:1px solid #eee;display:flex;justify-content:space-between\"><span>'+b.name+'</span><span style=\"color:#999\">'+b.count+'项</span></div>';});
        div.innerHTML=h||'<span style=\"color:#999\">暂无批次</span>';
    }).catch(function(){});
}
loadBatchList();
// Auto-refresh every 10s to sync with workshop

function autoRefresh(){loadJobs();loadBatchList();setTimeout(autoRefresh,3000);}
autoRefresh();
    </script></body></html>'''
    
    
    
# ====== 车间看板（分栏看板） ======
WORKSHOP_BOARD = '''<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=1.5"><title>车间看板</title><style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:"Microsoft YaHei","PingFang SC",sans-serif;background:#f0f2f5;color:#1a202c;min-height:100vh}
.hd{background:linear-gradient(135deg,#2d3748,#4a5568);color:#fff;padding:12px 20px}
.hd-top{display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:6px}
.hd h1{font-size:17px}.hd .sub{font-size:11px;color:#cbd5e0;margin-top:2px}
.people-bar{display:flex;align-items:center;gap:4px;font-size:12px;color:#e2e8f0}
.people-bar input{width:48px;padding:2px 6px;border:1px solid #cbd5e0;border-radius:4px;font-size:12px;text-align:center;background:rgba(255,255,255,.15);color:#fff}
.people-bar input:focus{outline:none;border-color:#63b3ed}
.nav-link{font-size:11px;color:#cbd5e0;text-decoration:none;margin-left:8px}
.nav-link:hover{color:#fff}
.stats{display:flex;gap:6px;padding:10px 16px;background:#fff;border-bottom:1px solid #e2e8f0;flex-wrap:wrap}
.stat-item{text-align:center;flex:1;min-width:60px;padding:4px 0}
.stat-item .num{font-size:20px;font-weight:700}
.stat-item .lbl{font-size:10px;color:#718096;margin-top:1px}
.stat-item.gray .num{color:#718096}.stat-item.blue .num{color:#3182ce}
.stat-item.orange .num{color:#d69e2e}.stat-item.green .num{color:#38a169}
.stat-item.red .num{color:#e53e3e}
.board{display:flex;gap:12px;padding:12px 16px;min-height:calc(100vh - 130px);overflow-x:auto;align-items:flex-start}
.col{flex:1;min-width:260px;max-width:420px;background:#f7fafc;border-radius:10px;padding:8px 10px}
.col-hd{font-size:13px;font-weight:600;padding:4px 6px 8px;display:flex;align-items:center;gap:6px}
.col-hd .cnt{background:#e2e8f0;color:#4a5568;font-size:10px;padding:1px 8px;border-radius:10px}
.card{background:#fff;border-radius:8px;padding:10px 12px;margin-bottom:8px;box-shadow:0 1px 3px rgba(0,0,0,.06);border-left:3px solid #e2e8f0}
.card.pending{border-left-color:#a0aec0}.card.processing{border-left-color:#d69e2e}.card.completed{border-left-color:#38a169}
.card .top{display:flex;align-items:center;gap:6px;margin-bottom:3px}
.card .sk{font-size:13px;font-weight:600;flex:1}
.card .nm{font-size:11px;color:#718096;margin-bottom:3px}
.card .qt{font-size:11px;color:#718096;margin-bottom:2px}
.card .info{display:flex;flex-wrap:wrap;gap:2px 10px;font-size:11px;color:#4a5568;margin:4px 0}
.card .info span{display:inline-flex;align-items:center;gap:3px}
.pri-badge{display:inline-block;background:#e53e3e;color:#fff;font-size:9px;padding:1px 5px;border-radius:3px;vertical-align:middle}
.btns{display:flex;gap:4px;margin-top:5px}
.btns button{flex:1;padding:5px;border:none;border-radius:5px;font-size:11px;cursor:pointer;font-weight:500}
.btn-s{background:#3182ce;color:#fff}.btn-c{background:#38a169;color:#fff}.btn-x{background:#e53e3e;color:#fff}
.btn-pri{background:#e53e3e;color:#fff;flex:0!important;padding:5px 8px!important;font-size:10px!important}
.btn-pri-off{background:#fff;color:#e53e3e;border:1px solid #e53e3e!important;flex:0!important;padding:4px 7px!important;font-size:10px!important}
.ov{position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,.5);z-index:50;display:none;align-items:center;justify-content:center}
.ov .bx{background:#fff;border-radius:12px;padding:20px;width:300px;max-width:90%}
.ov .bx h3{font-size:15px;margin-bottom:12px}
.ov .bx input{width:100%;padding:10px;border:2px solid #3182ce;border-radius:8px;font-size:16px;text-align:center;margin-bottom:12px}
.ov .bx button{padding:8px 20px;border:none;border-radius:6px;font-size:13px;cursor:pointer;margin:0 4px}
.empty{text-align:center;font-size:12px;color:#a0aec0;padding:16px 0}
    #pendingList{max-height:calc(100vh - 200px);overflow-y:auto}
    .est-time{background:#fefcbf;color:#d69e2e;padding:1px 6px;border-radius:3px;font-weight:600;font-size:11px}
    .progress{height:5px;background:#edf2f7;border-radius:3px;margin:6px 0 4px;overflow:hidden}
    .progress .bar{height:100%;border-radius:3px;transition:width .5s;background:linear-gradient(90deg,#d69e2e,#ecc94b)}
    .progress .bar.fast{background:linear-gradient(90deg,#38a169,#68d391)}
	    .progress .bar.slow{background:linear-gradient(90deg,#e53e3e,#fc8181)}
	.filter-bar{display:flex;gap:4px;padding:8px 16px;background:#fff;border-bottom:1px solid #e2e8f0;flex-wrap:wrap}
	.filter-bar button{flex:1;padding:6px;border:1px solid #3182ce;border-radius:6px;background:#fff;color:#3182ce;font-size:11px;cursor:pointer;font-weight:500}
	.filter-bar button.on{background:#e53e3e;border-color:#e53e3e;color:#fff}
	.collapse-col .col-body.collapsed{display:none}
	.collapse-col .cnt.clickable{cursor:pointer;user-select:none;transition:background .2s}
	.collapse-col .cnt.clickable:hover{background:#cbd5e0}
	.pri-cnt{display:none;background:#e53e3e;color:#fff;font-size:9px;padding:1px 5px;border-radius:8px;margin-left:4px}
	.del-bar{display:flex;align-items:center;gap:6px;margin:0 16px 8px;padding:6px 10px;background:#fff;border-radius:6px;font-size:11px;border:1px solid #fed7d7}
	.del-bar label{display:flex;align-items:center;gap:3px;cursor:pointer;color:#4a5568}
	.del-bar label input{cursor:pointer}
	.del-btn{background:#e53e3e;color:#fff;border:none;border-radius:4px;padding:4px 10px;font-size:11px;cursor:pointer}
	.del-btn:hover{background:#c53030}
	.cb-item{width:14px;height:14px;cursor:pointer;margin-right:4px;flex-shrink:0}
	.calc-btn{background:#e53e3e;color:#fff;border:none;border-radius:6px;padding:4px 10px;font-size:11px;cursor:pointer;font-weight:600}
	.calc-btn:hover{background:#c53030}
	#calcModal .bx{width:360px}.calc-row{margin-bottom:10px}.calc-row label{font-size:12px;color:#666;display:block;margin-bottom:3px}
	.calc-row input{width:100%;padding:7px 10px;border:1px solid #ddd;border-radius:6px;font-size:14px;outline:none}
	.calc-row input:focus{border-color:#4361ee;box-shadow:0 0 0 2px rgba(67,97,238,.15)}
	.calc-row .inl{display:flex;gap:10px}.calc-row .inl>div{flex:1}
	.calc-res{padding:12px;border-radius:6px;font-size:13px;line-height:1.8;display:none}
	.calc-res.ok{background:#e8f5e9;border:1px solid #a5d6a7;color:#1b5e20;display:block}
	.calc-res.err{background:#ffebee;border:1px solid #ef9a9a;color:#b71c1c;display:block}</style></head><body>
		<div class="hd"><div class="hd-top">
			<div style="display:flex;align-items:center;gap:10px"><div><h1>🔧 车间看板 <span style="font-size:11px;color:#a0aec0;font-weight:400">v1.2</span></h1><div class="sub" id="batchInfo">加载中...</div></div></div>
							<div class="people-bar"><span style="font-size:11px;color:#a0aec0">v1.2</span></div>
		</div></div>
		<div class="stats"><div class="stat-item gray"><div class="num" id="sTotal">0</div><div class="lbl">全部</div></div><div class="stat-item blue"><div class="num" id="sPending">0</div><div class="lbl">待处理</div></div><div class="stat-item orange"><div class="num" id="sProcessing">0</div><div class="lbl">加工中</div></div><div class="stat-item green"><div class="num" id="sToday">0</div><div class="lbl">今日完成</div></div><div class="stat-item red"><div class="num" id="sPriority">0</div><div class="lbl">⭐优先</div></div></div>
				<div class="filter-bar"><button id="bf_all" class="on" onclick="setFilter('all')">📋 全部</button><button id="bf_pending" onclick="setFilter('pending')">⏸ 待处理</button><button id="bf_processing" onclick="setFilter('processing')">🔧 加工中</button><button id="bf_today" onclick="setFilter('today')">✅ 今日完成</button><button id="bf_history" onclick="setFilter('history')">📋 历史完成</button><button id="bf_priority" onclick="setFilter('priority')">⭐ 优先</button></div>
				<div style="display:flex;gap:6px;padding:6px 16px;background:#fff;border-bottom:1px solid #e2e8f0"><input id="searchBox" type="text" placeholder="🔍 搜索SKU、品名、单号..." oninput="doSearch()" style="width:260px;padding:5px 8px;border:1px solid #e2e8f0;border-radius:6px;font-size:12px;outline:none"><button class="del-btn" onclick="document.getElementById('searchBox').value='';doSearch()" style="padding:4px 10px;font-size:11px">✕ 清除</button></div>
				<div class="del-bar"><label><input type="checkbox" id="selectAll" onchange="toggleAll()"> 全选</label><button class="del-btn" onclick="deleteSelected()">🗑 删除选中</button><button class="del-btn" style="background:#3182ce" onclick="document.getElementById('fuKanban').click()">📤 上传表格</button><span style="display:flex;align-items:center;gap:4px;font-size:12px;color:#4a5568">👥<input id="peopleInput" type="number" min="1" placeholder="上班人数" onchange="savePeople()" style="width:52px;padding:2px 4px;border:1px solid #e2e8f0;border-radius:4px;font-size:12px;text-align:center"><span id="peopleLabel" style="font-size:11px;color:#999"></span></span><input type="file" id="fuKanban" accept=".xlsx" style="display:none" onchange="uploadJobsDirect(this)"><span id="selCount" style="color:#999;margin-left:auto">0项</span></div>
	
	<div class="board"><div class="col"><div class="col-hd"><span>⏸ 待处理</span><span class="cnt" id="cntPending">0</span><span class="pri-cnt" id="pendingPri"></span></div><div id="pendingList"></div></div><div class="col"><div class="col-hd"><span>🔧 加工中</span><span class="cnt" id="cntProcessing">0</span></div><div id="processingList"></div></div><div class="col collapse-col"><div class="col-hd" onclick="toggleCol('todayList')"><span>✅ 今日完成</span><span class="cnt clickable" id="cntToday">0 ▶</span></div><div id="todayList" class="col-body collapsed"></div></div><div class="col collapse-col"><div class="col-hd" onclick="toggleCol('historyList')"><span>📋 历史完成</span><span class="cnt clickable" id="cntHistory">0 ▶</span></div><div id="historyList" class="col-body collapsed"></div></div></div>
<div class="ov" id="modal"><div class="bx"><h3>完成加工</h3><p style="font-size:12px;color:#666;margin-bottom:6px" id="mSku"></p><input id="mQty" type="number" placeholder="输入完成数量" min="1" onkeydown="if(event.key==='Enter')confirmComplete()"><button class="btn-s" onclick="confirmComplete()">确定完成</button><button onclick="closeModal()" style="background:#e2e8f0">取消</button></div></div>
<div class="ov" id="calcModal"><div class="bx" id="calcModalBox"><h3>⚡ 产能计算</h3><p style="font-size:11px;color:#888;margin-bottom:12px">输入SKU、数量和人数，自动估算用时</p>
<div class="calc-row"><label>SKU</label><input id="calcSku" placeholder="例：PW-MSG16-001" onkeydown="if(event.key==='Enter')doCalc()"></div>
<div class="calc-row"><div class="inl"><div><label>数量</label><input id="calcQty" type="number" min="1" placeholder="1000" onkeydown="if(event.key==='Enter')doCalc()"></div><div><label>人数</label><input id="calcPeople" type="number" min="1" value="1" onkeydown="if(event.key==='Enter')doCalc()"></div></div></div>
<button class="btn-s" onclick="doCalc()" style="width:100%;padding:10px;font-size:14px;margin-bottom:10px">▶ 计算用时</button>
<div class="calc-res" id="calcResult"></div>
	<button onclick="document.getElementById('calcModal').style.display='none'" style="width:100%;padding:8px;background:#e2e8f0;border:none;border-radius:6px;font-size:12px;cursor:pointer;margin-top:6px">关闭</button>
	</div></div>
	<div class="ov" id="etaModal"><div class="bx" style="width:380px"><h3>📝 产能录入</h3><p style="font-size:11px;color:#888;margin-bottom:12px">输入实际生产数据，自动计算效率（套/人/小时）</p>
	<div class="calc-row"><label>SKU</label><input id="etaSku" placeholder="例：PW-MSG16-001"></div>
	<div class="calc-row"><div class="inl"><div><label>完成数量</label><input id="etaQty" type="number" min="1" placeholder="300"></div><div><label>人数</label><input id="etaPeople" type="number" min="1" value="1" style="width:70px"></div><div><label>用时（小时）</label><input id="etaHours" type="number" step="0.1" min="0.1" placeholder="4" style="width:80px"></div></div></div>
	<div class="calc-row" style="font-size:12px;color:#666">自动计算效率：<span id="etaCalcPreview">-</span></div>
	<div class="calc-row"><label>备注</label><input id="etaNote" placeholder="如：3人4小时完成300个"></div>
	<button class="btn-s" onclick="saveEta()" style="width:100%;padding:10px;font-size:14px;margin-bottom:8px">💾 保存</button>
	<div class="calc-res" id="etaResult"></div>
	<div style="margin-top:8px"><div style="font-size:12px;color:#666;margin-bottom:4px">已录入的产能：</div><div id="etaList" style="max-height:150px;overflow-y:auto;font-size:12px;line-height:1.8"></div></div>
		<button onclick="document.getElementById('etaModal').style.display='none'" style="width:100%;padding:8px;background:#e2e8f0;border:none;border-radius:6px;font-size:12px;cursor:pointer;margin-top:6px">关闭</button>
		</div></div>
	<div class="ov" id="etaQueryModal"><div class="bx" style="width:420px"><h3>🔍 产能查询</h3>
	<div class="calc-row"><input id="etaQuerySku" type="text" placeholder="输入SKU搜索..." oninput="searchEtaQuery()" style="width:100%;padding:8px 10px;border:1px solid #ddd;border-radius:6px;font-size:14px;outline:none"></div>
	<div style="max-height:300px;overflow-y:auto;border:1px solid #e2e8f0;border-radius:6px;padding:4px 0">
		<div id="etaQueryList" style="font-size:13px;color:#999;text-align:center;padding:20px">输入SKU搜索...</div>
	</div>
	<div style="display:flex;justify-content:space-between;align-items:center;margin-top:10px">
		<button onclick="deleteSelectedEtaQuery()" id="etaQueryDelBtn" style="background:#e53e3e;color:#fff;border:none;border-radius:6px;padding:8px 16px;font-size:13px;cursor:pointer;display:none">🗑 删除选中</button>
		<span id="etaQueryCount" style="font-size:12px;color:#999"></span>
	</div>
	<button onclick="document.getElementById('etaQueryModal').style.display='none'" style="width:100%;padding:8px;background:#e2e8f0;border:none;border-radius:6px;font-size:12px;cursor:pointer;margin-top:8px">关闭</button>
	</div></div>
	<script>
	var curItemId=0, curFilter='all';
	function todayStr(){var d=new Date();return d.getFullYear()+'-'+(d.getMonth()+1).toString().padStart(2,'0')+'-'+d.getDate().toString().padStart(2,'0')}
	function toggleCol(id){if(window.getSelection().toString().length>0)return;var el=document.getElementById(id);if(!el)return;el.classList.toggle('collapsed');var hdr=el.parentElement.querySelector('.cnt');if(hdr)hdr.textContent=hdr.textContent.replace('▶','▼').replace('▼','▶')}
	function doSearch(){
	    var kw=document.getElementById('searchBox').value.trim().toLowerCase();
	    if(!kw || !window.allBoardItems){loadBoard();return}
	    var items=window.allBoardItems.filter(function(i){
	        return (i.sku&&i.sku.toLowerCase().indexOf(kw)>=0)||(i.name&&i.name.toLowerCase().indexOf(kw)>=0)||(i.job_number&&i.job_number.toLowerCase().indexOf(kw)>=0);
	    });
	    renderBoard(items);
	}
		function renderBoard(items){
		    var todayItems=items.filter(function(i){return i.status==='completed' && i.completed && i.completed.substr(0,10)===todayStr()});
		    var historyItems=items.filter(function(i){return i.status==='completed' && (!i.completed || i.completed.substr(0,10)!==todayStr)});
		    if(curFilter==='all' || curFilter==='pending') renderCol('pendingList',items.filter(function(i){return i.status==='pending'}),'pending'); else document.getElementById('pendingList').innerHTML='';
		    if(curFilter==='all' || curFilter==='processing') renderCol('processingList',items.filter(function(i){return i.status==='processing'}),'processing'); else document.getElementById('processingList').innerHTML='';
		    if(curFilter==='all' || curFilter==='today') renderCol('todayList',todayItems,'completed'); else document.getElementById('todayList').innerHTML='';
		    if(curFilter==='all' || curFilter==='history') renderCol('historyList',historyItems,'completed'); else document.getElementById('historyList').innerHTML='';
		    document.getElementById('batchInfo').textContent='全部加工单 ('+items.length+'项)';
		}
async function loadBoard(){
    var ppl=localStorage.getItem('default_people')||'1';
    var r=await fetch('/workshop_data?people='+ppl);var items=await r.json();
    window.allBoardItems=items;
    var total=items.length, pending=items.filter(function(i){return i.status==='pending'}).length;
    var processing=items.filter(function(i){return i.status==='processing'}).length;
    var todayItems=items.filter(function(i){return i.status==='completed' && i.completed && i.completed.substr(0,10)===todayStr()});
    var historyItems=items.filter(function(i){return i.status==='completed' && (!i.completed || i.completed.substr(0,10)!==todayStr)});
    var priority=items.filter(function(i){return i.priority==1}).length;
    document.getElementById('sTotal').textContent=total;document.getElementById('sPending').textContent=pending;
    document.getElementById('sProcessing').textContent=processing;document.getElementById('sToday').textContent=todayItems.length;
    document.getElementById('sPriority').textContent=priority;
    document.getElementById('batchInfo').textContent='全部加工单 ('+total+'项)  |  加工中 '+processing+' 项';
    document.getElementById('cntPending').textContent=pending;document.getElementById('cntProcessing').textContent=processing;
    var pendingPri=items.filter(function(i){return i.status==='pending' && i.priority==1}).length;
    var el=document.getElementById('pendingPri');
    if(el){if(pendingPri>0){el.textContent='⭐'+pendingPri;el.style.display='inline'}else el.style.display='none'}
    document.getElementById('cntToday').textContent=todayItems.length+' ▶';document.getElementById('cntHistory').textContent=historyItems.length+' ▶';
    var kw=document.getElementById('searchBox').value.trim().toLowerCase();
    if(kw){items=items.filter(function(i){return (i.sku&&i.sku.toLowerCase().indexOf(kw)>=0)||(i.name&&i.name.toLowerCase().indexOf(kw)>=0)||(i.job_number&&i.job_number.toLowerCase().indexOf(kw)>=0);});}
    // 保存勾选状态
    var checkedIds={};
    document.querySelectorAll('.cb-item:checked').forEach(function(c){checkedIds[c.getAttribute('data-id')]=true;});
    renderBoard(items);
    // 恢复勾选状态
    document.querySelectorAll('.cb-item').forEach(function(c){if(checkedIds[c.getAttribute('data-id')])c.checked=true;});
    updateSelCount();
    
}
	function setFilter(mode){
    curFilter=mode;
    document.querySelectorAll('.filter-bar button').forEach(function(b){b.className='';});
    var el=document.getElementById('bf_'+mode);
    if(el) el.className='on';
    // 今日/历史完成默认折叠，但点了筛选则展开对应列
    ['todayList','historyList'].forEach(function(id){var el=document.getElementById(id);if(el)el.classList.add('collapsed')});
    if(mode==='today'){var el=document.getElementById('todayList');if(el)el.classList.remove('collapsed')}
    if(mode==='history'){var el=document.getElementById('historyList');if(el)el.classList.remove('collapsed')}
    loadBoard();
}
function renderCol(elId,items,st){
    var el=document.getElementById(elId);el.innerHTML='';
    if(!items.length){el.innerHTML='<div class="empty">暂无</div>';return}
    items.forEach(function(i){
        var c=document.createElement('div');c.className='card '+st;
        var pri=i.priority?'<span class="pri-badge">⭐优先</span>':'';
        var jobNoHtml3 = i.job_number ? '<div style="font-size:10px;color:#718096;padding:0 18px">📋 #'+i.job_number+(i.notes?' <span style="color:#a0aec0">'+escHtml(i.notes)+'</span>':'')+'</div>' : '';
        var h='<div class="top"><span class="sk">'+i.sku+pri+'</span></div>';
        if(i.name) h+='<div class="nm">'+i.name+'</div>';
        h+=jobNoHtml3;
        h+='<div class="qt"><input type="checkbox" class="cb-item" data-id="'+i.id+'" onchange="updateSelCount()"> 数量: '+i.qty+'</div>';
        if(i.est_hours) h+='<div class="qt" style="color:#718096">⏱ 约 '+i.est_hours+'h</div>';
        if(st==='pending'){
            h+='<div class="btns"><button class="btn-s" onclick="startJob('+i.id+')">▶ 开始</button>'+
                (i.priority?'<button class="btn-pri" onclick="togglePri('+i.id+',0)">★优</button>':
                '<button class="btn-pri-off" onclick="togglePri('+i.id+',1)">☆设优</button>')+'</div>';
        }else if(st==='processing'){
            h+='<div class="info"><span>👥 '+(i.worker||'?')+'</span>';
            if(i.started) h+='<span>🕐 '+i.started.substr(11,5)+'</span>';
            if(i.est_time) h+='<span class="est-time">⏱ 预计完成：'+i.est_time+'</span>';
            else if(i.est_min) h+='<span>⏱ 预计完成：~'+i.est_min+'分钟</span>';
            else h+='<span style="color:#a0aec0">⏱ 暂无数据</span>';
            h+='</div>';
            if(i.started && i.est_min>0){
                var stTime=new Date(i.started).getTime();
                var elMin=(Date.now()-stTime)/60000;
                var pct=Math.min(100,Math.round(elMin/i.est_min*100));
                var pClass='';
                if(pct>=85) pClass='fast';
                h+='<div style="font-size:10px;color:#718096;margin:2px 0 1px">进度条</div><div class="progress"><div class="bar '+pClass+'" style="width:'+Math.max(3,pct)+'%"></div></div>';
            }
            h+='<div class="btns"><button class="btn-c" onclick="openComplete(this)" data-id="'+i.id+'" data-sku="'+i.sku+'">✔ 完成</button>'+
                '<button class="btn-x" onclick="cancelJob('+i.id+')">✖ 取消</button></div>';
        }else{
            h+='<div class="info">';
            if(i.worker) h+='<span>👥 '+i.worker+'</span>';
            if(i.started) h+='<span>🕐 '+i.started.substr(11,5)+'</span>';
            h+='<span>✔ 已完成</span>';
            if(i.done_qty) h+='<span>完成: '+i.done_qty+'件</span>';
            if(i.completed) h+='<span>🕐 '+i.completed.substr(11,5)+'</span>';
            h+='</div>';
        }
        c.innerHTML=h;el.appendChild(c);
    });
}
async function startJob(id){
    var ppl=document.getElementById('peopleInput')?document.getElementById('peopleInput').value:'';
    if(!ppl||parseInt(ppl)<1)ppl='1';localStorage.setItem('default_people',ppl);
    var wn='车间工人 x'+ppl;
    var fd=new FormData();fd.append('action','start_job');fd.append('batch_name',id);fd.append('worker',wn);
    try{var r=await fetch('/run',{method:'POST',body:fd});var d=await r.json();if(d.status==='ok'){alert('✅');loadBoard();}else alert('❌ '+d.message);}
    catch(e){alert('❌ '+e.message);}
}
function savePeople(){
    var v=document.getElementById('peopleInput').value;
    if(v&&parseInt(v)>=1){localStorage.setItem('default_people',v);document.getElementById('peopleLabel').textContent='上班: '+v+'人';}
    
    loadBoard();
}
async function uploadJobsDirect(fi){
    if(!fi.files.length) return;
    var name=new Date().toLocaleString('zh-CN',{month:'2-digit',day:'2-digit',hour:'2-digit',minute:'2-digit'}).replace(/[/:]/g,'-').replace(' ','')+'加工单';
    var fd=new FormData();fd.append('file',fi.files[0]);fd.append('action','import_jobs');fd.append('batch_name',name);
    fi.value='';
    try{
        var r=await fetch('/run',{method:'POST',body:fd});var d=await r.json();
        if(d.status==='ok'){alert('✅ '+d.message);loadBoard();}
        else alert('❌ '+d.message);
    }catch(e){alert('❌ '+e.message);}
}
async function cancelJob(id){
    if(!confirm('确定取消此加工任务？'))return;
    var fd=new FormData();fd.append('action','cancel_job');fd.append('batch_name',id);
    try{var r=await fetch('/run',{method:'POST',body:fd});var d=await r.json();if(d.status==='ok'){alert('✅');loadBoard();}else alert('❌ '+d.message);}
    catch(e){alert("❌ "+e.message);}
    }
async function togglePri(id,pri){
    var fd=new FormData();fd.append('action','set_priority');fd.append('batch_name',id);fd.append('priority',pri);
    try{var r=await fetch('/run',{method:'POST',body:fd});var d=await r.json();if(d.status==='ok')loadBoard();else alert(d.message);}
    catch(e){alert(e.message);}
}
function openComplete(btn){curItemId=btn.dataset.id;document.getElementById('mSku').textContent=btn.dataset.sku;document.getElementById('mQty').value='';document.getElementById('modal').style.display='flex';}
function closeModal(){document.getElementById('modal').style.display='none';}
async function confirmComplete(){
    var qty=document.getElementById('mQty').value;
    if(!qty||parseInt(qty)<=0){alert('请输入有效数量');return}
    var fd=new FormData();fd.append('action','complete_job');fd.append('batch_name',curItemId);fd.append('done_qty',qty);
    try{var r=await fetch('/run',{method:'POST',body:fd});var d=await r.json();closeModal();if(d.status==='ok'){alert(d.message);loadBoard();}else alert('❌ '+d.message);}
    catch(e){alert('❌ '+e.message);}
}
function toggleAll(){
    var checked=document.getElementById('selectAll').checked;
    document.querySelectorAll('.cb-item').forEach(function(c){c.checked=checked;});
    updateSelCount();
}
function updateSelCount(){
    var n=document.querySelectorAll('.cb-item:checked').length;
    var el=document.getElementById('selCount');
    if(el) el.textContent=n+'项';
}
async function deleteSelected(){
    var ids=[];
    document.querySelectorAll('.cb-item:checked').forEach(function(c){ids.push(c.getAttribute('data-id'));});
    if(!ids.length){alert('请选择要删除的项');return}
    if(!confirm('确定删除选中的 '+ids.length+' 项？'))return;
    var fd=new FormData();fd.append('action','delete_jobs');fd.append('ids',ids.join(','));
    try{var r=await fetch('/run',{method:'POST',body:fd});var d=await r.json();if(d.status==='ok'){alert('✅ 已删除 '+ids.length+' 项');loadBoard();}else alert('❌ '+d.message);}catch(e){alert('❌ '+e.message);}
}
var sp=localStorage.getItem('default_people');if(sp){document.getElementById('peopleInput').value=sp;document.getElementById('peopleLabel').textContent='上班: '+sp+'人';}
// 产能计算器
function fmtHours(h){var t=Math.round(h*60);var hr=Math.floor(t/60);var mi=t%60;if(hr>0&&mi>0)return hr+'小时'+mi+'分钟';if(hr>0)return hr+'小时';return mi+'分钟'}
function showCalc(){
    document.getElementById('calcModal').style.display='flex';
    document.getElementById('calcResult').className='calc-res';
    document.getElementById('calcResult').style.display='none';
    document.getElementById('calcSku').focus();
}
async function doCalc(){
    var sku=document.getElementById('calcSku').value.trim();
    var qty=parseInt(document.getElementById('calcQty').value);
    var ppl=parseInt(document.getElementById('calcPeople').value)||1;
    var res=document.getElementById('calcResult');
    if(!sku){res.className='calc-res err';res.style.display='block';res.textContent='❌ 请输入SKU';return}
    if(!qty||qty<1){res.className='calc-res err';res.style.display='block';res.textContent='❌ 请输入有效数量';return}
    res.className='calc-res';res.style.display='block';res.textContent='⏳ 计算中...';
    try{
        var r=await fetch('/calc_capacity?sku='+encodeURIComponent(sku)+'&qty='+qty+'&people='+ppl);
        var d=await r.json();
        if(d.status==='ok'){
            var html='✅ <b>预计耗时：'+fmtHours(d.hours)+'</b><br>';
            if(d.end_time) html+='📅 现在开始 → <b>'+d.end_time+' 完成</b><br>';
            html+='📊 数据来源：'+(d.source||'无历史记录');
            if(d.rate) html+='<br>⚡ 效率：'+d.rate+' 套/人/小时';
            res.className='calc-res ok';res.innerHTML=html;
        }else{
            res.className='calc-res err';res.innerHTML='❌ '+(d.message||'计算失败');
        }
    }catch(e){
        res.className='calc-res err';res.innerHTML='❌ 请求失败：'+e.message;
    }
}
// 产能录入
function showEta(){
    document.getElementById('etaModal').style.display='flex';
    document.getElementById('etaResult').className='calc-res';
    document.getElementById('etaResult').style.display='none';
    document.getElementById('etaSku').value='';
    document.getElementById('etaQty').value='';
    document.getElementById('etaPeople').value='1';
    document.getElementById('etaHours').value='';
    document.getElementById('etaNote').value='';
    document.getElementById('etaCalcPreview').textContent='-';
    document.getElementById('etaSku').focus();
    loadEtaList();
}
// 实时预览效率
document.getElementById('etaModal').addEventListener('input', function(e){
    if(e.target.id=='etaQty'||e.target.id=='etaPeople'||e.target.id=='etaHours'){
        var qty=parseFloat(document.getElementById('etaQty').value);
        var ppl=parseFloat(document.getElementById('etaPeople').value);
        var hrs=parseFloat(document.getElementById('etaHours').value);
        var pre=document.getElementById('etaCalcPreview');
        if(qty>0 && ppl>0 && hrs>0){
            pre.textContent=(qty/ppl/hrs).toFixed(2)+' 套/人/小时';
        }else{
            pre.textContent='-';
        }
    }
});
async function saveEta(){
    var sku=document.getElementById('etaSku').value.trim();
    var qty=parseFloat(document.getElementById('etaQty').value);
    var ppl=parseFloat(document.getElementById('etaPeople').value);
    var hrs=parseFloat(document.getElementById('etaHours').value);
    var note=document.getElementById('etaNote').value.trim();
    var res=document.getElementById('etaResult');
    if(!sku){res.className='calc-res err';res.style.display='block';res.textContent='❌ 请输入SKU';return}
    if(!qty||qty<1){res.className='calc-res err';res.style.display='block';res.textContent='❌ 请输入有效完成数量';return}
    if(!ppl||ppl<1){res.className='calc-res err';res.style.display='block';res.textContent='❌ 请输入有效人数';return}
    if(!hrs||hrs<0.1){res.className='calc-res err';res.style.display='block';res.textContent='❌ 请输入有效用时（小时）';return}
    var rate=qty/ppl/hrs;
    if(rate>200){res.className='calc-res err';res.style.display='block';res.textContent='❌ 计算出的效率异常（'+rate.toFixed(1)+'），请检查数据';return}
    res.className='calc-res';res.style.display='block';res.textContent='⏳ 保存中...';
    try{
        var fd=new FormData();
        fd.append('action','save_efficiency');
        fd.append('sku',sku);
        fd.append('rate',rate.toFixed(2));
        fd.append('note',note||(qty+'个/'+ppl+'人/'+hrs+'时'));
        var r=await fetch('/run',{method:'POST',body:fd});
        var d=await r.json();
        if(d.status==='ok'){
            res.className='calc-res ok';res.innerHTML='✅ '+sku+' → <b>'+rate.toFixed(2)+' 套/人/时</b> 已保存';
            document.getElementById('etaSku').value='';
            document.getElementById('etaQty').value='';
            document.getElementById('etaPeople').value='1';
            document.getElementById('etaHours').value='';
            document.getElementById('etaNote').value='';
            document.getElementById('etaCalcPreview').textContent='-';
            loadEtaList();
        }else{
            res.className='calc-res err';res.innerHTML='❌ '+(d.message||'保存失败');
        }
    }catch(e){
        res.className='calc-res err';res.innerHTML='❌ 请求失败：'+e.message;
    }
}
async function loadEtaList(){
    var el=document.getElementById('etaList');
    try{
        var r=await fetch('/get_efficiency');
        var items=await r.json();
        if(!items||items.length===0){el.innerHTML='<span style="color:#999">暂无录入数据</span>';return}
        var html='';
        for(var i=0;i<items.length;i++){
            var it=items[i];
            html+='<div style="display:flex;justify-content:space-between;align-items:center;padding:2px 0">';
            html+='<span><b>'+escHtml(it.sku)+'</b> → '+it.rate+' 套/人/时'+(it.note?' <span style="color:#999">('+escHtml(it.note)+')</span>':'')+'</span>';
            html+='<button class="del-eta-btn" data-sku="'+escHtml(it.sku)+'" style="background:none;border:none;color:#e53e3e;cursor:pointer;font-size:14px;padding:0 4px" title="删除">✕</button>';
            html+='</div>';
        }
        el.innerHTML=html;
    }catch(e){
        el.innerHTML='<span style="color:#e53e3e">加载失败：'+e.message+'</span>';
    }
}
document.getElementById('etaList').addEventListener('click', function(e){
    var btn=e.target.closest('.del-eta-btn');
    if(btn) delEta(btn.dataset.sku);
});
document.getElementById('etaQueryList').addEventListener('click', async function(e){
    var btn=e.target.closest('.del-eta-btn');
    if(btn){ delEtaQueryItem(btn.dataset.sku); }
});

async function delEta(sku){
    if(!confirm('确定删除 '+sku+' 的产能记录？')) return;
    try{
        var fd=new FormData();
        fd.append('action','delete_efficiency');
        fd.append('sku',sku);
        var r=await fetch('/run',{method:'POST',body:fd});
        var d=await r.json();
        if(d.status==='ok'){
            var res=document.getElementById('etaResult');
            res.className='calc-res ok';res.style.display='block';res.textContent='✅ 已删除';
            loadEtaList();
        }else{
            alert('❌ '+(d.message||'删除失败'));
        }
    }catch(e){
        alert('❌ 请求失败：'+e.message);
    }
}
function escHtml(s){if(!s)return '';return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;')}
// 产能查询
var etaQueryData=[];
function showEtaQuery(){
    document.getElementById('etaQueryModal').style.display='flex';
    document.getElementById('etaQuerySku').value='';
    document.getElementById('etaQueryList').innerHTML='<span style="color:#999">输入SKU搜索...</span>';
    document.getElementById('etaQueryDelBtn').style.display='none';
    document.getElementById('etaQueryCount').textContent='';
    etaQueryData=[];
    document.getElementById('etaQuerySku').focus();
}
async function searchEtaQuery(){
    var kw=document.getElementById('etaQuerySku').value.trim().toLowerCase();
    var el=document.getElementById('etaQueryList');
    if(!kw){el.innerHTML='<span style="color:#999">输入SKU搜索...</span>';document.getElementById('etaQueryDelBtn').style.display='none';document.getElementById('etaQueryCount').textContent='';return}
    try{
        // 同时获取手动录入和加工完成数据
        var [manualRes, jobRes]=await Promise.all([
            fetch('/get_efficiency').then(function(r){return r.json()}),
            fetch('/get_job_efficiency').then(function(r){return r.json()})
        ]);
        var allItems=[];
        // 手动录入（带删除功能）
        if(manualRes&&manualRes.length){
            manualRes.forEach(function(it){allItems.push({type:'manual',sku:it.sku,rate:it.rate,note:it.note||'',created:it.created});});
        }
        // 加工完成（只读，不可删除）
        if(jobRes&&jobRes.length){
            jobRes.forEach(function(it){allItems.push({type:'job',sku:it.sku,rate:it.rate,note:it.qty+'个/'+it.people+'人/'+it.hours+'时',created:it.completed});});
        }
        if(allItems.length===0){el.innerHTML='<span style="color:#999">暂无数据</span>';document.getElementById('etaQueryDelBtn').style.display='none';document.getElementById('etaQueryCount').textContent='';return}
        var filtered=allItems.filter(function(it){return it.sku.toLowerCase().indexOf(kw)>=0});
        if(filtered.length===0){el.innerHTML='<span style="color:#999">未找到匹配的SKU</span>';document.getElementById('etaQueryDelBtn').style.display='none';document.getElementById('etaQueryCount').textContent='';return}
        etaQueryData=filtered;
        var html='';
        for(var i=0;i<filtered.length;i++){
            var it=filtered[i];
            var isManual=(it.type==='manual');
            html+='<div style="display:flex;align-items:flex-start;gap:8px;padding:8px 10px;border-bottom:1px solid #f0f0f0">';
            if(isManual){
                html+='<input type="checkbox" class="eq-cb" value="'+escHtml(it.sku)+'" style="margin-top:3px;width:16px;height:16px;cursor:pointer">';
            }else{
                html+='<span style="width:16px;height:16px;margin-top:3px;flex-shrink:0"></span>';
            }
            html+='<div style="flex:1;font-size:13px;line-height:1.6"><b>'+escHtml(it.sku)+'</b> → '+it.rate+' 套/人/时';
            if(it.note) html+='<br><span style="color:#888;font-size:11px">'+(isManual?'📝 ':'✔ ')+escHtml(it.note)+'</span>';
            html+='<br><span style="color:#aaa;font-size:11px">🕐 '+escHtml(it.created)+'</span>';
            if(!isManual) html+=' <span style="color:#38a169;font-size:10px">[加工完成]</span>';
                        html+='<button class="del-eta-btn" data-sku="'+escHtml(it.sku)+'" style="background:none;border:none;color:#e53e3e;cursor:pointer;font-size:14px;padding:0 4px;flex-shrink:0" title="删除">✕</button>';
html+='</div></div>';
        }
        el.innerHTML=html;
        document.getElementById('etaQueryDelBtn').style.display=filtered.some(function(it){return it.type==='manual'})?'inline-block':'none';
        document.getElementById('etaQueryCount').textContent='共 '+filtered.length+' 条'+(filtered.some(function(it){return it.type==='manual'})?'（可勾选删除手动录入）':'');
    }catch(e){
        el.innerHTML='<span style="color:#e53e3e">加载失败：'+e.message+'</span>';
    }
}
async function deleteSelectedEtaQuery(){
    var cbs=document.querySelectorAll('.eq-cb:checked');
    var skus=[];
    cbs.forEach(function(cb){skus.push(cb.value);});
    if(!skus.length){alert('请选择要删除的条目');return}
    if(!confirm('确定删除选中的 '+skus.length+' 条产能记录？')) return;
    var ok=0,fail=0;
    for(var i=0;i<skus.length;i++){
        try{
            var fd=new FormData();
            fd.append('action','delete_efficiency');
            fd.append('sku',skus[i]);
            var r=await fetch('/run',{method:'POST',body:fd});
            var d=await r.json();
            if(d.status==='ok') ok++; else fail++;
        }catch(e){fail++;}
    }
    alert('✅ 已删除 '+ok+' 条'+(fail?'，❌ '+fail+' 条失败':''));
    searchEtaQuery();
}
async function delEtaQueryItem(sku){
    if(!confirm('确定删除 '+sku+' 的产能记录？')) return;
    try{
        var fd=new FormData();
        fd.append('action','delete_efficiency');
        fd.append('sku',sku);
        var r=await fetch('/run',{method:'POST',body:fd});
        var d=await r.json();
        if(d.status==='ok'){
            etaQueryData=etaQueryData.filter(function(it){return it.sku.toUpperCase()!==sku.toUpperCase();});
            renderEtaQueryResults();
        }
    }catch(e){alert('❌ '+e.message);}
}
function renderEtaQueryResults(){
    var el=document.getElementById('etaQueryList');
    var filtered=etaQueryData;
    if(filtered.length===0){el.innerHTML='<span style="color:#999">未找到匹配的SKU</span>';document.getElementById('etaQueryDelBtn').style.display='none';document.getElementById('etaQueryCount').textContent='';return}
    var html='';
    for(var i=0;i<filtered.length;i++){
        var it=filtered[i];
        var isManual=(it.type==='manual');
        html+='<div style="display:flex;align-items:flex-start;gap:8px;padding:8px 10px;border-bottom:1px solid #f0f0f0">';
        if(isManual){
            html+='<input type="checkbox" class="eq-cb" value="'+escHtml(it.sku)+'" style="margin-top:3px;width:16px;height:16px;cursor:pointer">';
        }else{
            html+='<span style="width:16px;height:16px;margin-top:3px;flex-shrink:0"></span>';
        }
        html+='<div style="flex:1;font-size:13px;line-height:1.6"><b>'+escHtml(it.sku)+'</b> → '+it.rate+' 套/人/时';
        if(it.note) html+='<br><span style="color:#888;font-size:11px">'+(isManual?'📝 ':'✔ ')+escHtml(it.note)+'</span>';
        html+='<br><span style="color:#aaa;font-size:11px">🕐 '+escHtml(it.created)+'</span>';
        if(!isManual) html+=' <span style="color:#38a169;font-size:10px">[加工完成]</span>';
        html+='<button class="del-eta-btn" data-sku="'+escHtml(it.sku)+'" style="background:none;border:none;color:#e53e3e;cursor:pointer;font-size:14px;padding:0 4px;flex-shrink:0" title="删除">✕</button>';
        html+='</div></div>';
    }
    el.innerHTML=html;
    document.getElementById('etaQueryDelBtn').style.display=filtered.some(function(it){return it.type==='manual'})?'inline-block':'none';
    document.getElementById('etaQueryCount').textContent='共 '+filtered.length+' 条'+(filtered.some(function(it){return it.type==='manual'})?'（可勾选删除手动录入）':'');
}
// sync disabled
document.addEventListener('visibilitychange', function() { if (!document.hidden) loadBoard(); });
loadBoard();setInterval(loadBoard,10000);
</script></body></html>'''

# ====== 主HTML ======
DIAG_PAGE = '''<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8"><title>POST诊断</title><style>
body{font-family:"Microsoft YaHei",sans-serif;padding:30px;max-width:600px;margin:0 auto}
button{padding:10px 20px;margin:5px;font-size:14px;cursor:pointer}
pre{background:#f5f5f5;padding:10px;border-radius:4px;font-size:12px;max-height:300px;overflow:auto}
input{margin:5px 0}
</style></head><body>
<h2>POST诊断工具</h2>
<p>此页用于独立测试后端POST功能，与主页逻辑完全隔离。</p>

<h3>1. 选文件测试 us</h3>
<input type="file" id="f1"><br>
<button onclick="doPost('us')">POST /run action=us</button>
<pre id="r1">等待点击...</pre>

<h3>2. 无文件POST</h3>
<button onclick="doPostNoFile('us')">POST /run (无文件)</button>
<pre id="r2">等待点击...</pre>

<h3>3. GET测试</h3>
<button onclick="doGet()">GET /health</button>
<pre id="r3">等待点击...</pre>

<script>
async function doPost(action){
    var fi=document.getElementById('f1');
    var r=document.getElementById('r1');
    if(!fi.files[0]){r.textContent='请先选文件';return}
    r.textContent='发送中...';
    try{
        var fd=new FormData();
        fd.append('file',fi.files[0]);
        fd.append('action',action);
        var t0=Date.now();
        var resp=await fetch('/run',{method:'POST',body:fd});
        var t1=Date.now();
        var d=await resp.json();
        r.textContent='OK ('+(t1-t0)+'ms): '+JSON.stringify(d,null,2);
    }catch(e){
        r.textContent='FAIL: '+e.message+'\n'+e.stack;
    }
}
async function doPostNoFile(action){
    var r=document.getElementById('r2');
    r.textContent='发送中...';
    try{
        var fd=new FormData();
        fd.append('action',action);
        var t0=Date.now();
        var resp=await fetch('/run',{method:'POST',body:fd});
        var t1=Date.now();
        var d=await resp.json();
        r.textContent='OK ('+(t1-t0)+'ms): '+JSON.stringify(d,null,2);
    }catch(e){
        r.textContent='FAIL: '+e.message+'\n'+e.stack;
    }
}
async function doGet(){
    var r=document.getElementById('r3');
    r.textContent='发送中...';
    try{
        var t0=Date.now();
        var resp=await fetch('/health');
        var t1=Date.now();
        var d=await resp.json();
        r.textContent='OK ('+(t1-t0)+'ms): '+JSON.stringify(d,null,2);
    }catch(e){
        r.textContent='FAIL: '+e.message+'\n'+e.stack;
    }
}
</script></body></html>
'''

US_PAGE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'us_page.html')
try:
    with open(US_PAGE_FILE, 'r', encoding='utf-8') as f:
        US_PAGE = f.read()
except:
    US_PAGE = '<h2>Error loading page</h2>'

CARDS_PAGE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'cards.html')
try:
    with open(CARDS_PAGE_FILE, 'r', encoding='utf-8') as f:
        CARDS_PAGE = f.read()
except:
    CARDS_PAGE = '<h2>Error loading page</h2>'



PACKING_PAGE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'packing.html')
try:
    with open(PACKING_PAGE_FILE, 'r', encoding='utf-8') as f:
        PACKING_PAGE = f.read()
except:
    PACKING_PAGE = '<h2>Error loading page</h2>'

POST_TEST_PAGE = '''<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8"><title>POST底层测试</title><style>
body{font-family:"Microsoft YaHei",sans-serif;padding:20px;font-size:14px}
button{padding:12px 24px;margin:8px;font-size:16px}
pre{background:#f0f0f0;padding:10px;margin:8px 0}
</style></head><body>
<h2>POST底层连通性测试</h2>

<p>测试1: 纯文本POST（不传文件）</p>
<button onclick="test1()">发送POST /run?q=test</button>
<pre id="r1">-</pre>

<p>测试2: 使用XMLHttpRequest POST</p>
<button onclick="test2()">XHR POST /run</button>
<pre id="r2">-</pre>

<p>测试3: GET (确认连通)</p>
<button onclick="test3()">GET /health</button>
<pre id="r3">-</pre>

<hr>
<p>控制台日志:</p>
<pre id="log" style="max-height:300px;overflow:auto"></pre>

<script>
function clog(msg){var l=document.getElementById('log');l.textContent+=msg+'\n';console.log(msg)}

async function test1(){
    var r=document.getElementById('r1');
    r.textContent='发送中...';
    try{
        var resp=await fetch('/run?q=test',{method:'POST'});
        clog('test1: status='+resp.status);
        var t=await resp.text();
        r.textContent='OK: '+t.substr(0,500);
    }catch(e){
        clog('test1 ERROR: '+e.message+' '+e.stack);
        r.textContent='FAIL: '+e.message;
    }
}
function test2(){
    var r=document.getElementById('r2');
    r.textContent='发送中...';
    try{
        var xhr=new XMLHttpRequest();
        xhr.open('POST','/run',true);
        xhr.onload=function(){
            clog('test2: status='+xhr.status);
            r.textContent='OK: '+xhr.status+' '+xhr.responseText.substr(0,500);
        };
        xhr.onerror=function(){
            clog('test2: network error');
            r.textContent='FAIL: network error';
        };
        xhr.send('hello');
    }catch(e){
        clog('test2 ERROR: '+e.message);
        r.textContent='FAIL: '+e.message;
    }
}
async function test3(){
    var r=document.getElementById('r3');
    r.textContent='发送中...';
    try{
        var resp=await fetch('/health');
        clog('test3: status='+resp.status);
        var t=await resp.text();
        r.textContent='OK: '+t;
    }catch(e){
        clog('test3 ERROR: '+e.message);
        r.textContent='FAIL: '+e.message;
    }
}
clog('页面已加载, 时间: '+new Date().toLocaleTimeString());
clog('当前URL: '+location.href);
</script></body></html>'''

HTML_MAIN_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'main_page.html')
try:
    with open(HTML_MAIN_FILE, 'r', encoding='utf-8') as f:
        HTML_MAIN = f.read()
except:
    HTML_MAIN = '<h2>Error loading main page</h2>'

# ====== HTTP Handler ======
class H(http.server.BaseHTTPRequestHandler):
    def do_GET(self):
        p = self.path
        if p == '/': return self._html(HTML_MAIN)
        if p == '/scan': return self._html(SCAN_PAGE)
        if p == '/scan_admin': return self._html(SCAN_ADMIN)
        if p == '/scan_history_page': return self._html(SCAN_HISTORY)
        if p == '/workshop': return self._html(WORKSHOP_PAGE)
        if p == '/workshop_board': return self._html(WORKSHOP_BOARD)
        if p == '/workshop_admin': return self._html(WORKSHOP_ADMIN)
        if p == '/health': return self._json({'status':'ok','port':PORT})
        if p == '/diag': return self._html(DIAG_PAGE)
        if p == '/us': return self._html(US_PAGE)
        if p == '/packing':
            try:
                with open(PACKING_PAGE_FILE, 'r', encoding='utf-8') as pf:
                    return self._html(pf.read())
            except:
                return self._html(PACKING_PAGE)
        if p == '/cards': return self._html(CARDS_PAGE)
        if p == '/pt': return self._html(POST_TEST_PAGE)
        if p == '/get_ip': return self._json({'ip':get_ip()})
        if p.startswith('/scan_info'):

            q = urllib.parse.parse_qs(urllib.parse.urlparse(p).query)
            bids = q.get('batch',['0'])
            bid = int(bids[0]) if bids[0].isdigit() else None
            if not bid: 
                b = get_latest_batch()
                if b: bid = b[0]
            if bid:
                b = sqlite3.connect(DB_PATH); c = b.cursor()
                c.execute('SELECT name,created_at,items_count,regions FROM batches WHERE id=?', (bid,))
                batch = c.fetchone()
                if batch:
                    regions = [r for r in batch[3].split(',') if r] if batch[3] else []
                    region_stats = {}
                    for rg in regions:
                        region_stats[rg] = get_region_stats(bid, rg)
                    b.close()
                    return self._json({'batch_name':batch[0],'total':batch[2],'regions':regions,'region_stats':region_stats})
                b.close()
            return self._json({'batch_name':'\u6682\u65e0\u6279\u6b21','total':0,'regions':[],'region_stats':{}})
        
        if p.startswith('/scan_check'):
            q = urllib.parse.parse_qs(urllib.parse.urlparse(p).query)
            code = q.get('code',[''])[0].strip(); worker = q.get('worker',[''])[0].strip(); region = q.get('region',[''])[0].strip()
            bid = get_latest_batch()
            if not bid: return self._json({'found':False,'history':[],'message':'暂无批次数据'})
            item = check_document(code, bid[0], region)
            conn = sqlite3.connect(DB_PATH); c = conn.cursor()
            
            if item and item[0] != 'wrong_region':
                # 正确匹配
                record_scan(bid[0], code, worker, 'correct', item[2], region)
                c.execute('SELECT doc_number,worker,result,expected_qty,scanned_at,region,note FROM scans WHERE batch_id=? AND region=? ORDER BY id DESC LIMIT 20', (bid[0], region))
                h = [{'code':h[0],'worker':h[1],'status':'✓' if h[2]=='correct' else '✗','time':h[3],'note':h[6] if h[6] else ''} for h in c.fetchall()]
                conn.close()
                return self._json({
                    'found':True, 'match':True,
                    'doc_number':item[0], 'region':region,
                    'expected_qty':item[2], 'total_weight':item[3],
                    'history':h})
            elif item and item[0] == 'wrong_region':
                # 扫到了其他区域
                wrong_region = item[1]
                wrong_boxes = item[2]
                record_scan(bid[0], code, worker, 'wrong_region', wrong_boxes, region, note='应属'+wrong_region)
                c.execute('SELECT doc_number,worker,result,expected_qty,scanned_at,region,note FROM scans WHERE batch_id=? AND region=? ORDER BY id DESC LIMIT 20', (bid[0], region))
                h = [{'code':h[0],'worker':h[1],'status':'✓' if h[2]=='correct' else '✗','time':h[3],'note':h[6] if h[6] else ''} for h in c.fetchall()]
                conn.close()
                return self._json({
                    'found':True, 'match':False, 'wrong_region':wrong_region,
                    'code12':code[:12].upper(), 'expected_qty':wrong_boxes,
                    'history':h})
            else:
                # 完全没找到
                record_scan(bid[0], code, worker, 'not_found', 0, region, note='未匹配'+code[:12].upper())
                c.execute('SELECT doc_number,worker,result,expected_qty,scanned_at,region,note FROM scans WHERE batch_id=? AND region=? ORDER BY id DESC LIMIT 20', (bid[0], region))
                h = [{'code':h[0],'worker':h[1],'status':'✓' if h[2]=='correct' else '✗','time':h[3],'note':h[6] if h[6] else ''} for h in c.fetchall()]
                conn.close()
                return self._json({'found':False, 'code12':code[:12].upper(), 'history':h})
        
        if p.startswith('/scan_stats'):
            q = urllib.parse.parse_qs(urllib.parse.urlparse(p).query)
            bid = int(q.get('batch',[0])[0])
            region = q.get('region',[''])[0].strip()
            stats = get_scan_stats(bid, region) if bid else {'total':0,'scanned':0,'correct':0,'wrong':0,'docs':[],'scans':[],'unscanned':[]}
            return self._json(stats)
        
        if p.startswith('/scan_batches'):
            return self._json(get_all_batches())
        
        if p == '/workshop': return self._html(WORKSHOP_PAGE)
        if p == '/workshop_board': return self._html(WORKSHOP_BOARD)
        if p == '/workshop_admin': return self._html(WORKSHOP_ADMIN)
        if p.startswith('/workshop_data'):
            q = urllib.parse.parse_qs(urllib.parse.urlparse(p).query)
            bid = q.get('batch',[0])[0]
            default_ppl = int(q.get('people',[0])[0]) or 0
            conn = sqlite3.connect(DB_PATH); c = conn.cursor()
            if bid and bid.isdigit():
                c.execute('SELECT id,sku,product_name,qty,customer,notes,status,worker,started_at,completed_at,completed_qty,priority,job_number FROM job_items WHERE batch_id=? ORDER BY priority DESC, id', (int(bid),))
            else:
                c.execute('SELECT id,sku,product_name,qty,customer,notes,status,worker,started_at,completed_at,completed_qty,priority,job_number FROM job_items ORDER BY priority DESC, id DESC')
            items = []
            for i in c.fetchall():
                item = {'id':i[0],'sku':i[1],'name':i[2],'qty':i[3],'customer':i[4],'notes':i[5],'status':i[6],'worker':i[7] or '','started':i[8] or '','completed':i[9] or '','done_qty':i[10] or 0,'priority':i[11] if i[11] else 0,'job_number':i[12] or ''}
                # 用实际人数或默认人数来估算预计用时
                if item['worker']:
                    actual_worker = item['worker']
                elif default_ppl > 0:
                    actual_worker = '车间工人 x' + str(default_ppl)
                else:
                    actual_worker = 'x1'
                est = calc_est_completion(item['sku'], actual_worker)
                if est and est['rate'] > 0:
                    item['est_hours'] = round(item['qty'] / (est['rate'] * 60 * est['cur_ppl']), 1)
                    if item['status'] == 'processing' and item['started']:
                        remain_min = item['qty'] / (est['rate'] * est['cur_ppl'])
                        if remain_min < 600:
                            try:
                                started = datetime.datetime.fromisoformat(item['started'])
                                est_time = calc_end_time(started, remain_min)
                                item['est_time'] = est_time.strftime('%H:%M')
                                item['est_min'] = round(remain_min)
                            except: pass
                items.append(item)
            conn.close()
            return self._json(items)
        if p.startswith('/calc_capacity'):
            q = urllib.parse.parse_qs(urllib.parse.urlparse(p).query)
            sku = q.get('sku',[''])[0].strip().upper()
            try: req_qty = int(q.get('qty',[0])[0])
            except: req_qty = 0
            try: req_ppl = int(q.get('people',[1])[0])
            except: req_ppl = 1
            if not sku or req_qty <= 0:
                return self._json({'status':'error','message':'参数错误'})
            est = calc_est_completion(sku, 'x'+str(req_ppl))
            if not est or not est.get('rate'):
                return self._json({'status':'error','message':'暂无「'+sku+'」的历史数据，无法估算'})
            rate = est['rate']
            hours = round(req_qty / (rate * 60 * req_ppl), 1)
            now = datetime.datetime.now()
            end = calc_end_time(now, req_qty / (rate * req_ppl))
            if end.date() == now.date():
                end_str = end.strftime('%H:%M')
            else:
                end_str = str(end.month)+'月'+str(end.day)+'日 '+end.strftime('%H:%M')
            return self._json({'status':'ok','hours':hours,'end_time':end_str,'rate':round(rate*60,1),'source':'历史记录'})
        if p.startswith('/get_efficiency'):
            conn = sqlite3.connect(DB_PATH); c = conn.cursor()
            c.execute('SELECT sku, rate, note, created_at FROM efficiency ORDER BY sku')
            rows = c.fetchall(); conn.close()
            return self._json([{'sku':r[0],'rate':r[1],'note':r[2] or '','created':r[3]} for r in rows])
        if p.startswith('/get_job_efficiency'):
            conn = sqlite3.connect(DB_PATH); c = conn.cursor()
            c.execute('SELECT sku, completed_qty, started_at, completed_at, worker FROM job_items WHERE status=\'completed\' AND completed_qty>0 AND started_at IS NOT NULL AND completed_at IS NOT NULL ORDER BY completed_at DESC')
            rows = c.fetchall(); conn.close()
            results = []
            for r in rows:
                sku, qty, started, completed, worker = r
                try:
                    s = datetime.datetime.fromisoformat(started)
                    e = datetime.datetime.fromisoformat(completed)
                    mins = max((e-s).total_seconds()/60, 1)
                    m = re.search(r'x(\d+)', worker or '')
                    ppl = int(m.group(1)) if m else 1
                    rate = round(qty / (mins/60) / ppl, 2)
                    results.append({'sku':sku,'rate':rate,'qty':qty,'people':ppl,'hours':round(mins/60,1),'completed':completed[:16],'worker':worker or ''})
                except: pass
            return self._json(results)
        if p.startswith('/job_batches'):
            conn = sqlite3.connect(DB_PATH); c = conn.cursor()
            c.execute('SELECT id,name,created_at,items_count FROM job_batches ORDER BY id DESC')
            bs = c.fetchall(); conn.close()
            return self._json([{'id':b[0],'name':b[1],'created':b[2],'count':b[3]} for b in bs])
        
        if p.startswith('/scan_history'):
            batches = get_all_batches()
            result = []
            for b in batches:
                summary = get_batch_summary(b[0])
                if summary:
                    result.append({
                        'id':b[0],'name':b[1],'created_at':b[2],
                        'items_count':b[3],'status':b[5],
                        'regions':summary['region_stats'],
                        'total_scans':summary['total_scans'],
                        'correct':summary['correct_scans'],
                        'wrong':summary['wrong_scans'],
                        'not_found':summary['not_found_scans'],
                        'corrections':summary['corrections'],
                        'still_wrong':len(summary['still_wrong'])
                    })
            return self._json(result)
        
        if p.startswith('/track'):
            q = urllib.parse.parse_qs(urllib.parse.urlparse(p).query)
            return self._json(track_logistics(q.get('num',[''])[0].strip()))
        
        if p.startswith('/q'):
            q = urllib.parse.parse_qs(urllib.parse.urlparse(p).query)
            return self._json({'status':'ok','message':answer_query(q.get('q',[''])[0].strip())})
        
        self._json({'status':'error','message':'Not found'})
    
    def do_POST(self):
        try:
            ct = self.headers.get('Content-Type','')
            print(f'[POST] CT={ct[:80]}', flush=True)
            print(f'[POST] CL={self.headers.get("Content-Length","?")}', flush=True)
            b = self.rfile.read(int(self.headers['Content-Length'])); print(f'[POST] read {len(b)} bytes, first 100: {b[:100]}', flush=True)
            boundary = re.search(r'boundary=(?:"([^"]+)"|([^;]+))', ct)
            if not boundary: print('[POST] NO boundary found in CT', flush=True); return self._json({'status':'error','message':'No boundary'})
            bnd = boundary.group(1) or boundary.group(2)
            parts = b.split(('--'+bnd).encode()); print(f'[POST] boundary={bnd}, parts={len(parts)}', flush=True)
            action = ''; fdata = None; fname = None; batch_name = ''
            self.post_data = {}
            for p in parts:
                if not p or p.strip() == b'--' or len(p) < 20: continue
                he = p.find(b'\r\n\r\n')
                if he == -1: continue
                hs = p[:he].decode('utf-8','replace')
                ds = p[he+4:].rstrip(b'\r\n-- ')
                nm = re.search(r'name="([^"]+)"', hs)
                if not nm: continue
                key = nm.group(1)
                val = ds.decode('utf-8','replace').strip()
                self.post_data[key] = val
                if key == 'action': action = val; print(f'[POST] action={action}', flush=True)
                elif key == 'batch_name': batch_name = val
                elif key == 'file':
                    fdata = ds
                    fm = re.search(r'filename="?([^"]+)"?', hs)
                    if fm: fname = fm.group(1).strip()
            # Debug logging
            print(f'[DEBUG] POST action=\"{action}\" fname=\"{fname}\" fdata_size={len(fdata) if fdata else 0} parts={len(parts)}', flush=True)
            # Non-upload actions don't need a file
            if action in ('start_job', 'complete_job', 'set_priority', 'cancel_job', 'delete_jobs', 'pause_job', 'resume_job', 'save_efficiency', 'delete_efficiency'):
                pass  # handle below
            elif not fdata or not fname:
                return self._json({'status':'error','message':'No file'})
            else:
                save_path = os.path.join(UPLOAD_DIR, os.path.basename(fname))
                os.makedirs(os.path.dirname(save_path), exist_ok=True)
                with open(save_path, 'wb') as f: f.write(fdata)
            if action == 'start_job':
                item_id = int(batch_name) if batch_name.isdigit() else 0
                worker = self._get_post('worker', '')
                if item_id:
                    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
                    c.execute('UPDATE job_items SET status=?, worker=?, started_at=? WHERE id=? AND status=\'pending\'', ('processing', worker, datetime.datetime.now().isoformat()[:19], item_id))
                    conn.commit(); conn.close()
                    self._json({'status':'ok','message':'\u2705 \u5df2开始加工'})
                    return
                self._json({'status':'error','message':'\u274c 无效ID'})
                return
            if action == 'complete_job':
                item_id = int(batch_name) if batch_name.isdigit() else 0
                done_qty = self._get_post('done_qty', '0')
                try: done_qty = int(float(done_qty))
                except: done_qty = 0
                if item_id and done_qty > 0:
                    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
                    now = datetime.datetime.now().isoformat()[:19]
                    c.execute("SELECT started_at, qty, sku, product_name, customer, notes, worker, job_number, priority, batch_id FROM job_items WHERE id=? AND status='processing'", (item_id,))
                    row = c.fetchone()
                    if not row:
                        conn.close()
                        self._json({'status':'error','message':'❌ 订单不在加工中'})
                        return
                    started, order_qty, sku, name, customer, notes, worker, job_number, priority, batch_id = row
                    if done_qty >= (order_qty or 0):
                        # 全部完成
                        c.execute('UPDATE job_items SET status=?, completed_at=?, completed_qty=? WHERE id=?', ('completed', now, done_qty, item_id))
                        msg = '✅ 全部完成！共' + str(done_qty) + '件'
                    else:
                        # 部分完成：拆分订单
                        remaining = order_qty - done_qty
                        c.execute('UPDATE job_items SET qty=?, status=?, completed_qty=0 WHERE id=?', (remaining, 'pending', item_id))
                        c.execute('INSERT INTO job_items (sku, product_name, qty, customer, notes, status, worker, started_at, completed_at, completed_qty, priority, job_number, batch_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)',
                            (sku, name, done_qty, customer, notes, 'completed', worker, started, now, done_qty, priority, job_number, batch_id))
                        msg = '✅ 已完成' + str(done_qty) + '件，剩余' + str(remaining) + '件已回到待处理'
                    conn.commit(); conn.close()
                    self._json({'status':'ok','message':msg})
                    return
                self._json({'status':'error','message':'❌ 无效ID或数量'})
                return
            if action == 'cancel_job':
                item_id = int(batch_name) if batch_name.isdigit() else 0
                if item_id:
                    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
                    c.execute("UPDATE job_items SET status='pending', worker='', started_at=NULL, completed_at=NULL, completed_qty=NULL WHERE id=? AND status='processing'", (item_id,))
                    conn.commit(); conn.close()
                    self._json({'status':'ok','message':'\u2705 \u5df2\u53d6\u6d88\u52a0\u5de5\uff0c\u53ef\u91cd\u65b0\u5f00\u59cb'})
                    return
                self._json({'status':'error','message':'\u274c 无效ID'})
                return
            if action == 'delete_jobs':
                ids_str = self._get_post('ids', '')
                if ids_str:
                    ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()]
                    if ids:
                        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
                        placeholders = ','.join(['?'] * len(ids))
                        c.execute('DELETE FROM job_items WHERE id IN (' + placeholders + ')', ids)
                        conn.commit(); conn.close()
                        self._json({'status':'ok','message':'\u2705 \u5df2\u5220\u9664 ' + str(len(ids)) + ' \u9879'})
                        return
                self._json({'status':'error','message':'\u274c \u65e0\u6548\u7684ID\u5217\u8868'})
                return
            if action == 'set_priority':
                iid = int(self.post_data.get('batch_name', 0))
                pri = int(self.post_data.get('priority', 0))
                conn = sqlite3.connect(DB_PATH); c = conn.cursor()
                c.execute('UPDATE job_items SET priority=? WHERE id=?', (pri, iid))
                conn.commit(); conn.close()
                return self._json({'status':'ok'})
            if action == 'pause_job':
                iid = int(self.post_data.get('batch_name', 0))
                if iid:
                    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
                    now = datetime.datetime.now().isoformat()[:19]
                    c.execute("UPDATE job_items SET status='paused', notes=? WHERE id=? AND status='processing'", ('PAUSED:'+now, iid))
                    conn.commit(); conn.close()
                    return self._json({'status':'ok','message':'⏸ 已暂停'})
                return self._json({'status':'error','message':'❌ 无效ID'})
            if action == 'resume_job':
                iid = int(self.post_data.get('batch_name', 0))
                if iid:
                    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
                    r = c.execute('SELECT notes, paused_seconds FROM job_items WHERE id=?', (iid,)).fetchone()
                    if r:
                        notes, paused_sec = r[0] or '', r[1] or 0
                        paused_at = notes.replace('PAUSED:','') if notes.startswith('PAUSED:') else ''
                        if paused_at:
                            try:
                                pt = datetime.datetime.fromisoformat(paused_at)
                                paused_sec += int((datetime.datetime.now() - pt).total_seconds())
                            except: pass
                        c.execute('UPDATE job_items SET status=\'processing\', notes=\'\', paused_seconds=? WHERE id=?', (paused_sec, iid))
                        conn.commit(); conn.close()
                        return self._json({'status':'ok','message':'▶ 已恢复'})
                    conn.close()
                    return self._json({'status':'error','message':'❌ 无效ID'})
            if action == 'complete':
                bid = int(batch_name) if batch_name.isdigit() else 0
                if bid:
                    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
                    c.execute('UPDATE batches SET status=? WHERE id=?', ('complete', bid))
                    conn.commit(); conn.close()
                    self._json({'status':'ok','message':'\u2705 \u5df2\u6807\u8bb0\u4e3a\u5b8c\u6210'})
                    return
                self._json({'status':'error','message':'\u274c \u65e0\u6548\u7684\u6279\u6b21ID'})
                return
            if action == 'import_jobs':
                res = import_jobs(save_path, batch_name)
                self._json({'status':'ok','message':res})
                return
            if action == 'shipment':
                bid, cnt, regions = import_shipment(save_path, batch_name or os.path.basename(fname).replace('.xlsx',''))
                ip = get_ip()
                region_info = '\n\u533a\u57df\uff1a' + ('\u3001'.join(sorted(regions)) if regions else '\u65e0')
                return self._json({'status':'ok','message':'\u2705 \u6279\u6b21\u5df2\u5bfc\u5165 (ID:'+str(bid)+')\n\u540d\u79f0\uff1a'+(batch_name or os.path.basename(fname).replace('.xlsx',''))+'\nSKU\u6761\u7801\uff1a'+str(cnt)+'\u9879'+region_info+'\n\n\U0001f4f1 \u5de5\u4eba\u626b\u7801\uff1ahttp://'+ip+':'+str(PORT)+'/scan\n\U0001f4ca \u7ba1\u7406\u540e\u53f0\uff1ahttp://'+ip+':'+str(PORT)+'/scan_admin'})
            if action == 'save_efficiency':
                sku = self._get_post('sku', '').strip().upper()
                rate = self._get_post('rate', '')
                note = self._get_post('note', '')
                if sku and rate:
                    try:
                        rate = float(rate)
                        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
                        c.execute('INSERT OR REPLACE INTO efficiency (sku, rate, note, created_at) VALUES (?,?,?,?)', (sku, rate, note, datetime.datetime.now().isoformat()[:19]))
                        conn.commit(); conn.close()
                        return self._json({'status':'ok','message':sku+' = '+str(rate)+' 已保存'})
                    except Exception as e:
                        return self._json({'status':'error','message':str(e)})
                return self._json({'status':'error','message':'参数错误'})
            if action == 'delete_efficiency':
                sku = self._get_post('sku', '').strip().upper()
                if sku:
                    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
                    c.execute('DELETE FROM efficiency WHERE sku=?', (sku,))
                    conn.commit(); conn.close()
                    return self._json({'status':'ok','message':'已删除'})
                return self._json({'status':'error','message':'参数错误'})
            print(f'[POST] dispatching action={action}', flush=True)
            func = {'lbl100':run_lbl100,'lbl30':run_lbl30,'us':run_us,'ca':run_ca,'rc':run_rc}.get(action)
            if func:
                print(f'[DEBUG] Calling {action} with save_path={save_path}', flush=True)
                res = func(save_path)
                print(f'[DEBUG] {action} completed successfully', flush=True)
                self._json({'status':'ok','message':res})
            else:
                print(f'[DEBUG] Unknown action: {action}', flush=True)
                self._json({'status':'error','message':'Unknown action: '+action})
        except Exception as e:
            import traceback
            print(f'[DEBUG] Exception in do_POST: {e}', flush=True)
            traceback.print_exc()
            self._json({'status':'error','message':str(e)+'\n'+traceback.format_exc()})
    
    def _get_post(self, key, default=''):
        if hasattr(self, 'post_data') and self.post_data and key in self.post_data:
            return self.post_data[key]
        return default
    
    def _html(self, s): 
        data = s.encode('utf-8')
        self.send_response(200)
        self.send_header('Cache-Control','no-cache, no-store, must-revalidate'); self.send_header('Pragma','no-cache'); self.send_header('Expires','0')
        self.send_header('Content-Type','text/html;charset=utf-8')
        self.send_header('Access-Control-Allow-Origin','*'); self.end_headers()
        self.wfile.write(data)
    def _json(self, d): 
        data = json.dumps(d,ensure_ascii=False).encode('utf-8')
        self.send_response(200); self.send_header('Content-Type','application/json;charset=utf-8')
        self.send_header('Content-Length',str(len(data)))
        self.send_header('Access-Control-Allow-Origin','*'); self.end_headers()
        self.wfile.write(data)
    def log_message(self, fmt, *a): print(fmt%a, flush=True)

# ====== 物流查询 ======
def track_logistics(num):
    if not num: return {'status':'error','message':'\u274c \u8f93\u5165\u5355\u53f7'}
    try:
        url = 'http://www.kuaidi100.com/autonumber/autoComNum?text='+urllib.parse.quote(num)
        req = urllib.request.Request(url, headers={'User-Agent':'Mozilla/5.0'})
        resp = urllib.request.urlopen(req, timeout=10)
        data = json.loads(resp.read().decode('utf-8'))
        autos = data.get('auto', [])
        if not autos:
            rules = {'SF':'shunfeng','YT':'yuantong','ST':'shentong','ZT':'zhongtong','YD':'yunda','JT':'jtexpress','JD':'jd','EMS':'ems'}
            for pfx, com in rules.items():
                if num.upper().startswith(pfx):
                    autos = [{'comCode':com,'name':CARRIER_NAMES.get(com,com)}]; break
            if not autos: return {'status':'error','message':'\u274c \u65e0\u6cd5\u8bc6\u522b\u7269\u6d41\u516c\u53f8'}
        for auto in autos:
            com = auto['comCode']; name = auto.get('name', CARRIER_NAMES.get(com, com))
            try:
                req2 = urllib.request.Request('https://www.kuaidi100.com/query?type='+com+'&postid='+urllib.parse.quote(num), headers={'User-Agent':'Mozilla/5.0'})
                resp2 = urllib.request.urlopen(req2, timeout=15)
                data2 = json.loads(resp2.read().decode('utf-8'))
                if data2.get('status') == '200':
                    traces = data2.get('data', []); state = str(data2.get('state',''))
                    sm = {'0':'\u5df2\u6536\u4ef6','1':'\u8fd0\u8f93\u4e2d','2':'\u6d3e\u4ef6\u4e2d','3':'\u5df2\u7b7e\u6536','4':'\u95ee\u9898\u4ef6','5':'\u9000\u4ef6','6':'\u5f85\u53d6\u4ef6'}
                    msg = '\U0001f4e6 '+name+'\n\u5355\u53f7\uff1a'+num+'\n\u72b6\u6001\uff1a'+sm.get(state,'')+'\n'
                    if traces:
                        msg += '\n\U0001f4cb \u7269\u6d41\u8f68\u8ff9\uff1a'
                        for t in traces[:5]: msg += '\n'+t.get('ftime','')+'  '+t.get('context','')
                    else: msg += '\n\u6682\u65e0\u8bb0\u5f55'
                    return {'status':'ok','message':msg}
            except: continue
        return {'status':'ok','message':'\u8bc6\u522b\u4e3a\uff1a'+autos[0].get('name','')+'\n\u6682\u65f6\u65e0\u6cd5\u83b7\u53d6\u8f68\u8ff9'}
    except Exception as e: return {'status':'error','message':'\u274c \u67e5\u8be2\u5931\u8d25\uff1a'+str(e)}

# ====== 处理函数 ======
def get_date_from_name(name):
    m = re.search(r'(\d{8})', name)
    return m.group(1)[:4]+'-'+m.group(1)[4:6]+'-'+m.group(1)[6:8] if m else None

def extract_name(name):
    parts = name.split('-')
    for i,p in enumerate(parts):
        if any('\u4e00'<=c<='\u9fff' for c in p):
            orig = '-'.join(parts[i:]).replace('.xlsx','')
            m = re.search(r'(\d{8})', orig)
            if m:
                return orig.replace(m.group(1),'').strip('-'), m.group(1)
            return orig, None
    return name.replace('.xlsx',''), None

def run_lbl100(fp):
    import openpyxl; from collections import defaultdict
    wb = openpyxl.load_workbook(fp); ws = wb.active
    bn = os.path.basename(fp); nc,fd = extract_name(bn)
    ff = get_date_from_name(bn) or '0000-00-00'
    boxes = defaultdict(list)
    # Read header row to map column names
    headers = {}
    for c in range(1, ws.max_column+1):
        h = str(ws.cell(1,c).value or '').strip()
        if h: headers[h] = c
    # Find columns by name
    def _col(*names):
        for n in names:
            if n in headers: return headers[n]
        return 0
    col_name = _col('品名')
    col_model = _col('型号','货位码')
    col_pk = _col('包数')
    col_qty = _col('单包数量','单包数')
    col_total = _col('采购量')
    col_box = _col('箱号','装箱序号')
    col_sku = _col('SKU')
    if not col_name or not col_box:
        return '\u274c \u627e\u4e0d\u5230\u201c\u54c1\u540d\u201d\u6216\u201c\u7bb1\u53f7\u201d\u5217\uff0c\u8bf7\u68c0\u67e5Excel\u8868\u5934'
    for r in range(2, ws.max_row+1):
        name = str(ws.cell(r,col_name).value or '').strip() if col_name else ''
        if not name: continue
        box = str(ws.cell(r,col_box).value or '').strip() if col_box else ''
        if not box.isdigit(): continue
        model = str(ws.cell(r,col_model).value or '').strip() if col_model else ''
        pk = str(ws.cell(r,col_pk).value or '') if col_pk else ''
        qty = str(ws.cell(r,col_qty).value or '') if col_qty else ''
        total = str(ws.cell(r,col_total).value or '') if col_total else ''
        sku = str(ws.cell(r,col_sku).value or '').strip() if col_sku else ''
        boxes[int(box)].append((sku,name,model,pk,qty,total))
    tb = len(boxes); ts = sum((len(v)+5)//6 for v in boxes.values())
    sp = [(b,len(boxes[b]),(len(boxes[b])+5)//6) for b in sorted(boxes.keys()) if (len(boxes[b])+5)//6>1]
    lh = ''
    for idx,bn_ in enumerate(sorted(boxes.keys()),1):
        for ps in range(0, len(boxes[bn_]), 6):
            ch = boxes[bn_][ps:ps+6]
            lh += '<div class="l"><div class="bh">(\u7b2c'+str(idx)+'\u7bb1/\u603b'+str(tb)+'\u7bb1)</div><table><tr><th>\u54c1\u540d</th><th>\u578b\u53f7</th><th>\u5305\u6570</th><th>\u5355\u5305\u6570\u91cf</th><th>\u91c7\u8d2d\u91cf</th><th>\u7bb1\u53f7</th></tr>'
            for sku,name,model,pk,qty,total in ch:
                lh += '<tr><td style="font-size:12px">'+name+'</td><td style="font-size:10px">'+model+'</td><td style="font-size:12px;text-align:center">'+pk+'</td><td style="font-size:12px;text-align:center">'+qty+'</td><td style="font-size:12px;text-align:center">'+total+'</td><td style="font-size:12px;text-align:center">'+str(idx)+'</td></tr>'
            lh += '</table></div>'
    s = '<style>@page{size:100mm 100mm;margin:0}body{font-family:"Microsoft YaHei","PingFang SC",sans-serif;margin:0;padding:0}.l{width:100mm;height:100mm;padding:1.5mm;page-break-after:always;overflow:hidden;display:flex;flex-direction:column}.l:last-child{page-break-after:auto}.bh{text-align:center;font-weight:bold;font-size:11px;padding:1mm 0}table{width:100%;border-collapse:collapse;table-layout:fixed}th,td{padding:0.8mm 0.5mm;border:0.7px solid #000;line-height:1.15;word-break:break-all}th{font-size:9px;text-align:center}th:nth-child(1),td:nth-child(1){width:34%}th:nth-child(2),td:nth-child(2){width:17%}th:nth-child(3),td:nth-child(3){width:10%}th:nth-child(4),td:nth-child(4){width:12%}th:nth-child(5),td:nth-child(5){width:12%}th:nth-child(6),td:nth-child(6){width:15%}.np{text-align:center;padding:8px;background:#fff3cd;border-bottom:2px solid #ffc107}@media print{.np{display:none}}</style>'
    html = '<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8"><title>'+nc+'\u6807\u7b7e</title>'+s+'</head><body><div class="np"><strong>'+nc+'\u6807\u7b7e</strong> | 10x10cm | '+str(ts)+'\u5f20 | <button onclick="window.print()" style="font-size:15px;padding:5px 18px">\u6253\u5370</button></div>'+lh+'</body></html>'
    fout = os.path.join(DESKTOP, datetime.datetime.now().strftime('%m-%d').lstrip('0').replace('-0','-')+'-\u6807\u7b7e-'+nc+'.html')
    with open(fout,'w',encoding='utf-8') as f: f.write(html)
    msg = '\u2705 '+nc+' \u5171 '+str(ts)+' \u5f20\n\u6587\u4ef6\uff1a'+fout
    if sp: msg += '\n\u62c6\u7bb1\uff1a'+', '.join(['\u7bb1'+str(b)+'('+str(c)+'\u6b3e\u2192'+str(p)+'\u5f20)' for b,c,p in sp])
    return msg

def run_lbl30(fp):
    import openpyxl
    wb = openpyxl.load_workbook(fp); ws = wb.active
    labels = []
    for r in range(2, ws.max_row+1):
        if not str(ws.cell(r,1).value or '').strip(): continue
        labels.append([str(ws.cell(r,c).value or '').strip() for c in range(1,12)])
    num = len(labels); pages = (num+1)//2
    lbs = ''
    for pi in range(0, num, 2):
        lbs += '<div class="p">'
        for j in range(2):
            i = pi+j
            if i < num:
                l = labels[i]
                lbs += '<div class="lb"><div class="rg">'+l[0]+'</div><div class="bk">'+l[1]+'</div><div class="dt">'+l[2]+'\u7bb1--'+l[3]+'kg--'+l[4]+'m3</div><div class="dt tm">'+l[5][:10]+'--'+l[6][:5]+'</div><div class="dt">'+l[7]+'--'+l[8]+'--'+l[9]+'</div><div class="ph">'+l[10]+'</div></div>'
        lbs += '</div>'
    s = '<style>@page{size:100mm 30mm;margin:0}body{font-family:"Microsoft YaHei","PingFang SC",sans-serif;margin:0;padding:0}.p{width:100mm;height:30mm;display:flex;page-break-after:always}.p:last-child{page-break-after:auto}.lb{width:50mm;height:30mm;padding-left:3mm;padding-top:2.5mm;overflow:hidden;line-height:1.15}.rg{font-size:14px;font-weight:bold}.bk{font-size:12px}.dt{font-size:12px}.tm{font-weight:bold}.ph{font-size:12px}.np{text-align:center;padding:8px;background:#fff3cd;border-bottom:2px solid #ffc107}@media print{.np{display:none}}</style>'
    html = '<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8"><title>\u6807\u7b7e-\u63d0\u8d27\u4fe1\u606f</title>'+s+'</head><body><div class="np"><strong>\u6807\u7b7e-\u63d0\u8d27\u4fe1\u606f</strong> | 100x30mm | '+str(num)+'\u5f20 | 2\u5217/\u9875 | <button onclick="window.print()" style="font-size:15px;padding:5px 18px">\u6253\u5370</button></div>'+lbs+'</body></html>'
    fout = os.path.join(DESKTOP, datetime.datetime.now().strftime('%m-%d').lstrip('0').replace('-0','-')+'-\u6807\u7b7e-\u63d0\u8d27\u4fe1\u606f.html')
    with open(fout,'w',encoding='utf-8') as f: f.write(html)
    return '\u2705 '+str(num)+'\u5f20\u6807\u7b7e\uff0c'+str(pages)+'\u9875\n\u6587\u4ef6\uff1a'+fout

def run_us(fp):
    import openpyxl; from collections import OrderedDict,defaultdict
    from openpyxl.styles import Font,PatternFill,Border,Side,Alignment; from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook(fp); ws1=wb[wb.sheetnames[0]]
    try: ws5=wb[wb.sheetnames[4]]
    except: ws5=wb.active
    d=OrderedDict()
    for r in range(2,ws1.max_row+1):
        doc=str(ws1.cell(r,1).value or '').strip()
        if not doc or doc in d: continue
        d[doc]=[str(ws1.cell(r,c).value or '').strip() for c in [2,3,4,5,7,8]]
    pk={}
    for r in range(2,ws5.max_row+1):
        doc=str(ws5.cell(r,1).value or '').strip()
        if not doc: continue
        k=doc+'|'+str(ws5.cell(r,5).value or '')
        if k not in pk: pk[k]={'b':0,'w':0.0,'v':0.0}
        try: pk[k]['b']+=int(ws5.cell(r,2).value or 0)
        except: pass
        try: pk[k]['w']+=float(str(ws5.cell(r,3).value or '0').replace(',','').strip())
        except: pass
        try: pk[k]['v']+=float(str(ws5.cell(r,4).value or '0').replace(',','').strip())
        except: pass
    out=openpyxl.Workbook(); ows=out.active; ows.title='\u53d1\u8d27\u6c47\u603b'
    h=['\u53d1\u8d27\u5355\u53f7','\u7269\u6d41\u4e2d\u5fc3\u7f16\u7801','\u7269\u6d41\u5546','\u7269\u6d41\u6e20\u9053','\u7269\u6d41\u5546\u5355\u53f7','\u56fd\u5bb6','\u603b\u7bb1\u6570','\u603b\u91cd\u91cf(kg)','\u603b\u4f53\u79ef(m3)','\u8d27\u4ef6\u5355\u53f7']
    tn=Side(style='thin',color='000000'); bd=Border(left=tn,right=tn,top=tn,bottom=tn)
    hf=Font(bold=True,size=11,color='FFFFFF'); hf2=PatternFill(start_color='4472C4',end_color='4472C4',fill_type='solid')
    nf=Font(size=11); bf=Font(bold=True,size=12); sfl=PatternFill(start_color='D9E2F3',end_color='D9E2F3',fill_type='solid')
    for i,hh in enumerate(h,1): c=ows.cell(1,i,hh); c.font=hf; c.fill=hf2; c.alignment=Alignment(horizontal='center',vertical='center'); c.border=bd
    ro=['\u4e1c\u5317\u90e8','\u4e1c\u5357\u90e8','\u897f\u90e8','\u4e2d\u5317\u90e8','\u4e2d\u5357\u90e8']
    rd=defaultdict(list)
    for doc,v in d.items():
        ch=v[2]; rg='\u5176\u4ed6'
        for r2 in ro:
            if ch.startswith(r2): rg=r2; break
        b=0; w=0.0; vol=0.0
        for k,pv in pk.items():
            if k.startswith(doc): b+=pv['b']; w+=pv['w']; vol+=pv['v']
        rd[rg].append((doc,v[0],v[1],ch,v[3],v[4],b,round(w,2),round(vol,2),v[5]))
    row=2; gb=0; gw=0.0; gv=0.0; rg_stats={}
    for rg in ro:
        items=rd.get(rg,[])
        if not items: continue
        for it in items:
            for i,v in enumerate(it,1): c=ows.cell(row,i,v); c.font=nf; c.border=bd; c.alignment=Alignment(vertical='center')
            row+=1
        sb=sum(i[6] for i in items); sw=sum(i[7] for i in items); sv=sum(i[8] for i in items)
        ows.cell(row,1,rg+' \u5c0f\u8ba1 '+str(len(items))+'\u5355').font=Font(size=11); ows.cell(row,7,sb).font=Font(size=11); ows.cell(row,8,round(sw,2)).font=Font(size=11); ows.cell(row,9,round(sv,2)).font=Font(size=11)
        for i in range(1,11): ows.cell(row,i).border=bd; ows.cell(row,i).fill=sfl; ows.cell(row,i).alignment=Alignment(horizontal='center',vertical='center')
        row+=1; gb+=sb; gw+=sw; gv+=sv; rg_stats[rg]=(len(items),sb,round(sw,2),round(sv,2))
    ows.cell(row,1,'\u5408\u8ba1 '+str(len(d))+'\u5355').font=bf; ows.cell(row,7,gb).font=bf; ows.cell(row,8,round(gw,2)).font=bf; ows.cell(row,9,round(gv,2)).font=bf
    for i in range(1,11): ows.cell(row,i).border=bd; ows.cell(row,i).alignment=Alignment(horizontal='center',vertical='center')
    for i,w in enumerate([18,12,10,22,22,8,10,14,14,20],1): ows.column_dimensions[get_column_letter(i)].width=w
    fn=re.sub(r'[-]*\d+(?:[-]*\d+)*$','',os.path.basename(fp).replace('.xlsx','')).rstrip('-').replace('\ufffd','')
    fout=os.path.join(DESKTOP,fn+'\u6c47\u603b.xlsx')
    out.save(fout)
    return '\u2705 \u7f8e\u56fd\u53d1\u8d27\u6c47\u603b\n\u5408\u8ba1\uff1a'+str(len(d))+'\u5355\uff0c'+str(gb)+'\u7bb1\uff0c'+str(round(gw,2))+'kg\uff0c'+str(round(gv,2))+'m3\n'+'\n'.join([rg+': '+str(rg_stats[rg][0])+'\u5355\uff0c'+str(rg_stats[rg][1])+'\u7bb1\uff0c'+str(rg_stats[rg][2])+'kg\uff0c'+str(rg_stats[rg][3])+'m\u00b3' for rg in ro if rg in rg_stats if rd[rg]])+'\n\u6587\u4ef6\uff1a'+fout

def run_ca(fp):
    import openpyxl; from collections import OrderedDict
    from openpyxl.styles import Font,PatternFill,Border,Side,Alignment; from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook(fp); ws1=wb[wb.sheetnames[0]]
    try: ws5=wb[wb.sheetnames[4]]
    except: ws5=wb.active
    d=OrderedDict()
    for r in range(2,ws1.max_row+1):
        doc=str(ws1.cell(r,1).value or '').strip()
        if not doc or doc in d: continue
        d[doc]=[str(ws1.cell(r,c).value or '').strip() for c in [2,3,7,8]]
    pk={}
    for r in range(2,ws5.max_row+1):
        doc=str(ws5.cell(r,1).value or '').strip()
        if not doc: continue
        k=doc+'|'+str(ws5.cell(r,5).value or '')
        if k not in pk: pk[k]={'b':0,'w':0.0,'v':0.0}
        try: pk[k]['b']+=int(ws5.cell(r,2).value or 0)
        except: pass
        try: pk[k]['w']+=float(str(ws5.cell(r,3).value or '0').replace(',','').strip())
        except: pass
        try: pk[k]['v']+=float(str(ws5.cell(r,4).value or '0').replace(',','').strip())
        except: pass
    out=openpyxl.Workbook(); ows=out.active; ows.title='\u52a0\u62ff\u5927\u53d1\u8d27\u6c47\u603b'
    h=['\u53d1\u8d27\u5355\u53f7','\u7269\u6d41\u4e2d\u5fc3\u7f16\u7801','\u7269\u6d41\u5546','\u56fd\u5bb6','\u603b\u7bb1\u6570','\u603b\u91cd\u91cf(kg)','\u603b\u4f53\u79ef(m3)','\u8d27\u4ef6\u5355\u53f7']
    tn=Side(style='thin',color='000000'); bd=Border(left=tn,right=tn,top=tn,bottom=tn)
    hf=Font(bold=True,size=11,color='FFFFFF'); hf2=PatternFill(start_color='4472C4',end_color='4472C4',fill_type='solid')
    nf=Font(size=11); bf=Font(bold=True,size=12)
    for i,hh in enumerate(h,1): c=ows.cell(1,i,hh); c.font=hf; c.fill=hf2; c.alignment=Alignment(horizontal='center',vertical='center'); c.border=bd
    row=2; gb=0; gw=0.0; gv=0.0
    for doc,v in d.items():
        b=0; w=0.0; vol=0.0
        for k,pv in pk.items():
            if k.startswith(doc): b+=pv['b']; w+=pv['w']; vol+=pv['v']
        vs=[doc,v[0],v[1],v[2],b,round(w,2),round(vol,2),v[3]]
        for i,val in enumerate(vs,1): c=ows.cell(row,i,val); c.font=nf; c.border=bd; c.alignment=Alignment(vertical='center')
        row+=1; gb+=b; gw+=w; gv+=vol
    ows.cell(row,1,'\u5408\u8ba1 '+str(len(d))+'\u5355').font=bf; ows.cell(row,5,gb).font=bf; ows.cell(row,6,round(gw,2)).font=bf; ows.cell(row,7,round(gv,2)).font=bf
    for i in range(1,9): ows.cell(row,i).border=bd; ows.cell(row,i).alignment=Alignment(horizontal='center',vertical='center')
    for i,w in enumerate([18,14,14,10,10,14,14,20],1): ows.column_dimensions[get_column_letter(i)].width=w
    fn=os.path.basename(fp).replace('.xlsx','');idx=fn.find('\u53d1\u8d27\u5355');fn=(fn[:idx+3] if idx>=0 else fn).rstrip('-')
    fout=os.path.join(DESKTOP,fn+'\u6c47\u603b.xlsx')
    out.save(fout)
    return '\u2705 \u52a0\u62ff\u5927\u53d1\u8d27\u6c47\u603b\n\u5408\u8ba1\uff1a'+str(len(d))+'\u5355\uff0c'+str(gb)+'\u7bb1\uff0c'+str(round(gw,2))+'kg\uff0c'+str(round(gv,2))+'m3\n\u6587\u4ef6\uff1a'+fout

def run_rc(fp):
    import openpyxl; from collections import defaultdict
    wb=openpyxl.load_workbook(fp,data_only=True)
    # 自动选择工作表
    if 'Sheet2' in wb.sheetnames:
        ws = wb['Sheet2']
    elif 'sheet1' in wb.sheetnames:
        ws = wb['sheet1']
    else:
        ws = wb[wb.sheetnames[0]]
    left={}
    for r in range(3,ws.max_row+1):
        sku=str(ws.cell(r,1).value or '').strip(); qty=ws.cell(r,3).value
        if sku and qty is not None:
            try: left[sku]=left.get(sku,0)+int(float(str(qty)))
            except: pass
    right={}
    for r in range(3,ws.max_row+1):
        sku=str(ws.cell(r,5).value or '').strip(); qty=ws.cell(r,7).value
        if sku and qty is not None:
            try: right[sku]=right.get(sku,0)+int(float(str(qty)))
            except: pass
    names={}
    for r in range(3,ws.max_row+1):
        for sc,nc in [(1,2),(5,6)]:
            sku=str(ws.cell(r,sc).value or '').strip(); nm=str(ws.cell(r,nc).value or '').strip()
            if sku: names[sku]=nm
    all_s=set(list(left.keys())+list(right.keys()))
    m=sum(1 for s in all_s if left.get(s,0)==right.get(s,0))
    dif=[(s,names.get(s,''),left.get(s,0),right.get(s,0)) for s in sorted(all_s) if left.get(s,0)!=right.get(s,0)]
    msg='\u2705 \u6838\u5bf9\u5b8c\u6210\n\u603bSKU: '+str(len(all_s))+' | \u4e00\u81f4: '+str(m)+' | \u5dee\u5f02: '+str(len(dif))
    if dif:
        msg+='\n\n\u5dee\u5f02\u660e\u7ec6\uff1a'
        for s,nm,l,r in dif:
            msg+='\n'+s+' | '+nm[:20]+'\n  \u901a\u77e5 '+format(l,',')+' \u2192 \u5230\u8d27 '+format(r,',')+' (\u5dee\u989d '+format(l-r,',')+')'
    return msg

# ====== 数据问答 ======
def calc_work_minutes(start_dt, end_dt):
    """计算两个时间之间的实际工作分钟数（排除午休12:00-13:00和下班时段）"""
    if start_dt >= end_dt: return 0
    total = 0
    cur = start_dt
    while cur.date() <= end_dt.date():
        day_start = datetime.datetime.combine(cur.date(), datetime.time(8, 30))
        lunch_s = datetime.datetime.combine(cur.date(), datetime.time(12, 0))
        lunch_e = datetime.datetime.combine(cur.date(), datetime.time(13, 0))
        day_end = datetime.datetime.combine(cur.date(), datetime.time(17, 30))
        
        if cur.date() == start_dt.date():
            seg_start = max(cur, day_start)
        else:
            seg_start = day_start
        if cur.date() == end_dt.date():
            seg_end = min(end_dt, day_end)
        else:
            seg_end = day_end
        
        if seg_start < seg_end:
            # 减去午休重叠部分
            if seg_start < lunch_e and seg_end > lunch_s:
                if seg_start >= lunch_s:
                    pass  # 全程在午休内，不计
                elif seg_end <= lunch_e:
                    pass  # 全程在午休内，不计
                else:
                    total += (lunch_s - seg_start).total_seconds() / 60
                    total += (seg_end - lunch_e).total_seconds() / 60
            else:
                total += (seg_end - seg_start).total_seconds() / 60
        
        cur += datetime.timedelta(days=1)
        cur = datetime.datetime.combine(cur.date(), datetime.time(8, 30))
    return max(0, total)

def get_workshop_efficiency(sku):
    try:
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        sku_upper = sku.upper()
        c.execute('SELECT completed_qty, completed_at, started_at FROM job_items WHERE sku=? AND status=\'completed\' AND completed_qty>0', (sku_upper,))
        rows = c.fetchall()
        conn.close()
        if not rows: return None
        total_qty = sum(r[0] for r in rows)
        total_minutes = 0
        for r in rows:
            try:
                s = datetime.datetime.fromisoformat(r[2])
                e = datetime.datetime.fromisoformat(r[1])
                total_minutes += calc_work_minutes(s, e)
            except: pass
        if total_minutes <= 0: return None
        total_hours = total_minutes / 60
        rate = total_qty / total_hours if total_hours > 0 else None
        return {'qty': total_qty, 'hours': round(total_hours, 2), 'rate': round(rate, 2) if rate else None, 'count': len(rows)}
    except: return None

def calc_est_completion(sku, current_worker):
    """根据历史完成记录或预设效率，估算当前加工任务的预计完成时间"""
    try:
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        sku_upper = sku.upper()
        # 先查手动录入的产能
        c.execute('SELECT rate FROM efficiency WHERE sku=?', (sku_upper,))
        m = c.fetchone()
        if m:
            conn.close()
            cur_ppl = int(re.search(r'x(\d+)', current_worker or '').group(1)) if re.search(r'x(\d+)', current_worker or '') else 1
            return {'rate': round(m[0] / 60, 4), 'cur_ppl': cur_ppl}
        # 取所有完成记录（含人数和暂停时间）
        c.execute('SELECT completed_qty, completed_at, started_at, worker, IFNULL(paused_seconds,0) FROM job_items WHERE sku=? AND status=\'completed\' AND completed_qty>0 AND started_at IS NOT NULL AND completed_at IS NOT NULL ORDER BY completed_at DESC', (sku_upper,))
        rows = c.fetchall()
        conn.close()
        # 当前人数
        m2 = re.search(r'x(\d+)', current_worker or '')
        cur_ppl = int(m2.group(1)) if m2 else 1
        
        if rows:
            # 有历史记录：取平均人均每分钟效率
            rates = []
            for prev_qty, prev_end, prev_start, prev_worker, paused_sec in rows:
                s = datetime.datetime.fromisoformat(prev_start)
                e = datetime.datetime.fromisoformat(prev_end)
                dur_min = max(calc_work_minutes(s, e) - (paused_sec or 0) / 60, 1)  # 扣除暂停时间
                m = re.search(r'x(\d+)', prev_worker or '')
                prev_ppl = int(m.group(1)) if m else 1
                rates.append(prev_qty / dur_min / prev_ppl)
            rates.sort()
            if len(rates) >= 3:
                rates = rates[1:-1]  # 去掉最高最低
            rate_per_min = sum(rates) / len(rates)
        else:
            # 无历史记录：使用预设效率（套/人/小时 → 套/人/分钟）
            preset = {'pw-msg16-001':21.35,'pw-msg12-001':16.31,'cb-cmsg-01a':29.20,'cb-msg08-a01':13.50,
                'hw-pnp36-001':27.88,'gn-jp500-002':21.59,'gn-pq115-001':21.57,'gn-jp750-001':21.43,
                'gn-fqj15-001':15.00,'hs-msg32-001':15.56,'co-msg08-a01':10.50,'sp-cms04-002':9.60,'sp-msg04-002':9.60}.get(sku_upper.lower())
            if not preset: return None
            rate_per_min = preset / 60  # 转换为每分钟效率
        
        return {'rate': round(rate_per_min, 4), 'cur_ppl': cur_ppl}
    except: return None

def calc_end_time(start_dt, work_minutes):
    """从开始时间算起，经过 work_minutes 分钟工作时间后的实际结束时间（考虑上班时段）"""
    cur = start_dt
    remaining = work_minutes
    while remaining > 0:
        day_start = datetime.datetime.combine(cur.date(), datetime.time(8, 30))
        lunch_s = datetime.datetime.combine(cur.date(), datetime.time(12, 0))
        lunch_e = datetime.datetime.combine(cur.date(), datetime.time(13, 0))
        day_end = datetime.datetime.combine(cur.date(), datetime.time(17, 30))
        
        if cur < day_start:
            cur = day_start
            continue
        if lunch_s <= cur < lunch_e:
            cur = lunch_e
            continue
        if cur >= day_end:
            cur = datetime.datetime.combine(cur.date(), datetime.time(8, 30)) + datetime.timedelta(days=1)
            continue
        
        if cur < lunch_s:
            avail = (lunch_s - cur).total_seconds() / 60
        else:
            avail = (day_end - cur).total_seconds() / 60
        
        if avail >= remaining:
            cur += datetime.timedelta(minutes=remaining)
            remaining = 0
        else:
            remaining -= avail
            cur = lunch_e if cur < lunch_s else datetime.datetime.combine(cur.date(), datetime.time(8, 30)) + datetime.timedelta(days=1)
    
    return cur

def get_workshop_sku_list():
    try:
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute('SELECT DISTINCT sku FROM job_items ORDER BY sku')
        skus = [r[0] for r in c.fetchall()]
        conn.close()
        return skus
    except: return []

def answer_query(query):
    q = query.lower(); import re
    sku_match = re.search(r'([a-z]{2}[-_][a-z0-9][a-z0-9-]*)', q)
    people_match = re.search(r'(\d+)\u4eba', q); qty_match = re.search(r'(\d+)\u5957', q)
    vol_match = re.search(r'(\d+)\u65b9', q)
    calc_kw = ['\u591a\u4e45','\u65f6\u95f4','\u5c0f\u65f6','\u505a\u5b8c']
    is_calc = any(k in q for k in calc_kw)
    sku_rates = {'pw-msg16-001':21.35,'pw-msg12-001':16.31,'cb-cmsg-01a':29.20,'cb-msg08-a01':13.50,
        'hw-pnp36-001':27.88,'gn-jp500-002':21.59,'gn-pq115-001':21.57,'gn-jp750-001':21.43,'gn-fqj15-001':15.00,
        'hs-msg32-001':15.56,'co-msg08-a01':10.50,'sp-cms04-002':9.60,'sp-msg04-002':9.60}
    if sku_match and is_calc and (people_match or qty_match):
        sku = sku_match.group(1).replace('_','-').lower()
        people = int(people_match.group(1)) if people_match else 1
        qty = int(qty_match.group(1)) if qty_match else 0
        rate = sku_rates.get(sku)
        ws = get_workshop_efficiency(sku)
        if ws and ws['rate']: rate = ws['rate']
        if rate and qty:
            hours = qty / (people * rate); h,mm = int(hours), int((hours-int(hours))*60)
            src = '(\u52a0\u5de5\u5355\u5b9e\u6d4b)' if ws else '(\u9884\u8bbe\u503c)'
            return '\U0001f4c8 \u3010'+sku.upper()+'\u3011\u6548\u7387\uff1a'+str(rate)+' \u5957/\u4eba/\u5c0f\u65f6 '+src+'\n'+str(people)+'\u4eba\u00d7'+str(rate)+' = '+str(round(people*rate,1))+' \u5957/\u5c0f\u65f6\n'+str(qty)+'\u5957 \u00f7 '+str(round(people*rate,1))+' = '+str(round(hours,2))+' \u5c0f\u65f6 = '+str(h)+'\u5c0f\u65f6'+str(mm)+'\u5206\u949f'
        if rate:
            src = '(\u52a0\u5de5\u5355\u5b9e\u6d4b)' if ws else '(\u9884\u8bbe\u503c)'
            return '\U0001f4c8 \u3010'+sku.upper()+'\u3011\u6548\u7387\uff1a'+str(rate)+' \u5957/\u4eba/\u5c0f\u65f6 '+src
        ws2 = get_workshop_efficiency(sku)
        if ws2:
            ans = '\U0001f4c8 \u3010'+sku.upper()+'\u3011\u52a0\u5de5\u5355\u6570\u636e\n\u7d2f\u8ba1: '+str(ws2['qty'])+'\u4ef6, '+str(ws2['hours'])+'\u5c0f\u65f6, '+str(ws2['count'])+'\u6b21'
            if ws2['rate']: ans += '\n\u6548\u7387: '+str(ws2['rate'])+' \u5957/\u4eba/\u5c0f\u65f6'
            return ans
        return '\u26a0\ufe0f \u672a\u77e5SKU\uff1a'+sku+'\\n\u5df2\u77e5\uff1aPW-MSG16-001=21.35'
    if is_calc and people_match and vol_match:
        people = int(people_match.group(1)); vol = int(vol_match.group(1))
        hours = vol / (people * 0.77); h,mm = int(hours), int((hours-int(hours))*60)
        return '\U0001f4c8 \u6253\u5305\u6548\u7387\uff1a0.77 \u65b9/\u4eba/\u5c0f\u65f6\n'+str(people)+'\u4eba\u00d70.77='+str(round(people*0.77,1))+' \u65b9/\u5c0f\u65f6\n'+str(vol)+'\u65b9\u00f7'+str(round(people*0.77,1))+'='+str(round(hours,2))+' \u5c0f\u65f6='+str(h)+'\u5c0f\u65f6'+str(mm)+'\u5206\u949f'
    if '\u52a0\u5de5' in q or '\u8f66\u95f4' in q or '\u751f\u4ea7' in q:
        ws_skus = get_workshop_sku_list()
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute("SELECT COUNT(*), COALESCE(SUM(completed_qty),0) FROM job_items WHERE status='completed'")
        done = c.fetchone()
        c.execute("SELECT COUNT(*) FROM job_items WHERE status='processing'")
        proc = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM job_items WHERE status='pending'")
        pend = c.fetchone()[0]
        conn.close()
        msg = '\U0001f527 \u52a0\u5de5\u5355\u7edf\u8ba1\n\n\u5f85\u5904\u7406: '+str(pend)+'\u9879\n\u52a0\u5de5\u4e2d: '+str(proc)+'\u9879\n\u5df2\u5b8c\u6210: '+str(done[0])+'\u9879, '+str(done[1])+'\u4ef6'
        if ws_skus: msg += '\n\nSKU: '+', '.join(ws_skus[:15])
        return msg
    
    if sku_match:
        sku = sku_match.group(1).replace('_','-').lower()
        rate = sku_rates.get(sku)
        ws = get_workshop_efficiency(sku)
        if ws:
            ans = '\U0001f4c8 \u3010'+sku.upper()+'\u3011\u52a0\u5de5\u5355\u6570\u636e'
            ans += '\n\u7d2f\u8ba1: '+str(ws['qty'])+'\u4ef6, '+str(ws['hours'])+'\u5c0f\u65f6, '+str(ws['count'])+'\u6b21'
            if ws['rate']:
                ans += '\n\u6548\u7387: '+str(ws['rate'])+' \u5957/\u4eba/\u5c0f\u65f6 (\u52a0\u5de5\u5355\u5b9e\u6d4b)'
                if rate: ans += '\n\u9884\u8bbe: '+str(rate)+' \u5957/\u4eba/\u5c0f\u65f6'
            elif rate:
                ans += '\n\u6548\u7387: '+str(rate)+' \u5957/\u4eba/\u5c0f\u65f6 (\u9884\u8bbe\u503c)'
            return ans
        if rate:
            return '\U0001f4c8 \u3010'+sku.upper()+'\u3011\u6548\u7387\uff1a'+str(rate)+' \u5957/\u4eba/\u5c0f\u65f6'
    
    m_data = {
        'sku':'\U0001f4c8 SKU\u6548\u7387\u5168\u89c8\n\n\u5305\u88c5\u57fa\u51c6: 0.77 \u65b9/\u4eba/\u5c0f\u65f6\n\n\u5e38\u7528SKU:\nPW-MSG16-001  \u8bd5\u7ba1  21.35\nPW-MSG12-001  \u8bd5\u7ba1  16.31\nCB-CMSG-01A  \u5f69\u8272\u8bd5\u7ba1  29.20\nCB-MSG08-A01  \u5f69\u8272\u8bd5\u7ba1  13.50\nHW-PNP36-001  \u6309\u538b\u74f6  27.88\nGN-JP500-002  \u5c06\u519b\u74f6  21.59\nGN-PQ115-001  \u5c06\u519b\u74f6  21.57\nGN-JP750-001  \u5c06\u519b\u74f6  21.43\nGN-FQJ15-001  \u5c06\u519b\u74f6  15.00\nHS-MSG32-001  \u9501\u8272\u8bd5\u7ba1  15.56\nCO-MSG08-A01  \u679d\u88c5\u7ba1  10.50\nSP-CMS04-002  \u7279\u6b8a\u7ba1  9.60\nSP-MSG04-002  \u7279\u6b8a\u7ba1  9.60\n\n\u516c\u5f0f: \u6548\u7387=\u603b\u4ea7\u91cf\u00f7(\u4eba\u6570\u00d7\u5c0f\u65f6\u6570)\n\u524d\u7f00: PW=\u8bd5\u7ba1 CB=\u5f69\u8272\u8bd5\u7ba1 HW=\u6309\u538b\u74f6 SP=\u7279\u6b8a\u7ba1 GN=\u5c06\u519b\u74f6 HS=\u9501\u8272\u8bd5\u7ba1 CO=\u679d\u88c5\u7ba1',
        '\u7269\u6d41':'\U0001f4e6 \u7269\u6d41\u89c4\u5219\n\n\u6d77\u8fd0(\u6309\u91cd\u91cf): \u5b9e\u91cd\u226422kg, \u4f53\u79ef\u91cd\u226422kg\n\u6700\u957f\u8fb9\u2264120cm(\u8d85150\u5143/\u4ef6)\n\u6b21\u957f\u8fb9\u226475cm(\u8d85150\u5143/\u4ef6)\n\u56f4\u957fL+2W+2H\u2264260cm(260~330cm=180\u5143/\u4ef6)\n\u6d77\u8fd0(\u6309\u65b9): \u65e0\u91cd\u91cf/\u5c3a\u5bf8\u8981\u6c42\n\u5feb\u9012: \u5c3a\u5bf8\u226460cm, \u91cd\u91cf\u226424kg\n\u7f8e\u68ee: \u4f53\u79ef\u91cd\u226422kg(\u4e0d\u8db312kg\u630912kg\u8ba1)',
        '\u6d77\u8fd0':'\U0001f4e6 \u6d77\u8fd0\u89c4\u5219\n\n\u6309\u91cd\u91cf: \u5b9e\u91cd\u226422kg, \u4f53\u79ef\u91cd\u226422kg\n\u6700\u957f\u8fb9\u2264120cm(\u8d85150)\n\u6b21\u957f\u8fb9\u226475cm(\u8d85150)\n\u56f4\u957f\u2264260cm(260~330=180)\n\u6309\u65b9: \u65e0\u8981\u6c42',
        '\u5feb\u9012':'\U0001f4e6 \u5feb\u9012\u89c4\u5219\n\n\u5c3a\u5bf8\u226460cm\n\u91cd\u91cf\u226424kg',
        '\u7f8e\u68ee':'\U0001f4e6 \u7f8e\u68ee\n\n\u4f53\u79ef\u91cd\u226422kg(\u4e0d\u8db312kg\u630912kg\u8ba1)',
        '\u6548\u7387':'\U0001f4c8 \u6548\u7387\n\n\u5305\u88c5: 0.77 \u65b9/\u4eba/\u5c0f\u65f6\nSKU\u4ee5\u6570\u636e\u95ee\u7b54\u4e3a\u51c6\nPW-MSG16-001: 21.35 \u5957/\u4eba/\u5c0f\u65f6',
        '\u65fa\u5b63':'\U0001f4c5 \u65fa\u5b63\n\n9-10\u6708\u5907\u65fa\u5b63, 2\u7ec4\u52a0\u73ed, \u6709\u65f63\u7ec4\n\u6bcf\u7ec410+\u4eba\n\u66ae\u5047\u524d1\u4e2a\u6708\u5148\u505a\u91cf\u5927\u7684',
        '\u6625\u8282':'\U0001f4c5 \u6625\u8282\n\n\u6625\u8282\u524d1\u5468\u53d1\u5b8c\u8d27, \u63d0\u524d2\u5468\u5efa\u5355\n12\u670816\u4e0b\u5e74\u524d\u8d27\n12\u670820\u4e0b\u5e74\u540e\u8d27',
        '\u8fd0\u8425':'\U0001f4c5 \u8fd0\u8425\n\n2025: 9-10\u6708\u5907\u65fa\u5b63, 2\u7ec4\u52a0\u73ed\n11\u6708\u95f2\u4e0b\u6765, \u767d\u592917\u4eba\n\u6625\u8282\u524d1\u5468\u53d1\u5b8c, \u63d0\u524d2\u5468\u5efa\u5355\n12\u670816\u4e0b\u5e74\u524d, 12\u670820\u4e0b\u5e74\u540e',
        '\u4e0b\u5355':'\U0001f4c5 \u4e0b\u5355\n\n12\u670816\u4e0b\u5e74\u524d\u8d27\n12\u670820\u4e0b\u5e74\u540e\u8d27\n\u6625\u8282\u524d2\u5468\u5efa\u5355',
        '\u5de5\u4eba':'\U0001f464 \u4eba\u5458\n\n\u65fa\u5b63(9-10\u6708): 2\u7ec4\u52a0\u73ed, \u6bcf\u7ec410+\u4eba\n\u6de1\u5b63(11\u6708): \u767d\u592917\u4eba, \u665a\u4e0a\u4e0d\u52a0\u73ed',
    }
    for kw, ans in m_data.items():
        if kw in q: return ans
    return '\U0001f4a1 \u8bd5\u8bd5\u95ee:\n- \u7269\u6d41\u89c4\u5219 / \u6d77\u8fd0 / \u5feb\u9012\n- \u6548\u7387 / SKU / \u5168\u90e8SKU\n- \u65fa\u5b63 / \u8fd0\u8425 / \u6625\u8282\n- \u5de5\u4eba / \u4e0b\u5355\n- \u8f66\u95f4 / \u52a0\u5de5 / \u751f\u4ea7\n- \u8ba1\u7b97: PW-MSG16-001,10\u4eba1000\u5957\u505a\u591a\u4e45'

if __name__ == '__main__':
    url = 'http://localhost:'+str(PORT)
    print('='*45)
    print('  \u5de5\u5382\u5de5\u4f5c\u53f0 v3.1 (\u542b\u5206\u533a\u57df\u626b\u7801)')
    print('  '+url)
    print('  \u6309 Ctrl+C \u505c\u6b62\u670d\u52a1\u5668')
    print('='*45)
    # 关闭旧数据库重新创建新schema
    # (new schema with region columns auto-creates)
    threading.Timer(0.5, lambda: os.startfile(url)).start()
    print('\u6d4f\u89c8\u5668\u672a\u81ea\u52a8\u6253\u5f00\uff1f '+url)
    print('\u5de5\u4eba\u626b\u7801\uff1ahttp://'+get_ip()+':'+str(PORT)+'/scan')
    print('\u7ba1\u7406\u540e\u53f0\uff1ahttp://'+get_ip()+':'+str(PORT)+'/scan_admin')
    server = http.server.ThreadingHTTPServer(('0.0.0.0', PORT), H)
    print('服务器已启动，等待连接...')
    try: server.serve_forever()
    except KeyboardInterrupt: server.shutdown()
    except Exception as e:
        print('服务器异常:', e)
        import traceback
        traceback.print_exc()
