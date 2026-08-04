"""
美国发货汇总生成功能
======================
功能: 上传发货单Excel，生成按区域分组的美国发货汇总表
"""

import os, openpyxl
from collections import OrderedDict, defaultdict
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


REGIONS = ['东北部', '东南部', '西部', '中北部', '中南部']


def run_us(fp):
    """
    生成美国发货汇总Excel
    
    参数:
        fp: Excel文件路径
        
    返回:
        str: 结果消息
        
    Excel格式:
        Sheet1: 发货单信息 (第1行=表头)
            A=货件单号 B=物流中心编码 C=物流商 
            D=物流渠道 E=物流商单号 G=国家 H=货件单号
        
        Sheet5: 箱数/重量/体积 (第1行=表头)
            A=货件单号 B=箱数 C=重量(kg) D=体积(m3) E=类型
    """
    wb = openpyxl.load_workbook(fp)
    ws1 = wb[wb.sheetnames[0]]
    try:
        ws5 = wb[wb.sheetnames[4]]
    except:
        ws5 = wb.active
    
    # 读取发货单信息
    d = OrderedDict()
    for r in range(2, ws1.max_row + 1):
        doc = str(ws1.cell(r, 1).value or '').strip()
        if not doc or doc in d:
            continue
        d[doc] = [str(ws1.cell(r, c).value or '').strip() for c in [2, 3, 4, 5, 7, 8]]
    
    # 读取箱数/重量/体积
    pk = {}
    for r in range(2, ws5.max_row + 1):
        doc = str(ws5.cell(r, 1).value or '').strip()
        if not doc:
            continue
        k = doc + '|' + str(ws5.cell(r, 5).value or '')
        if k not in pk:
            pk[k] = {'b': 0, 'w': 0.0, 'v': 0.0}
        try:
            pk[k]['b'] += int(ws5.cell(r, 2).value or 0)
        except:
            pass
        try:
            pk[k]['w'] += float(str(ws5.cell(r, 3).value or '0').replace(',', '').strip())
        except:
            pass
        try:
            pk[k]['v'] += float(str(ws5.cell(r, 4).value or '0').replace(',', '').strip())
        except:
            pass
    
    # 创建输出Excel
    out = openpyxl.Workbook()
    ows = out.active
    ows.title = '发货汇总'
    
    # 表头
    headers = ['发货单号', '物流中心编码', '物流商', '物流渠道', '物流商单号', 
               '国家', '总箱数', '总重量(kg)', '总体积(m3)', '货件单号']
    tn = Side(style='thin', color='000000')
    bd = Border(left=tn, right=tn, top=tn, bottom=tn)
    hf = Font(bold=True, size=11, color='FFFFFF')
    hf2 = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    nf = Font(size=11)
    bf = Font(bold=True, size=12)
    sfl = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    
    for i, hh in enumerate(headers, 1):
        c = ows.cell(1, i, hh)
        c.font = hf
        c.fill = hf2
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = bd
    
    # 按区域分组
    rd = defaultdict(list)
    for doc, v in d.items():
        ch = v[2]
        rg = '其他'
        for r2 in REGIONS:
            if ch.startswith(r2):
                rg = r2
                break
        b = 0
        w = 0.0
        vol = 0.0
        for k, pv in pk.items():
            if k.startswith(doc):
                b += pv['b']
                w += pv['w']
                vol += pv['v']
        rd[rg].append((doc, v[0], v[1], ch, v[3], v[4], b, round(w, 2), round(vol, 2), v[5]))
    
    # 写入数据
    row = 2
    gb = 0
    gw = 0.0
    gv = 0.0
    for rg in REGIONS:
        items = rd.get(rg, [])
        if not items:
            continue
        for it in items:
            for i, val in enumerate(it, 1):
                c = ows.cell(row, i, val)
                c.font = nf
                c.border = bd
                c.alignment = Alignment(vertical='center')
            row += 1
        
        sb = sum(i[6] for i in items)
        sw = sum(i[7] for i in items)
        sv = sum(i[8] for i in items)
        
        ows.cell(row, 1, rg + ' 小计').font = Font(size=11)
        ows.cell(row, 7, sb).font = Font(size=11)
        ows.cell(row, 8, round(sw, 2)).font = Font(size=11)
        ows.cell(row, 9, round(sv, 2)).font = Font(size=11)
        for i in range(1, 11):
            ows.cell(row, i).border = bd
            ows.cell(row, i).fill = sfl
            ows.cell(row, i).alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        gb += sb
        gw += sw
        gv += sv
    
    # 合计行
    ows.cell(row, 1, '合计').font = bf
    ows.cell(row, 7, gb).font = bf
    ows.cell(row, 8, round(gw, 2)).font = bf
    ows.cell(row, 9, round(gv, 2)).font = bf
    for i in range(1, 11):
        ows.cell(row, i).border = bd
        ows.cell(row, i).alignment = Alignment(horizontal='center', vertical='center')
    
    # 列宽
    for i, w in enumerate([18, 12, 10, 22, 22, 8, 10, 14, 14, 20], 1):
        ows.column_dimensions[get_column_letter(i)].width = w
    
    # 保存
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    fout = os.path.join(output_dir, os.path.basename(fp).replace('.xlsx', '-汇总.xlsx'))
    out.save(fout)
    
    region_summary = '\n'.join([rg + ': ' + str(len(rd[rg])) + '单' for rg in REGIONS if rd[rg]])
    return ('✅ 美国发货汇总\n合计：' + str(len(d)) + '单，' + str(gb) + '箱，'
            + str(round(gw, 2)) + 'kg，' + str(round(gv, 2)) + 'm3\n'
            + region_summary + '\n文件：' + os.path.basename(fout))


# ====== 独立测试 ======
if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        result = run_us(sys.argv[1])
        print(result)
    else:
        print("用法: python 功能_美国发货汇总.py <Excel文件路径>")
