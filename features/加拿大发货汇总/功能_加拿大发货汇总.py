"""
加拿大发货汇总生成功能
======================
功能: 上传发货单Excel，生成加拿大发货汇总表（直接清单）
"""

import os, openpyxl
from collections import OrderedDict
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def run_ca(fp):
    """
    生成加拿大发货汇总Excel
    
    参数:
        fp: Excel文件路径
        
    返回:
        str: 结果消息
        
    Excel格式:
        Sheet1: 发货单信息
            A=货件单号 B=物流中心编码 C=物流商 G=国家 H=货件单号
        
        Sheet5: 箱数/重量/体积
            A=货件单号 B=箱数 C=重量(kg) D=体积(m3) E=类型
    """
    wb = openpyxl.load_workbook(fp)
    ws1 = wb[wb.sheetnames[0]]
    try:
        ws5 = wb[wb.sheetnames[4]]
    except:
        ws5 = wb.active
    
    d = OrderedDict()
    for r in range(2, ws1.max_row + 1):
        doc = str(ws1.cell(r, 1).value or '').strip()
        if not doc or doc in d:
            continue
        d[doc] = [str(ws1.cell(r, c).value or '').strip() for c in [2, 3, 7, 8]]
    
    pk = {}
    for r in range(2, ws5.max_row + 1):
        doc = str(ws5.cell(r, 1).value or '').strip()
        if not doc:
            continue
        k = doc + '|' + str(ws5.cell(r, 5).value or '')
        if k not in pk:
            pk[k] = {'b': 0, 'w': 0.0, 'v': 0.0}
        try:
            pk[k]['b'] += int(ws5.cell(r, 2).value or 0)
        except:
            pass
        try:
            pk[k]['w'] += float(str(ws5.cell(r, 3).value or '0').replace(',', '').strip())
        except:
            pass
        try:
            pk[k]['v'] += float(str(ws5.cell(r, 4).value or '0').replace(',', '').strip())
        except:
            pass
    
    out = openpyxl.Workbook()
    ows = out.active
    ows.title = '加拿大发货汇总'
    
    headers = ['发货单号', '物流中心编码', '物流商', '国家', '总箱数', '总重量(kg)', '总体积(m3)', '货件单号']
    tn = Side(style='thin', color='000000')
    bd = Border(left=tn, right=tn, top=tn, bottom=tn)
    hf = Font(bold=True, size=11, color='FFFFFF')
    hf2 = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    nf = Font(size=11)
    bf = Font(bold=True, size=12)
    
    for i, hh in enumerate(headers, 1):
        c = ows.cell(1, i, hh)
        c.font = hf
        c.fill = hf2
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = bd
    
    row = 2
    gb = 0
    gw = 0.0
    gv = 0.0
    for doc, v in d.items():
        b = 0
        w = 0.0
        vol = 0.0
        for k, pv in pk.items():
            if k.startswith(doc):
                b += pv['b']
                w += pv['w']
                vol += pv['v']
        
        vs = [doc, v[0], v[1], v[2], b, round(w, 2), round(vol, 2), v[3]]
        for i, val in enumerate(vs, 1):
            c = ows.cell(row, i, val)
            c.font = nf
            c.border = bd
            c.alignment = Alignment(vertical='center')
        row += 1
        gb += b
        gw += w
        gv += vol
    
    ows.cell(row, 1, '合计').font = bf
    ows.cell(row, 5, gb).font = bf
    ows.cell(row, 6, round(gw, 2)).font = bf
    ows.cell(row, 7, round(gv, 2)).font = bf
    for i in range(1, 9):
        ows.cell(row, i).border = bd
        ows.cell(row, i).alignment = Alignment(horizontal='center', vertical='center')
    
    for i, w in enumerate([18, 14, 14, 10, 10, 14, 14, 20], 1):
        ows.column_dimensions[get_column_letter(i)].width = w
    
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    fout = os.path.join(output_dir, os.path.basename(fp).replace('.xlsx', '-加拿大汇总.xlsx'))
    out.save(fout)
    
    return ('✅ 加拿大发货汇总\n合计：' + str(len(d)) + '单，' + str(gb) + '箱，'
            + str(round(gw, 2)) + 'kg，' + str(round(gv, 2)) + 'm3\n文件：' + os.path.basename(fout))


# ====== 独立测试 ======
if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        result = run_ca(sys.argv[1])
        print(result)
    else:
        print("用法: python 功能_加拿大发货汇总.py <Excel文件路径>")
