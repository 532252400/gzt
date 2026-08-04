import openpyxl, sys, os
sys.path.insert(0, "D:/桌面")

# Create test file
wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Sheet1"
ws1.cell(1,1,"货件单号")
ws1.cell(1,2,"物流中心编码")
ws1.cell(1,3,"物流商")
ws1.cell(1,4,"物流渠道")
ws1.cell(1,5,"物流商单号")
ws1.cell(1,7,"国家")
ws1.cell(1,8,"货件单号")
ws1.cell(2,1,"TEST001")
ws1.cell(2,2,"东北部中心1")
ws1.cell(2,3,"FedEx")
ws1.cell(2,4,"Ground")
ws1.cell(2,5,"FDX123")
ws1.cell(2,7,"US")
ws1.cell(2,8,"SHIP001")

ws5 = wb.create_sheet("Sheet5")
ws5.cell(1,1,"货件单号")
ws5.cell(1,2,"箱数")
ws5.cell(1,3,"重量")
ws5.cell(1,4,"体积")
ws5.cell(1,5,"类型")
ws5.cell(2,1,"TEST001")
ws5.cell(2,2,10)
ws5.cell(2,3,100.5)
ws5.cell(2,4,2.5)
ws5.cell(2,5,"A")

test_path = "C:/Users/pc/ZCodeProject/test_us.xlsx"
wb.save(test_path)
print(f"Test file created: {test_path}")

# Import and test run_us
import importlib.util
spec = importlib.util.spec_from_file_location("factory", "D:/桌面/工厂工作台.py")
mod = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)

try:
    result = mod.run_us(test_path)
    print(f"SUCCESS: {result}")
except Exception as e:
    import traceback
    print(f"ERROR: {e}")
    traceback.print_exc()
