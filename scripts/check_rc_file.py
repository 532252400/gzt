import openpyxl
fp = r'D:\桌面\7-8Receiving-Record-20260708-933381695180312576.xlsx'
wb = openpyxl.load_workbook(fp)
print('Sheet names:', wb.sheetnames)
ws = wb.active
print('Active sheet:', ws.title)
print('First 5 rows:')
for r in range(1, min(6, ws.max_row+1)):
    row_data = []
    for c in range(1, min(10, ws.max_column+1)):
        val = ws.cell(r, c).value
        row_data.append(str(val)[:30] if val else '')
    print(f'  Row {r}: {row_data}')
