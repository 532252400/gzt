import openpyxl
fp = r'D:\桌面\7-8Receiving-Record-20260708-933381695180312576.xlsx'
wb = openpyxl.load_workbook(fp, data_only=True)
ws = wb['sheet1']
print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
print('Row 3 data:')
for c in range(1, 10):
    val = ws.cell(3, c).value
    print(f'  Col {c}: {repr(val)[:50]}')
