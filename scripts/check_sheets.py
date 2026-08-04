import openpyxl
fp = r'D:\桌面\7-8Receiving-Record-20260708-933381695180312576.xlsx'
wb = openpyxl.load_workbook(fp)
print('Sheet names:', wb.sheetnames)
for name in wb.sheetnames:
    print(f'  "{name}" - exact bytes: {name.encode("utf-8")}')
print(f"'Sheet2' in names: {'Sheet2' in wb.sheetnames}")
print(f"'sheet1' in names: {'sheet1' in wb.sheetnames}")
print(f"Lower: {[n.lower() for n in wb.sheetnames]}")
